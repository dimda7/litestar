import asyncio
import json
import logging
import tempfile
import time
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path

import openpyxl
from dateutil import parser as date_parser
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession
from litestar import Controller, get, post
from litestar.connection.request import Request
from litestar.datastructures import UploadFile
from litestar.enums import RequestEncodingType
from litestar.params import Body
from litestar.response import Template, Response, Redirect

from db_manager import get_session_maker
from models import Actives, CounterType, Ptoir, PtoirLevelWarning
from schemas import SelectSheetRequest
from sql_utils import sql_escape
from parser_storage import (
    LOG_DIR,
    load_data as _load_data,
    save_data as _save_data,
    cleanup_old_files as _cleanup_old_files,
)

logger = logging.getLogger("ptoir_parser")

PREFIX = "ptoir_parser"

# Excel-даты вводятся по московскому времени, в БД date_activation/zero_point_value
# хранятся в UTC (см. старую функцию update_ptoir) — поэтому везде сдвиг на 3 часа.
MSK_OFFSET = timedelta(hours=3)

# Каждая строка требует нескольких запросов к БД, на файлах в тысячи строк
# операция идёт секундами-минутами — прогресс отдаётся через отдельный опрос,
# чтобы не держать один HTTP-запрос открытым всё это время.
PROGRESS_TTL_SECONDS = 15 * 60
_progress: dict[str, dict] = {}
# asyncio хранит только слабую ссылку на fire-and-forget задачи — без явного
# хранения задача может быть собрана GC до завершения.
_tasks: dict[str, asyncio.Task] = {}


def _cleanup_progress() -> None:
    cutoff = time.time() - PROGRESS_TTL_SECONDS
    stale = [tid for tid, state in _progress.items() if state["created_at"] < cutoff]
    for tid in stale:
        _progress.pop(tid, None)


class PtoirParserController(Controller):
    path = "/ptoir-parser"

    @get("/")
    async def index(
        self,
        request: Request,
        page: int = 1,
        per_page: int = 10,
        select_sheet: bool = False,
    ) -> Template:
        page = max(page, 1)
        per_page = min(per_page, 200)
        error: str = request.session.pop(f"{PREFIX}_error", "")

        pending_sheets: list[str] = []
        pending_filename: str = ""
        if select_sheet:
            pending_sheets = request.session.get(f"{PREFIX}_pending_sheets", [])
            pending_filename = request.session.get(f"{PREFIX}_pending_filename", "")

        session_id = request.session.get(f"{PREFIX}_session_id", "")
        stored = _load_data(session_id) if session_id else None

        all_rows: list[dict] = stored["rows"] if stored else []
        headers: list[str] = stored["headers"] if stored else []
        filename: str = stored["filename"] if stored else ""

        total = len(all_rows)
        total_pages = max((total + per_page - 1) // per_page, 1)
        if page > total_pages:
            page = total_pages

        offset = (page - 1) * per_page
        rows = all_rows[offset:offset + per_page]

        return Template(
            template_name="ptoir_parser.html",
            context={
                "headers": headers,
                "rows": rows,
                "all_rows": all_rows,
                "filename": filename,
                "error": error,
                "page": page,
                "per_page": per_page,
                "total": total,
                "total_pages": total_pages,
                "user_id": request.session.get("user_id"),
                "fullname": request.session.get("fullname", ""),
                "active_page": "ptoir_parser",
                "pending_sheets": pending_sheets,
                "pending_filename": pending_filename,
            },
        )

    @post("/upload")
    async def upload(self, request: Request) -> Redirect:
        form = await request.form()
        upload_file: UploadFile | None = form.get("file")

        if not upload_file or not upload_file.filename:
            request.session[f"{PREFIX}_error"] = "Файл не выбран"
            return Redirect("/ptoir-parser")

        suffix = Path(upload_file.filename).suffix.lower()
        if suffix not in (".xlsx", ".xls"):
            request.session[f"{PREFIX}_error"] = "Поддерживаются только .xlsx и .xls файлы"
            return Redirect("/ptoir-parser")

        try:
            _cleanup_old_files()

            content = await upload_file.read()
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                tmp.write(content)
                tmp_path = tmp.name

            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()

            if len(sheet_names) > 1:
                request.session[f"{PREFIX}_pending_file"] = tmp_path
                request.session[f"{PREFIX}_pending_sheets"] = sheet_names
                request.session[f"{PREFIX}_pending_filename"] = upload_file.filename
                return Redirect("/ptoir-parser?select_sheet=1")

            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            ws = wb.active

            rows: list[dict[str, str | None]] = []
            headers: list[str] = []

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c) if c else f"col_{i}" for i, c in enumerate(row)]
                    continue
                if all(c is None or str(c).strip() == "" for c in row):
                    continue
                rows.append({headers[j]: row[j] for j in range(len(row))})

            wb.close()
            Path(tmp_path).unlink(missing_ok=True)

            session_id = uuid.uuid4().hex
            _save_data(session_id, {
                "headers": headers,
                "rows": rows,
                "filename": upload_file.filename,
            })
            request.session[f"{PREFIX}_session_id"] = session_id

            return Redirect("/ptoir-parser")
        except Exception as e:
            request.session[f"{PREFIX}_error"] = f"Ошибка чтения файла: {e}"
            return Redirect("/ptoir-parser")

    @post("/select-sheet")
    async def select_sheet(
        self,
        request: Request,
        data: SelectSheetRequest = Body(media_type=RequestEncodingType.URL_ENCODED),
    ) -> Redirect:
        sheet_name = data.sheet_name

        tmp_path = request.session.get(f"{PREFIX}_pending_file", "")
        filename = request.session.get(f"{PREFIX}_pending_filename", "")
        sheet_names = request.session.get(f"{PREFIX}_pending_sheets", [])

        if not tmp_path or not Path(tmp_path).exists():
            request.session[f"{PREFIX}_error"] = "Временный файл истёк. Загрузите файл заново."
            return Redirect("/ptoir-parser")

        if sheet_name not in sheet_names:
            request.session[f"{PREFIX}_error"] = "Выбранный лист не найден в файле."
            return Redirect("/ptoir-parser")

        try:
            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            ws = wb[sheet_name]

            rows: list[dict[str, str | None]] = []
            headers: list[str] = []

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c) if c else f"col_{i}" for i, c in enumerate(row)]
                    continue
                if all(c is None or str(c).strip() == "" for c in row):
                    continue
                rows.append({headers[j]: row[j] for j in range(len(row))})

            wb.close()
            Path(tmp_path).unlink(missing_ok=True)

            request.session.pop(f"{PREFIX}_pending_file", None)
            request.session.pop(f"{PREFIX}_pending_sheets", None)
            request.session.pop(f"{PREFIX}_pending_filename", None)

            session_id = uuid.uuid4().hex
            _save_data(session_id, {
                "headers": headers,
                "rows": rows,
                "filename": f"{filename} [{sheet_name}]",
            })
            request.session[f"{PREFIX}_session_id"] = session_id

            return Redirect("/ptoir-parser")
        except Exception as e:
            request.session[f"{PREFIX}_error"] = f"Ошибка чтения листа: {e}"
            return Redirect("/ptoir-parser")

    async def _validate_and_build_rows(
        self, db_session: AsyncSession, rows: list[dict],
        progress: dict | None = None,
    ) -> tuple[list[dict], list[tuple[int, datetime, int, int, int]]]:
        """Валидирует строки Excel для запуска ПТОиР.

        Возвращает (errors, valid_rows), где valid_rows — список кортежей
        (id_ptoir, date_activation, interval, id_level_warning, zero_point_value).
        Если передан progress-словарь, каждые несколько строк в него пишется
        (processed, total, phase="validating") для опроса с фронтенда.
        """
        errors: list[dict] = []
        valid_rows: list[tuple[int, datetime, int, int, int]] = []

        if progress is not None:
            progress.update(processed=0, total=len(rows), phase="validating")

        for idx, row in enumerate(rows):
            row_num = idx + 1
            if progress is not None and (idx % 20 == 0 or row_num == len(rows)):
                progress["processed"] = row_num
            number_ptoir = str(row.get("ПТОиР", "") or "").strip()
            active_number = str(row.get("Актив", "") or "").strip()
            type_counter = str(row.get("Тип счетчика", "") or "").strip()
            interval_raw = row.get("Интервал")
            date_raw = row.get("Дата активации")
            service_raw = row.get("Данные последнего обслуживания")

            if not number_ptoir:
                errors.append({"row": row_num, "field": "ПТОиР", "message": "Поле 'ПТОиР' пустое"})
                continue

            result = await db_session.execute(
                select(Ptoir.id, Ptoir.id_active).where(Ptoir.number_ptoir == number_ptoir)
            )
            ptoir_row = result.first()
            if ptoir_row is None:
                errors.append({"row": row_num, "field": "ПТОиР",
                                "message": f"ПТОиР не найден: '{number_ptoir}'"})
                continue
            ptoir_id, ptoir_active_id = ptoir_row

            if active_number:
                result = await db_session.execute(
                    select(Actives.id).where(Actives.active_number == active_number)
                )
                active_id = result.scalar_one_or_none()
                if active_id is None:
                    errors.append({"row": row_num, "field": "Актив",
                                    "message": f"Актив не найден: '{active_number}'"})
                    continue
                if ptoir_active_id is not None and active_id != ptoir_active_id:
                    errors.append({"row": row_num, "field": "Актив",
                                    "message": (f"Актив '{active_number}' не соответствует "
                                                f"активу ПТОиР '{number_ptoir}'")})
                    continue

            if not type_counter:
                errors.append({"row": row_num, "field": "Тип счетчика", "message": "Поле 'Тип счетчика' пустое"})
                continue

            result = await db_session.execute(
                select(CounterType.id).where(CounterType.type == type_counter)
            )
            counter_type_id = result.scalar_one_or_none()
            if counter_type_id is None:
                errors.append({"row": row_num, "field": "Тип счетчика",
                                "message": f"Тип счетчика не найден: '{type_counter}'"})
                continue

            result = await db_session.execute(
                select(PtoirLevelWarning.id).where(
                    (PtoirLevelWarning.id_ptoir == ptoir_id)
                    & (PtoirLevelWarning.id_counter_type == counter_type_id)
                )
            )
            level_warning_id = result.scalar_one_or_none()
            if level_warning_id is None:
                errors.append({"row": row_num, "field": "*",
                                "message": (f"Уровень предупреждения не найден для ПТОиР "
                                            f"'{number_ptoir}' и типа счетчика '{type_counter}'")})
                continue

            try:
                if isinstance(date_raw, datetime):
                    date_activation = date_raw - MSK_OFFSET
                else:
                    date_activation = date_parser.parse(str(date_raw), dayfirst=True) - MSK_OFFSET
            except Exception:
                errors.append({"row": row_num, "field": "Дата активации",
                                "message": f"Некорректная дата активации: '{date_raw}'"})
                continue

            try:
                interval = int(interval_raw)
            except (TypeError, ValueError):
                errors.append({"row": row_num, "field": "Интервал",
                                "message": f"Некорректный интервал: '{interval_raw}'"})
                continue

            if isinstance(service_raw, (int, float)) and not isinstance(service_raw, bool):
                zero_point_value = int(service_raw)
            else:
                try:
                    service_dt = date_parser.parse(str(service_raw), dayfirst=True) - MSK_OFFSET
                    service_dt = service_dt.replace(tzinfo=timezone.utc)
                    zero_point_value = int(service_dt.timestamp())
                except Exception:
                    errors.append({"row": row_num, "field": "Данные последнего обслуживания",
                                    "message": f"Некорректное значение обслуживания: '{service_raw}'"})
                    continue

            valid_rows.append((ptoir_id, date_activation, interval, level_warning_id, zero_point_value))

        return errors, valid_rows

    @post("/generate-sql/start")
    async def generate_sql_start(
        self,
        request: Request,
        data: dict = Body(media_type=RequestEncodingType.MULTI_PART),
    ) -> Response:
        rows: list[dict] = json.loads(data.get("rows", "[]"))
        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_generate(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task
        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_generate(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                errors, valid_rows = await self._validate_and_build_rows(session, rows, progress=progress)
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка валидации: {e}"}])
            return

        if errors:
            progress.update(status="error", errors=errors)
            return

        sql_lines = self._build_sql_lines(valid_rows)
        progress.update(status="done", sql="\n".join(sql_lines), count=len(valid_rows))

    @post("/execute-sql/start")
    async def execute_sql_start(
        self,
        request: Request,
        data: dict = Body(media_type=RequestEncodingType.MULTI_PART),
    ) -> Response:
        rows: list[dict] = json.loads(data.get("rows", "[]"))
        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_execute(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task
        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_execute(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                try:
                    errors, valid_rows = await self._validate_and_build_rows(session, rows, progress=progress)
                except Exception as e:
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка валидации: {e}"}])
                    return

                if errors:
                    progress.update(status="error", errors=errors)
                    return

                if not valid_rows:
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": "Нет валидных строк для обновления"}])
                    return

                progress.update(processed=0, total=len(valid_rows), phase="executing")
                try:
                    for i, (ptoir_id, date_activation, interval, level_warning_id, zero_point_value) in enumerate(valid_rows, start=1):
                        await session.execute(
                            text(
                                "UPDATE public.ptoir SET date_activation = :da, interval = :iv, is_active = TRUE "
                                "WHERE id = :id"
                            ),
                            {"da": date_activation, "iv": interval, "id": ptoir_id},
                        )
                        await session.execute(
                            text("UPDATE public.ptoir_level_warning SET zero_point_value = :zp WHERE id = :id"),
                            {"zp": zero_point_value, "id": level_warning_id},
                        )
                        if i % 20 == 0 or i == len(valid_rows):
                            progress["processed"] = i
                    await session.commit()
                except Exception as e:
                    await session.rollback()
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
                    return
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
            return

        now = datetime.now()
        log_lines = [
            f"=== Execute PTOиR update: {now.strftime('%Y-%m-%d %H:%M:%S')} ===",
            f"Rows updated: {len(valid_rows)}",
            "",
            *self._build_sql_lines(valid_rows),
            "",
        ]

        log_file = LOG_DIR / f"update_ptoir_{now.strftime('%Y-%m-%d_%H-%M-%S')}.log"
        with open(log_file, "a", encoding="utf-8") as f:
            f.write("\n".join(log_lines))
        logger.info("Updated ptoir for %d rows, log: %s", len(valid_rows), log_file)

        progress.update(status="done", count=len(valid_rows),
                         message=f"Успешно обновлено {len(valid_rows)} ПТОиР")

    @get("/progress/{task_id:str}")
    async def get_progress(self, task_id: str) -> Response:
        state = _progress.get(task_id)
        if state is None:
            return Response(
                content=json.dumps({"status": "error", "errors": [{"row": 0, "field": "*", "message": "Задача не найдена или устарела"}]}),
                status_code=200,
                media_type="application/json",
            )
        return Response(
            content=json.dumps({k: v for k, v in state.items() if k != "created_at"}),
            status_code=200,
            media_type="application/json",
        )

    @staticmethod
    def _build_sql_lines(valid_rows: list[tuple[int, datetime, int, int, int]]) -> list[str]:
        sql_lines: list[str] = []
        for ptoir_id, date_activation, interval, level_warning_id, zero_point_value in valid_rows:
            date_str = date_activation.strftime("%Y-%m-%d %H:%M:%S")
            sql_lines.append(
                f"UPDATE public.ptoir SET date_activation = '{sql_escape(date_str)}', "
                f"interval = {interval}, is_active = TRUE WHERE id = {ptoir_id};"
            )
            sql_lines.append(
                f"UPDATE public.ptoir_level_warning SET zero_point_value = {zero_point_value} "
                f"WHERE id = {level_warning_id};"
            )
        return sql_lines
