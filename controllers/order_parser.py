import asyncio
import json
import logging
import tempfile
import time
import uuid
from datetime import datetime
from pathlib import Path

import openpyxl
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession
from litestar import Controller, get, post
from litestar.connection.request import Request
from litestar.datastructures import UploadFile
from litestar.enums import RequestEncodingType
from litestar.params import Body
from litestar.response import Template, Response, Redirect

from db_manager import get_session_maker
from models import Orders
from schemas import SelectSheetRequest
from sql_utils import sql_escape
from parser_storage import (
    LOG_DIR,
    load_data as _load_data,
    save_data as _save_data,
    cleanup_old_files as _cleanup_old_files,
)

logger = logging.getLogger("order_parser")

PREFIX = "order_parser"

# Every row needs database queries, and on files of hundreds to thousands of
# rows the operation takes seconds — progress is served through a separate poll
# (see ptoir_parser.py).
PROGRESS_TTL_SECONDS = 15 * 60
_progress: dict[str, dict] = {}
_tasks: dict[str, asyncio.Task] = {}


def _cleanup_progress() -> None:
    cutoff = time.time() - PROGRESS_TTL_SECONDS
    stale = [tid for tid, state in _progress.items() if state["created_at"] < cutoff]
    for tid in stale:
        _progress.pop(tid, None)


class OrderParserController(Controller):
    path = "/order-parser"

    @get("/")
    async def index(
        self,
        request: Request,
        page: int = 1,
        per_page: int = 10,
        select_sheet: bool = False,
    ) -> Template:
        page = max(page, 1)
        per_page = min(per_page, 200)
        error: str = request.session.pop(f"{PREFIX}_error", "")

        pending_sheets: list[str] = []
        pending_filename: str = ""
        if select_sheet:
            pending_sheets = request.session.get(f"{PREFIX}_pending_sheets", [])
            pending_filename = request.session.get(f"{PREFIX}_pending_filename", "")

        session_id = request.session.get(f"{PREFIX}_session_id", "")
        stored = _load_data(session_id) if session_id else None

        all_rows: list[dict] = stored["rows"] if stored else []
        headers: list[str] = stored["headers"] if stored else []
        filename: str = stored["filename"] if stored else ""

        total = len(all_rows)
        total_pages = max((total + per_page - 1) // per_page, 1)
        if page > total_pages:
            page = total_pages

        offset = (page - 1) * per_page
        rows = all_rows[offset:offset + per_page]

        return Template(
            template_name="order_parser.html",
            context={
                "headers": headers,
                "rows": rows,
                "all_rows": all_rows,
                "filename": filename,
                "error": error,
                "page": page,
                "per_page": per_page,
                "total": total,
                "total_pages": total_pages,
                "user_id": request.session.get("user_id"),
                "fullname": request.session.get("fullname", ""),
                "active_page": "order_parser",
                "pending_sheets": pending_sheets,
                "pending_filename": pending_filename,
            },
        )

    @post("/upload")
    async def upload(self, request: Request) -> Redirect:
        form = await request.form()
        upload_file: UploadFile | None = form.get("file")

        if not upload_file or not upload_file.filename:
            request.session[f"{PREFIX}_error"] = "Файл не выбран"
            return Redirect("/order-parser")

        suffix = Path(upload_file.filename).suffix.lower()
        if suffix not in (".xlsx", ".xls"):
            request.session[f"{PREFIX}_error"] = "Поддерживаются только .xlsx и .xls файлы"
            return Redirect("/order-parser")

        try:
            _cleanup_old_files()

            content = await upload_file.read()
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                tmp.write(content)
                tmp_path = tmp.name

            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()

            if len(sheet_names) > 1:
                request.session[f"{PREFIX}_pending_file"] = tmp_path
                request.session[f"{PREFIX}_pending_sheets"] = sheet_names
                request.session[f"{PREFIX}_pending_filename"] = upload_file.filename
                return Redirect("/order-parser?select_sheet=1")

            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            ws = wb.active

            rows: list[dict[str, str | None]] = []
            headers: list[str] = []

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c) if c else f"col_{i}" for i, c in enumerate(row)]
                    continue
                if all(c is None or str(c).strip() == "" for c in row):
                    continue
                rows.append({headers[j]: row[j] for j in range(len(row))})

            wb.close()
            Path(tmp_path).unlink(missing_ok=True)

            session_id = uuid.uuid4().hex
            _save_data(session_id, {
                "headers": headers,
                "rows": rows,
                "filename": upload_file.filename,
            })
            request.session[f"{PREFIX}_session_id"] = session_id

            return Redirect("/order-parser")
        except Exception as e:
            request.session[f"{PREFIX}_error"] = f"Ошибка чтения файла: {e}"
            return Redirect("/order-parser")

    @post("/select-sheet")
    async def select_sheet(
        self,
        request: Request,
        data: SelectSheetRequest = Body(media_type=RequestEncodingType.URL_ENCODED),
    ) -> Redirect:
        sheet_name = data.sheet_name

        tmp_path = request.session.get(f"{PREFIX}_pending_file", "")
        filename = request.session.get(f"{PREFIX}_pending_filename", "")
        sheet_names = request.session.get(f"{PREFIX}_pending_sheets", [])

        if not tmp_path or not Path(tmp_path).exists():
            request.session[f"{PREFIX}_error"] = "Временный файл истёк. Загрузите файл заново."
            return Redirect("/order-parser")

        if sheet_name not in sheet_names:
            request.session[f"{PREFIX}_error"] = "Выбранный лист не найден в файле."
            return Redirect("/order-parser")

        try:
            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            ws = wb[sheet_name]

            rows: list[dict[str, str | None]] = []
            headers: list[str] = []

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c) if c else f"col_{i}" for i, c in enumerate(row)]
                    continue
                if all(c is None or str(c).strip() == "" for c in row):
                    continue
                rows.append({headers[j]: row[j] for j in range(len(row))})

            wb.close()
            Path(tmp_path).unlink(missing_ok=True)

            request.session.pop(f"{PREFIX}_pending_file", None)
            request.session.pop(f"{PREFIX}_pending_sheets", None)
            request.session.pop(f"{PREFIX}_pending_filename", None)

            session_id = uuid.uuid4().hex
            _save_data(session_id, {
                "headers": headers,
                "rows": rows,
                "filename": f"{filename} [{sheet_name}]",
            })
            request.session[f"{PREFIX}_session_id"] = session_id

            return Redirect("/order-parser")
        except Exception as e:
            request.session[f"{PREFIX}_error"] = f"Ошибка чтения листа: {e}"
            return Redirect("/order-parser")

    async def _validate_and_build_rows(
        self, db_session: AsyncSession, rows: list[dict],
        progress: dict | None = None,
    ) -> tuple[list[dict], list[tuple[int, int, str, str]]]:
        """Validate the Excel rows for assigning a parent work order.

        Returns (errors, valid_rows), where valid_rows is a list of tuples
        (child_id, parent_id, child_number, parent_number).
        When a progress dict is passed, (processed, total, phase="validating")
        is written into it every few rows for the frontend to poll.
        """
        errors: list[dict] = []
        valid_rows: list[tuple[int, int, str, str]] = []
        batch_children: set[str] = set()

        if rows and "Заказ-наряд" not in rows[0]:
            errors.append({"row": 0, "field": "Заказ-наряд",
                            "message": "В файле не найдена колонка 'Заказ-наряд'"})
            return errors, valid_rows

        if rows and "Родительский ЗН" not in rows[0]:
            errors.append({"row": 0, "field": "Родительский ЗН",
                            "message": "В файле не найдена колонка 'Родительский ЗН'"})
            return errors, valid_rows

        if progress is not None:
            progress.update(processed=0, total=len(rows), phase="validating")

        for idx, row in enumerate(rows):
            row_num = idx + 1
            if progress is not None and (idx % 20 == 0 or row_num == len(rows)):
                progress["processed"] = row_num

            child_number = str(row.get("Заказ-наряд", "") or "").strip()
            parent_number = str(row.get("Родительский ЗН", "") or "").strip()

            if not child_number:
                errors.append({"row": row_num, "field": "Заказ-наряд", "message": "Поле 'Заказ-наряд' пустое"})
                continue

            if not parent_number:
                errors.append({"row": row_num, "field": "Родительский ЗН", "message": "Поле 'Родительский ЗН' пустое"})
                continue

            if child_number == parent_number:
                errors.append({"row": row_num, "field": "Родительский ЗН",
                                "message": f"Заказ-наряд не может быть родителем самому себе: '{child_number}'"})
                continue

            if child_number in batch_children:
                errors.append({"row": row_num, "field": "Заказ-наряд",
                                "message": f"Дубликат внутри файла: '{child_number}'"})
                continue

            result = await db_session.execute(select(Orders.id).where(Orders.order_number == child_number))
            child_id = result.scalar_one_or_none()
            if child_id is None:
                errors.append({"row": row_num, "field": "Заказ-наряд",
                                "message": f"Заказ-наряд не найден: '{child_number}'"})
                continue

            result = await db_session.execute(select(Orders.id).where(Orders.order_number == parent_number))
            parent_id = result.scalar_one_or_none()
            if parent_id is None:
                errors.append({"row": row_num, "field": "Родительский ЗН",
                                "message": f"Заказ-наряд не найден: '{parent_number}'"})
                continue

            batch_children.add(child_number)
            valid_rows.append((child_id, parent_id, child_number, parent_number))

        return errors, valid_rows

    @post("/generate-sql/start")
    async def generate_sql_start(
        self,
        request: Request,
        data: dict = Body(media_type=RequestEncodingType.MULTI_PART),
    ) -> Response:
        rows: list[dict] = json.loads(data.get("rows", "[]"))
        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_generate(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task
        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_generate(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                errors, valid_rows = await self._validate_and_build_rows(session, rows, progress=progress)
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка валидации: {e}"}])
            return

        if errors:
            progress.update(status="error", errors=errors)
            return

        sql_lines = self._build_sql_lines(valid_rows)
        progress.update(status="done", sql="\n".join(sql_lines), count=len(valid_rows))

    @post("/execute-sql/start")
    async def execute_sql_start(
        self,
        request: Request,
        data: dict = Body(media_type=RequestEncodingType.MULTI_PART),
    ) -> Response:
        rows: list[dict] = json.loads(data.get("rows", "[]"))
        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_execute(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task
        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_execute(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                try:
                    errors, valid_rows = await self._validate_and_build_rows(session, rows, progress=progress)
                except Exception as e:
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка валидации: {e}"}])
                    return

                if errors:
                    progress.update(status="error", errors=errors)
                    return

                if not valid_rows:
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": "Нет валидных строк для обновления"}])
                    return

                progress.update(processed=0, total=len(valid_rows), phase="executing")
                try:
                    for i, (child_id, parent_id, child_number, parent_number) in enumerate(valid_rows, start=1):
                        await session.execute(
                            text(
                                "INSERT INTO public.order_to_order (id_parent, id_child) "
                                "VALUES (:parent_id, :child_id) "
                                "ON CONFLICT (id_child) DO UPDATE SET id_parent = EXCLUDED.id_parent"
                            ),
                            {"parent_id": parent_id, "child_id": child_id},
                        )
                        if i % 20 == 0 or i == len(valid_rows):
                            progress["processed"] = i
                    await session.commit()
                except Exception as e:
                    await session.rollback()
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
                    return
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
            return

        now = datetime.now()
        log_lines = [
            f"=== Execute assign parent order: {now.strftime('%Y-%m-%d %H:%M:%S')} ===",
            f"Rows updated: {len(valid_rows)}",
            "",
            *self._build_sql_lines(valid_rows),
            "",
        ]

        log_file = LOG_DIR / f"assign_parent_order_{now.strftime('%Y-%m-%d_%H-%M-%S')}.log"
        with open(log_file, "a", encoding="utf-8") as f:
            f.write("\n".join(log_lines))
        logger.info("Assigned parent order for %d rows, log: %s", len(valid_rows), log_file)

        progress.update(status="done", count=len(valid_rows),
                         message=f"Успешно присвоено {len(valid_rows)} родительских ЗН")

    @get("/progress/{task_id:str}")
    async def get_progress(self, task_id: str) -> Response:
        state = _progress.get(task_id)
        if state is None:
            return Response(
                content=json.dumps({"status": "error", "errors": [{"row": 0, "field": "*", "message": "Задача не найдена или устарела"}]}),
                status_code=200,
                media_type="application/json",
            )
        return Response(
            content=json.dumps({k: v for k, v in state.items() if k != "created_at"}),
            status_code=200,
            media_type="application/json",
        )

    @staticmethod
    def _build_sql_lines(valid_rows: list[tuple[int, int, str, str]]) -> list[str]:
        sql_lines: list[str] = []
        for child_id, parent_id, child_number, parent_number in valid_rows:
            sql_lines.append(
                f"INSERT INTO public.order_to_order (id_parent, id_child) VALUES ({parent_id}, {child_id}) "
                f"ON CONFLICT (id_child) DO UPDATE SET id_parent = EXCLUDED.id_parent; "
                f"-- '{sql_escape(child_number)}' -> '{sql_escape(parent_number)}'"
            )
        return sql_lines
