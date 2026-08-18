import asyncio
import json
import logging
import re
import tempfile
import time
import uuid
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from sqlalchemy import bindparam, select, text
from sqlalchemy.ext.asyncio import AsyncSession
from litestar import Controller, get, post
from litestar.connection.request import Request
from litestar.datastructures import UploadFile
from litestar.enums import RequestEncodingType
from litestar.params import Body
from litestar.response import Template, Response, Redirect

from db_manager import get_session_maker
from models import TrainType, CarPlace, DesignNumber, Models, Train, Storage, Consignment, User
from schemas import (
    SelectSheetRequest,
    GenerateSQLResponse, ExecuteSQLResponse,
)
from sql_utils import sql_escape
from parser_storage import (
    LOG_DIR,
    load_data as _load_data,
    save_data as _save_data,
    cleanup_old_files as _cleanup_old_files,
)

logger = logging.getLogger("parser")

# Validating "set serial='none' lcn" issues a database query per unique
# id_train_type, which on large files takes seconds — progress is served through
# a separate poll rather than holding one HTTP request open all that time.
PROGRESS_TTL_SECONDS = 15 * 60
_progress: dict[str, dict] = {}
# asyncio only keeps a weak reference to fire-and-forget tasks — without
# storing it explicitly the task can be garbage collected before it finishes.
_tasks: dict[str, asyncio.Task] = {}

# A model lcn like 'M9.6.5': the digits after the letter prefix and before the
# first dot are id_train_type; the rest of the path is carried over as is.
_MODEL_LCN_RE = re.compile(r"^\D*(\d+)(?:\.(.*))?$")

# Excel dates in the project's other parsers are entered in Moscow time while
# the database (relocate.date/date_current) stores UTC — the same shift as in
# ptoir_parser.py (MSK_OFFSET). Here no date comes from the file, so the shift is
# applied once to "now" for the whole move batch.
MOVE_TZ_SHIFT = timedelta(hours=3)


def _cleanup_progress() -> None:
    cutoff = time.time() - PROGRESS_TTL_SECONDS
    stale = [tid for tid, state in _progress.items() if state["created_at"] < cutoff]
    for tid in stale:
        _progress.pop(tid, None)


def _parse_model_lcn(lcn: str) -> tuple[int, str] | None:
    """Extract (id_train_type, rest_of_path) from an lcn like 'M9.6.5' -> (9, '6.5'); 'M9' -> (9, '')."""
    match = _MODEL_LCN_RE.match(lcn)
    if not match:
        return None
    return int(match.group(1)), match.group(2) or ""


class ParserController(Controller):
    path = "/parser"

    @get("/")
    async def index(self, request: Request, page: int = 1, per_page: int = 10, select_sheet: bool = False) -> Template:
        page = max(page, 1)
        per_page = min(per_page, 200)
        error: str = request.session.pop("parser_error", "")

        pending_sheets: list[str] = []
        pending_filename: str = ""
        if select_sheet:
            pending_sheets = request.session.get("parser_pending_sheets", [])
            pending_filename = request.session.get("parser_pending_filename", "")

        session_id = request.session.get("parser_session_id", "")
        stored = _load_data(session_id) if session_id else None

        all_rows: list[dict] = stored["rows"] if stored else []
        headers: list[str] = stored["headers"] if stored else []
        filename: str = stored["filename"] if stored else ""

        total = len(all_rows)
        total_pages = max((total + per_page - 1) // per_page, 1)
        if page > total_pages:
            page = total_pages

        offset = (page - 1) * per_page
        rows = all_rows[offset:offset + per_page]

        return Template(
            template_name="parser.html",
            context={
                "headers": headers,
                "rows": rows,
                "all_rows": all_rows,
                "filename": filename,
                "error": error,
                "page": page,
                "per_page": per_page,
                "total": total,
                "total_pages": total_pages,
                "user_id": request.session.get("user_id"),
                "fullname": request.session.get("fullname", ""),
                "active_page": "parser",
                "pending_sheets": pending_sheets,
                "pending_filename": pending_filename,
            },
        )

    @post("/upload")
    async def upload(self, request: Request) -> Redirect:
        """Upload of an Excel file (.xlsx/.xls).

        Parses the file and extracts headers and rows.
        If the file holds several sheets, redirects to the sheet picker.
        """
        form = await request.form()
        upload_file: UploadFile | None = form.get("file")

        if not upload_file or not upload_file.filename:
            request.session["parser_error"] = "Файл не выбран"
            return Redirect("/parser")

        suffix = Path(upload_file.filename).suffix.lower()
        if suffix not in (".xlsx", ".xls"):
            request.session["parser_error"] = "Поддерживаются только .xlsx и .xls файлы"
            return Redirect("/parser")

        try:
            _cleanup_old_files()

            content = await upload_file.read()
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                tmp.write(content)
                tmp_path = tmp.name

            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()

            if len(sheet_names) > 1:
                request.session["parser_pending_file"] = tmp_path
                request.session["parser_pending_sheets"] = sheet_names
                request.session["parser_pending_filename"] = upload_file.filename
                return Redirect("/parser?select_sheet=1")

            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            ws = wb.active

            rows: list[dict[str, str | None]] = []
            headers: list[str] = []

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c) if c else f"col_{i}" for i, c in enumerate(row)]
                    continue
                rows.append({headers[j]: row[j] for j in range(len(row))})

            wb.close()
            Path(tmp_path).unlink(missing_ok=True)

            session_id = uuid.uuid4().hex
            _save_data(session_id, {
                "headers": headers,
                "rows": rows,
                "filename": upload_file.filename,
            })
            request.session["parser_session_id"] = session_id

            return Redirect("/parser")
        except Exception as e:
            request.session["parser_error"] = f"Ошибка чтения файла: {e}"
            return Redirect("/parser")

    @post("/select-sheet")
    async def select_sheet(
        self,
        request: Request,
        data: SelectSheetRequest = Body(media_type=RequestEncodingType.URL_ENCODED),
    ) -> Redirect:
        """Sheet choice for a multi-sheet Excel file.

        Parses the chosen sheet and stores the data for further processing.
        """
        sheet_name = data.sheet_name

        tmp_path = request.session.get("parser_pending_file", "")
        filename = request.session.get("parser_pending_filename", "")
        sheet_names = request.session.get("parser_pending_sheets", [])

        if not tmp_path or not Path(tmp_path).exists():
            request.session["parser_error"] = "Временный файл истёк. Загрузите файл заново."
            return Redirect("/parser")

        if sheet_name not in sheet_names:
            request.session["parser_error"] = "Выбранный лист не найден в файле."
            return Redirect("/parser")

        try:
            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            ws = wb[sheet_name]

            rows: list[dict[str, str | None]] = []
            headers: list[str] = []

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c) if c else f"col_{i}" for i, c in enumerate(row)]
                    continue
                rows.append({headers[j]: row[j] for j in range(len(row))})

            wb.close()
            Path(tmp_path).unlink(missing_ok=True)

            request.session.pop("parser_pending_file", None)
            request.session.pop("parser_pending_sheets", None)
            request.session.pop("parser_pending_filename", None)

            session_id = uuid.uuid4().hex
            _save_data(session_id, {
                "headers": headers,
                "rows": rows,
                "filename": f"{filename} [{sheet_name}]",
            })
            request.session["parser_session_id"] = session_id

            return Redirect("/parser")
        except Exception as e:
            request.session["parser_error"] = f"Ошибка чтения листа: {e}"
            return Redirect("/parser")

    async def _validate_and_build_rows(
        self, db_session: AsyncSession, rows: list[dict], progress: dict | None = None,
    ) -> tuple[list[dict[str, str]], list[tuple[int, int, int, str, bool]]]:
        if progress is not None:
            progress.update(processed=0, total=len(rows), phase="validating")

        existing_rows = await db_session.execute(
            select(Models.id_train_type, Models.lcn, Models.id_car_place,
                   Models.id_design_number, Models.is_default)
        )
        existing_set: set[tuple] = set()
        for er in existing_rows.all():
            existing_set.add((er[0], er[1], er[2], er[3], er[4]))

        existing_default_lcn_car: set[tuple] = set()
        existing_default_car_type_design: set[tuple] = set()
        for er in existing_set:
            if er[4]:
                existing_default_lcn_car.add((er[1], er[2]))
                existing_default_car_type_design.add((er[2], er[0], er[3]))

        errors: list[dict[str, str]] = []
        valid_rows: list[tuple[int, int, int, str, bool]] = []

        batch_full: set[tuple] = set()
        batch_default_lcn_car: set[tuple] = set()
        batch_default_car_type_design: set[tuple] = set()

        for idx, row in enumerate(rows):
            row_num = idx + 1
            if progress is not None:
                progress["processed"] = row_num
            model_name = str(row.get("model", "")).strip()
            position = str(row.get("position", "")).strip()
            itemnum = str(row.get("itemnum", "")).strip()
            lcn = str(row.get("lsn", "") or row.get("lcn", "")).strip()
            isdefault = str(row.get("isdefault", "")).strip().lower()
            is_default = isdefault == "true"

            train_type_id: int | None = None
            if model_name:
                result = await db_session.execute(
                    select(TrainType.id).where(TrainType.name == model_name)
                )
                r = result.scalar_one_or_none()
                if r is not None:
                    train_type_id = r
                else:
                    errors.append({"row": row_num, "field": "model",
                                   "message": f"train_type не найден: '{model_name}'"})

            car_place_id: int | None = None
            if position and position != "null":
                result = await db_session.execute(
                    select(CarPlace.id).where(CarPlace.name == position)
                )
                matches = result.scalars().all()
                if len(matches) == 1:
                    car_place_id = matches[0]
                elif len(matches) == 0:
                    errors.append({"row": row_num, "field": "position",
                                   "message": f"car_place не найден: '{position}'"})
                else:
                    errors.append({"row": row_num, "field": "position",
                                   "message": (f"car_place неоднозначен: найдено {len(matches)} записей "
                                               f"с именем '{position}' (id: {matches})")})

            design_number_id: int | None = None
            if itemnum:
                result = await db_session.execute(
                    select(DesignNumber.id).where(DesignNumber.number == itemnum)
                )
                r = result.scalar_one_or_none()
                if r is not None:
                    design_number_id = r
                else:
                    errors.append({"row": row_num, "field": "itemnum",
                                   "message": f"design_number не найден: '{itemnum}'"})

            if train_type_id is None or car_place_id is None or design_number_id is None:
                continue

            full_tuple = (train_type_id, lcn, car_place_id, design_number_id, is_default)
            if full_tuple in existing_set or full_tuple in batch_full:
                errors.append({
                    "row": row_num, "field": "*",
                    "message": (f"Дубликат: строка (train_type={train_type_id}, lcn='{lcn}', "
                                f"car_place={car_place_id}, design_number={design_number_id}, "
                                f"is_default={is_default}) уже существует"),
                })
                continue

            if is_default:
                if (lcn, car_place_id) in existing_default_lcn_car or (lcn, car_place_id) in batch_default_lcn_car:
                    errors.append({
                        "row": row_num, "field": "lcn",
                        "message": (f"Конфликт unique (lcn, car_place) WHERE is_default=true: "
                                    f"lcn='{lcn}', car_place={car_place_id} уже заняты"),
                    })
                    continue
                if ((car_place_id, train_type_id, design_number_id) in existing_default_car_type_design
                        or (car_place_id, train_type_id, design_number_id) in batch_default_car_type_design):
                    errors.append({
                        "row": row_num, "field": "*",
                        "message": (f"Конфликт unique (car_place, train_type, design_number) WHERE is_default=true: "
                                    f"car_place={car_place_id}, train_type={train_type_id}, "
                                    f"design_number={design_number_id} уже заняты"),
                    })
                    continue

            batch_full.add(full_tuple)
            if is_default:
                batch_default_lcn_car.add((lcn, car_place_id))
                batch_default_car_type_design.add((car_place_id, train_type_id, design_number_id))

            valid_rows.append((train_type_id, car_place_id, design_number_id, lcn, is_default))

        return errors, valid_rows

    async def _validate_and_build_serial_none_rows(
        self, db_session: AsyncSession, rows: list[dict], progress: dict | None = None,
    ) -> tuple[list[dict], list[dict]]:
        """Validate the Excel rows for 'set serial=none lcn'.

        In an lcn like 'M9.6.5', 9 is id_train_type and '6.5' is the rest of the
        path. For every train of that id_train_type its own id replaces 'M9',
        yielding the list of asset lcns ('lcn_trains') that need
        serial_number='none'.
        """
        errors: list[dict] = []
        valid_rows: list[dict] = []

        lcn_column: str | None = next(
            (k for k in (rows[0] if rows else {}) if str(k).strip().lower() in ("lsn", "lcn")),
            None,
        )
        if rows and lcn_column is None:
            errors.append({"row": 0, "field": "lcn", "message": "В файле не найдена колонка 'lsn' (или 'lcn')"})
            return errors, valid_rows

        if progress is not None:
            progress.update(processed=0, total=len(rows), phase="validating")

        train_ids_by_type: dict[int, list[int]] = {}

        for idx, row in enumerate(rows):
            row_num = idx + 1
            if progress is not None and (idx % 20 == 0 or row_num == len(rows)):
                progress["processed"] = row_num

            lcn_raw = str(row.get(lcn_column, "") or "").strip()
            if not lcn_raw:
                errors.append({"row": row_num, "field": "lcn", "message": "Пустой lcn"})
                continue

            parsed = _parse_model_lcn(lcn_raw)
            if parsed is None:
                errors.append({"row": row_num, "field": "lcn",
                               "message": f"Не удалось распознать id_train_type в lcn '{lcn_raw}'"})
                continue
            id_train_type, rest = parsed

            if id_train_type not in train_ids_by_type:
                result = await db_session.execute(select(Train.id).where(Train.id_train_type == id_train_type))
                train_ids_by_type[id_train_type] = [r[0] for r in result.all()]
            train_ids = train_ids_by_type[id_train_type]

            if not train_ids:
                errors.append({"row": row_num, "field": "lcn",
                               "message": f"Поезда с id_train_type={id_train_type} не найдены"})
                continue

            lcn_trains = [f"{tid}.{rest}" if rest else str(tid) for tid in train_ids]

            valid_rows.append({
                "row": row_num,
                "lcn": lcn_raw,
                "id_train_type": id_train_type,
                "lcn_trains": lcn_trains,
            })

        return errors, valid_rows

    async def _validate_and_build_move_no_relocate_rows(
        self, db_session: AsyncSession, rows: list[dict], progress: dict | None = None,
    ) -> tuple[list[dict], list[dict]]:
        """Validate the Excel rows for 'Переместить активы без relocate'.

        Each row gives the old and the new model lcn ('Старый lsn' -> 'lsn',
        both like 'M9.1.6') for one and the same position. The id_train_type
        (which must match on both) finds every train of that type, and for each
        a pair of real asset lcns is built (old -> new). Duplicate or
        overlapping file rows collapse into one pair; the same old lcn with
        different new ones is a conflict (an error).
        """
        errors: list[dict] = []
        pairs: dict[str, str] = {}
        pair_list: list[dict] = []

        new_column: str | None = next(
            (k for k in (rows[0] if rows else {}) if str(k).strip().lower() in ("lsn", "lcn")),
            None,
        )
        old_column: str | None = next(
            (k for k in (rows[0] if rows else {}) if str(k).strip().lower() in ("старый lsn", "старый lcn")),
            None,
        )
        if rows and new_column is None:
            errors.append({"row": 0, "field": "lcn", "message": "В файле не найдена колонка 'lsn' (или 'lcn')"})
            return errors, pair_list
        if rows and old_column is None:
            errors.append({"row": 0, "field": "lcn", "message": "В файле не найдена колонка 'Старый lsn' (или 'Старый lcn')"})
            return errors, pair_list

        if progress is not None:
            progress.update(processed=0, total=len(rows), phase="validating")

        train_ids_by_type: dict[int, list[int]] = {}

        for idx, row in enumerate(rows):
            row_num = idx + 1
            if progress is not None and (idx % 20 == 0 or row_num == len(rows)):
                progress["processed"] = row_num

            new_raw = str(row.get(new_column, "") or "").strip()
            old_raw = str(row.get(old_column, "") or "").strip()
            if not new_raw:
                errors.append({"row": row_num, "field": "lcn", "message": "Пустой lcn"})
                continue
            if not old_raw:
                errors.append({"row": row_num, "field": "lcn", "message": "Пустой Старый lcn"})
                continue

            new_parsed = _parse_model_lcn(new_raw)
            if new_parsed is None:
                errors.append({"row": row_num, "field": "lcn",
                               "message": f"Не удалось распознать id_train_type в lcn '{new_raw}'"})
                continue
            old_parsed = _parse_model_lcn(old_raw)
            if old_parsed is None:
                errors.append({"row": row_num, "field": "lcn",
                               "message": f"Не удалось распознать id_train_type в Старый lcn '{old_raw}'"})
                continue

            id_train_type_new, new_path = new_parsed
            id_train_type_old, old_path = old_parsed
            if id_train_type_new != id_train_type_old:
                errors.append({"row": row_num, "field": "lcn",
                               "message": (f"id_train_type не совпадает у 'Старый lsn' и 'lsn': "
                                           f"'{old_raw}' ({id_train_type_old}) vs '{new_raw}' ({id_train_type_new})")})
                continue

            if id_train_type_new not in train_ids_by_type:
                result = await db_session.execute(select(Train.id).where(Train.id_train_type == id_train_type_new))
                train_ids_by_type[id_train_type_new] = [r[0] for r in result.all()]
            train_ids = train_ids_by_type[id_train_type_new]

            if not train_ids:
                errors.append({"row": row_num, "field": "lcn",
                               "message": f"Поезда с id_train_type={id_train_type_new} не найдены"})
                continue

            row_pairs: list[tuple[str, str]] = []
            conflict = False
            for tid in train_ids:
                old_lcn = f"{tid}.{old_path}" if old_path else str(tid)
                new_lcn = f"{tid}.{new_path}" if new_path else str(tid)
                if old_lcn in pairs and pairs[old_lcn] != new_lcn:
                    errors.append({"row": row_num, "field": "lcn",
                                   "message": (f"Конфликт: '{old_lcn}' уже сопоставлен другому lcn "
                                               f"('{pairs[old_lcn]}', а не '{new_lcn}')")})
                    conflict = True
                    break
                row_pairs.append((old_lcn, new_lcn))
            if conflict:
                continue

            for old_lcn, new_lcn in row_pairs:
                if old_lcn not in pairs:
                    pairs[old_lcn] = new_lcn
                    pair_list.append({"old_lcn": old_lcn, "new_lcn": new_lcn})

        return errors, pair_list

    @staticmethod
    async def _check_lcn_collisions(db_session: AsyncSession, pair_list: list[dict]) -> list[dict]:
        """Find assets already holding a pair's target new_lcn that are not in the file.

        The two-phase UPDATE (see _build_two_phase_lcn_update_lines) is safe for
        chains WITHIN one batch (A->B, B->C), but does not help when the target
        new_lcn is already taken by an asset the file never mentions — then a
        UniqueViolationError surfaces mid-execution with no clear cause. It is a
        separate method (not part of _validate_and_build_move_no_relocate_rows)
        because it uses lcn::text, a Postgres-specific cast the SQLite test
        database does not support (see test_parser_change_lcn_validation.py).
        """
        if not pair_list:
            return []
        old_lcns = {vr["old_lcn"] for vr in pair_list}
        new_lcns = [vr["new_lcn"] for vr in pair_list]
        stmt = text(
            "SELECT active_number, lcn::text FROM public.actives WHERE lcn::text IN :lcns"
        ).bindparams(bindparam("lcns", expanding=True))
        result = await db_session.execute(stmt, {"lcns": new_lcns})
        errors: list[dict] = []
        for active_number, current_lcn in result.all():
            if current_lcn not in old_lcns:
                errors.append({
                    "row": 0, "field": "lcn",
                    "message": (f"Конфликт: целевой lcn '{current_lcn}' уже занят активом "
                                f"'{active_number}', которого нет в файле"),
                })
        return errors

    async def _validate_and_build_change_model_lcn_rows(
        self, db_session: AsyncSession, rows: list[dict], progress: dict | None = None,
    ) -> tuple[list[dict], list[dict]]:
        """Validate the Excel rows for 'изменить lcn в модели' (the public.models table).

        The 'id' column is models.id (the same scheme as the "Удалить из
        моделей"/"Добавить в модели" sheets in this very file); 'lsn'/'lcn' is
        the new models.lcn value. The optional 'Старый lsn'/'Старый lcn' is
        checked against the current models.lcn — a guard against a stale or wrong
        file; it takes no part in the update (matching always goes through the
        stable id, never the lcn text).
        """
        errors: list[dict] = []
        valid_rows: list[dict] = []

        id_column: str | None = next(
            (k for k in (rows[0] if rows else {}) if str(k).strip().lower() == "id"),
            None,
        )
        new_column: str | None = next(
            (k for k in (rows[0] if rows else {}) if str(k).strip().lower() in ("lsn", "lcn")),
            None,
        )
        old_column: str | None = next(
            (k for k in (rows[0] if rows else {}) if str(k).strip().lower() in ("старый lsn", "старый lcn")),
            None,
        )
        if rows and id_column is None:
            errors.append({"row": 0, "field": "id", "message": "В файле не найдена колонка 'id'"})
            return errors, valid_rows
        if rows and new_column is None:
            errors.append({"row": 0, "field": "lcn", "message": "В файле не найдена колонка 'lsn' (или 'lcn')"})
            return errors, valid_rows

        if progress is not None:
            progress.update(processed=0, total=len(rows), phase="validating")

        batch_ids: dict[int, str] = {}

        for idx, row in enumerate(rows):
            row_num = idx + 1
            if progress is not None and (idx % 20 == 0 or row_num == len(rows)):
                progress["processed"] = row_num

            id_raw = str(row.get(id_column, "") or "").strip()
            new_raw = str(row.get(new_column, "") or "").strip()
            old_raw = str(row.get(old_column, "") or "").strip() if old_column else ""

            if not id_raw:
                errors.append({"row": row_num, "field": "id", "message": "Поле 'id' пустое"})
                continue
            try:
                model_id = int(float(id_raw))
            except ValueError:
                errors.append({"row": row_num, "field": "id", "message": f"Некорректный id: '{id_raw}'"})
                continue

            if not new_raw:
                errors.append({"row": row_num, "field": "lcn", "message": "Пустой lcn"})
                continue

            if model_id in batch_ids and batch_ids[model_id] != new_raw:
                errors.append({"row": row_num, "field": "id",
                               "message": (f"Конфликт: id={model_id} уже сопоставлен другому lcn "
                                           f"('{batch_ids[model_id]}', а не '{new_raw}')")})
                continue

            result = await db_session.execute(
                text("SELECT lcn::text FROM public.models WHERE id = :id"), {"id": model_id}
            )
            current_lcn = result.scalar_one_or_none()
            if current_lcn is None:
                errors.append({"row": row_num, "field": "id", "message": f"Модель с id={model_id} не найдена"})
                continue

            if old_raw and current_lcn != old_raw:
                errors.append({"row": row_num, "field": "lcn",
                               "message": (f"Текущий lcn модели id={model_id} ('{current_lcn}') "
                                           f"не совпадает со 'Старый lsn' ('{old_raw}')")})
                continue

            if model_id not in batch_ids:
                batch_ids[model_id] = new_raw
                valid_rows.append({"id": model_id, "new_lcn": new_raw})

        return errors, valid_rows

    @staticmethod
    def _build_two_phase_lcn_update_lines(valid_rows: list[dict], extra_set: str = "") -> list[str]:
        """Two-phase lcn UPDATE: old -> temporary ('Z'+old) -> new.

        Files often contain chains (one pair's new lcn equals another pair's old
        one — a child position moving after its parent, say). Updating in a
        single FROM(VALUES...) statement fails with UniqueViolationError on
        actives_lcn_key: until every row has moved, the intermediate state holds
        a duplicate. The temporary 'Z' prefix cannot collide with real lcns
        (train_id, 'M' model ones, 'S' storage ones) — after the first pass no
        live lcn matches another pair's old or new value.

        extra_set holds additional ", column = value" clauses for the final
        UPDATE (resetting id_actves_parent/id_actives_root on a move, say).
        """
        if not valid_rows:
            return []
        old_list = ", ".join(f"'{sql_escape(vr['old_lcn'])}'" for vr in valid_rows)
        tmp_values_list = ", ".join(
            f"('Z{sql_escape(vr['old_lcn'])}', '{sql_escape(vr['new_lcn'])}')" for vr in valid_rows
        )
        return [
            f"UPDATE public.actives SET lcn = ('Z' || lcn::text)::ltree WHERE lcn::text IN ({old_list});",
            f"UPDATE public.actives AS act SET lcn = v.new_lcn::ltree{extra_set} "
            f"FROM (VALUES {tmp_values_list}) AS v(tmp_lcn, new_lcn) WHERE act.lcn::text = v.tmp_lcn;",
        ]

    @staticmethod
    def _build_move_no_relocate_sql_lines(valid_rows: list[dict]) -> list[str]:
        """The same two-phase lcn UPDATE as in 'Изменить lcn в модели' in spirit,
        but on actives here (no id_location change and no relocate)."""
        return ParserController._build_two_phase_lcn_update_lines(valid_rows)

    @staticmethod
    def _build_change_lcn_sql_lines(valid_rows: list[dict]) -> list[str]:
        """Two-phase UPDATE of public.models.lcn, matching on models.id.

        Unlike the actives variant, matching goes through the stable id
        (models.id does not change, lcn does) rather than the lcn text — the id
        alone identifies the row. The 'Z' prefix is still needed: lcn is part of
        a composite UNIQUE (id_train_type, lcn, id_car_place, id_design_number,
        is_default), and the per-row processing order within one UPDATE is not
        guaranteed — without a temporary value chains can raise
        UniqueViolationError (see _build_two_phase_lcn_update_lines).
        """
        if not valid_rows:
            return []
        id_list = ", ".join(str(vr["id"]) for vr in valid_rows)
        values_list = ", ".join(f"({vr['id']}, '{sql_escape(vr['new_lcn'])}')" for vr in valid_rows)
        return [
            f"UPDATE public.models SET lcn = ('Z' || lcn::text)::ltree WHERE id IN ({id_list});",
            "UPDATE public.models AS m SET lcn = v.new_lcn::ltree "
            f"FROM (VALUES {values_list}) AS v(mid, new_lcn) WHERE m.id = v.mid;",
        ]

    async def _validate_and_build_is_default_rows(
        self, db_session: AsyncSession, rows: list[dict], progress: dict | None = None,
    ) -> tuple[list[dict], list[dict]]:
        """Validate the Excel rows for 'Изменить серийность в модели' (models.is_default).

        The 'id' column is models.id and 'isdefault' the new value
        (true/false/1/0/да/нет, as in _validate_is_serial_1c in
        design_number_parser.py). Switching to true additionally checks both
        partial UNIQUE indexes on models over is_default=true (see
        _validate_and_build_rows) — otherwise the conflict surfaces as a bare
        UniqueViolationError during execution instead of a clear per-row error.
        """
        errors: list[dict] = []

        id_column: str | None = next(
            (k for k in (rows[0] if rows else {}) if str(k).strip().lower() == "id"),
            None,
        )
        isdefault_column: str | None = next(
            (k for k in (rows[0] if rows else {}) if str(k).strip().lower() in ("isdefault", "is_default")),
            None,
        )
        if rows and id_column is None:
            errors.append({"row": 0, "field": "id", "message": "В файле не найдена колонка 'id'"})
            return errors, []
        if rows and isdefault_column is None:
            errors.append({"row": 0, "field": "isdefault", "message": "В файле не найдена колонка 'isdefault'"})
            return errors, []

        if progress is not None:
            progress.update(processed=0, total=len(rows), phase="validating")

        models_result = await db_session.execute(
            select(Models.id, Models.id_train_type, Models.lcn, Models.id_car_place,
                   Models.id_design_number, Models.is_default)
        )
        models_by_id = {m[0]: m for m in models_result.all()}

        # First pass: parse the columns and dedupe by id — it does not look at the
        # current default sets, so the order of the file's rows cannot influence the
        # result (which matters for the second pass, see below).
        parsed: dict[int, dict] = {}
        order: list[int] = []

        for idx, row in enumerate(rows):
            row_num = idx + 1
            if progress is not None and (idx % 20 == 0 or row_num == len(rows)):
                progress["processed"] = row_num

            id_raw = str(row.get(id_column, "") or "").strip()
            isdefault_raw = str(row.get(isdefault_column, "") or "").strip().lower()

            if not id_raw:
                errors.append({"row": row_num, "field": "id", "message": "Поле 'id' пустое"})
                continue
            try:
                model_id = int(float(id_raw))
            except ValueError:
                errors.append({"row": row_num, "field": "id", "message": f"Некорректный id: '{id_raw}'"})
                continue

            if isdefault_raw not in ("true", "false", "1", "0", "да", "нет"):
                errors.append({"row": row_num, "field": "isdefault",
                               "message": f"Неверное значение isdefault: '{isdefault_raw}' (ожидается true/false)"})
                continue
            is_default = isdefault_raw in ("true", "1", "да")

            if model_id in parsed:
                if parsed[model_id]["is_default"] != is_default:
                    errors.append({"row": row_num, "field": "id",
                                   "message": (f"Конфликт: id={model_id} уже сопоставлен другому значению isdefault "
                                               f"({parsed[model_id]['is_default']}, а не {is_default})")})
                continue

            if model_id not in models_by_id:
                errors.append({"row": row_num, "field": "id", "message": f"Модель с id={model_id} не найдена"})
                continue

            parsed[model_id] = {"row": row_num, "is_default": is_default}
            order.append(model_id)

        # Second pass: UNIQUE collisions on is_default=true (see
        # _validate_and_build_rows). Models from this same file count by their NEW
        # value from the file, not the current one in the database — otherwise a file
        # that clears the old default and sets a new one at the same (lcn, car_place)
        # or (car_place, train_type, design_number) would raise a false conflict from
        # the arbitrary row order (the same case that makes _build_is_default_sql_lines
        # sort FALSE before TRUE).
        existing_default_lcn_car: dict[tuple, int] = {}
        existing_default_car_type_design: dict[tuple, int] = {}
        for mid, m in models_by_id.items():
            if mid in parsed or not m[5]:
                continue
            existing_default_lcn_car[(m[2], m[3])] = mid
            existing_default_car_type_design[(m[3], m[1], m[4])] = mid

        batch_lcn_car: dict[tuple, int] = {}
        batch_car_type_design: dict[tuple, int] = {}
        valid_ids: list[int] = []

        for model_id in order:
            info = parsed[model_id]
            row_num = info["row"]
            is_default = info["is_default"]
            _, id_train_type, lcn, id_car_place, id_design_number, _ = models_by_id[model_id]

            if is_default:
                lcn_car = (lcn, id_car_place)
                car_type_design = (id_car_place, id_train_type, id_design_number)
                owner = existing_default_lcn_car.get(lcn_car, batch_lcn_car.get(lcn_car))
                if owner is not None and owner != model_id:
                    errors.append({"row": row_num, "field": "isdefault",
                                   "message": (f"Конфликт unique (lcn, car_place) WHERE is_default=true: "
                                               f"lcn='{lcn}', car_place={id_car_place} уже заняты")})
                    continue
                owner2 = existing_default_car_type_design.get(car_type_design, batch_car_type_design.get(car_type_design))
                if owner2 is not None and owner2 != model_id:
                    errors.append({"row": row_num, "field": "isdefault",
                                   "message": (f"Конфликт unique (car_place, train_type, design_number) WHERE is_default=true: "
                                               f"car_place={id_car_place}, train_type={id_train_type}, "
                                               f"design_number={id_design_number} уже заняты")})
                    continue
                batch_lcn_car[lcn_car] = model_id
                batch_car_type_design[car_type_design] = model_id

            valid_ids.append(model_id)

        valid_rows = [{"id": mid, "is_default": parsed[mid]["is_default"]} for mid in valid_ids]
        return errors, valid_rows

    @staticmethod
    def _build_is_default_sql_lines(valid_rows: list[dict]) -> list[str]:
        """FALSE rows go before TRUE ones: the partial UNIQUE index (WHERE is_default=true)

        is checked row by row as the individual UPDATEs run inside one transaction
        rather than at the end — so if a file clears the old default and sets a new one
        at the same (lcn, car_place) or (car_place, train_type, design_number), the
        clearing must run before the setting.
        """
        ordered = sorted(valid_rows, key=lambda vr: vr["is_default"])
        return [
            f"UPDATE public.models SET is_default = {'TRUE' if vr['is_default'] else 'FALSE'} WHERE id = {vr['id']};"
            for vr in ordered
        ]

    @staticmethod
    def _merge_serial_none_lcns(valid_rows: list[dict]) -> list[str]:
        """Merge the lcn_trains of every file row into one duplicate-free list.

        Different Excel rows with the same lsn (or different lsns yielding
        overlapping lcns) would otherwise produce several identical or
        overlapping UPDATEs in a row — this makes it one query for the whole file.
        """
        seen: set[str] = set()
        merged: list[str] = []
        for vr in valid_rows:
            for l in vr["lcn_trains"]:
                if l not in seen:
                    seen.add(l)
                    merged.append(l)
        return merged

    @staticmethod
    def _build_serial_none_sql_lines(valid_rows: list[dict]) -> list[str]:
        lcns = ParserController._merge_serial_none_lcns(valid_rows)
        if not lcns:
            return []
        lcn_list = ", ".join(f"'{sql_escape(l)}'" for l in lcns)
        return [f"UPDATE public.actives SET serial_number = 'none' WHERE lcn::text IN ({lcn_list});"]

    @staticmethod
    async def _resolve_user_by_fullname(db_session: AsyncSession, fullname: str) -> int | None:
        """Look up fdw_users by the full name built exactly as session['fullname'] in auth.py:

        ' '.join(filter(None, [lastname, firstname, middlename])) or username.
        """
        result = await db_session.execute(select(User.id, User.lastname, User.firstname, User.middlename, User.username))
        for uid, lastname, firstname, middlename, username in result.all():
            built = " ".join(filter(None, [lastname, firstname, middlename])) or username or ""
            if built.strip() == fullname:
                return uid
        return None

    async def _validate_and_build_move_rows(
        self, db_session: AsyncSession, rows: list[dict],
        storage_name: str, consignment_name: str, user_fullname: str, set_nocm: bool,
        progress: dict | None = None,
    ) -> tuple[list[dict], list[dict], int | None, int | None, int | None, int | None]:
        """Validate the Excel rows for 'Переместить активы' (the move_active equivalent).

        Assets are located through lsn/lcn by the same logic as the
        "set serial='none' lcn" button (_validate_and_build_serial_none_rows):
        a model lcn ('M9.6.5') yields every train of that type and their asset
        lcns, among which the assets that actually exist are then looked up.
        Rows with no asset at their lcn simply do not appear in the result —
        not an error (the position may already have been removed from some
        trains).

        Storage, consignment and user are shared by the whole file (set once in
        the modal). set_nocm is the "Установить позицию ТМЦ = 'NOCM'" checkbox:
        when on, id_design_number for every moved asset is resolved through
        design_number.number == 'NOCM' and additionally written by the UPDATE.
        Returns (errors, valid_rows, id_storage, id_consignment, id_user, id_design_number).
        """
        errors: list[dict] = []
        valid_rows: list[dict] = []

        storage_name = storage_name.strip()
        consignment_name = consignment_name.strip()
        user_fullname = user_fullname.strip()

        id_storage: int | None = None
        if not storage_name:
            errors.append({"row": 0, "field": "Склад", "message": "Поле 'Склад' пустое"})
        else:
            id_storage = await db_session.scalar(select(Storage.id).where(Storage.name == storage_name))
            if id_storage is None:
                errors.append({"row": 0, "field": "Склад", "message": f"Склад не найден: '{storage_name}'"})

        id_consignment: int | None = None
        if not consignment_name:
            errors.append({"row": 0, "field": "Партия", "message": "Поле 'Партия' пустое"})
        else:
            id_consignment = await db_session.scalar(select(Consignment.id).where(Consignment.name == consignment_name))
            if id_consignment is None:
                errors.append({"row": 0, "field": "Партия", "message": f"Партия не найдена: '{consignment_name}'"})

        id_user: int | None = None
        if not user_fullname:
            errors.append({"row": 0, "field": "Пользователь", "message": "Поле 'Пользователь' пустое"})
        else:
            id_user = await self._resolve_user_by_fullname(db_session, user_fullname)
            if id_user is None:
                errors.append({"row": 0, "field": "Пользователь", "message": f"Пользователь не найден: '{user_fullname}'"})

        id_design_number: int | None = None
        if set_nocm:
            id_design_number = await db_session.scalar(select(DesignNumber.id).where(DesignNumber.number == "NOCM"))
            if id_design_number is None:
                errors.append({"row": 0, "field": "ТМЦ", "message": "Позиция ТМЦ 'NOCM' не найдена"})

        if errors:
            return errors, valid_rows, id_storage, id_consignment, id_user, id_design_number

        lcn_errors, lcn_rows = await self._validate_and_build_serial_none_rows(db_session, rows, progress=progress)
        if lcn_errors:
            return lcn_errors, valid_rows, id_storage, id_consignment, id_user, id_design_number

        merged_lcns = self._merge_serial_none_lcns(lcn_rows)
        if not merged_lcns:
            errors.append({"row": 0, "field": "lcn", "message": "Не найдено ни одного lcn для перемещения"})
            return errors, valid_rows, id_storage, id_consignment, id_user, id_design_number

        stmt = text(
            "SELECT id, active_number, id_location FROM public.actives WHERE lcn::text IN :lcns"
        ).bindparams(bindparam("lcns", expanding=True))
        result = await db_session.execute(stmt, {"lcns": merged_lcns})
        for row_num, (id_active, active_number, id_location_old) in enumerate(result.all(), start=1):
            valid_rows.append({
                "row": row_num,
                "active_number": active_number,
                "id_active": id_active,
                "id_location_old": id_location_old,
            })

        return errors, valid_rows, id_storage, id_consignment, id_user, id_design_number

    @staticmethod
    def _build_move_actives_sql_body(
        valid_rows: list[dict], id_storage: int, id_consignment: int, id_user: int,
        reason: str, move_date: datetime, id_design_number: int | None = None,
    ) -> list[str]:
        """Build the DO block moving assets into storage (the move_active equivalent).

        One destination storage for the whole batch (unlike the original script,
        where 'Куда' was taken per row) — which reduces the lcn counter to a
        single variable instead of a dict keyed by storage id. The lcn is taken
        from storage.last_lcn under FOR UPDATE and written back at the end of the
        DO block (unlike the original move_active, which incremented the counter
        only in Python memory and never saved it back — a rerun would hand out
        the same lcn numbers again).

        id_design_number (the "Установить позицию ТМЦ = 'NOCM'" checkbox), when
        set, is additionally written by the same UPDATE on actives.
        """
        total = len(valid_rows)
        date_str = move_date.strftime("%Y-%m-%d %H:%M:%S")
        reason_val = f"'{sql_escape(reason)}'" if reason else "NULL"

        sql_lines: list[str] = ["DO $$", "DECLARE"]
        sql_lines.append(
            f"    loc_ids bigint[] := ARRAY(SELECT nextval('public.location_id_seq') "
            f"FROM generate_series(1, {total}));"
        )
        sql_lines.append("    lcn_new bigint;")
        sql_lines.append("BEGIN")
        sql_lines.append(f"    SELECT last_lcn INTO lcn_new FROM public.storage WHERE id = {id_storage} FOR UPDATE;")

        for i, vr in enumerate(valid_rows, start=1):
            loc_ref = f"loc_ids[{i}]"
            old_loc_val = str(vr["id_location_old"]) if vr["id_location_old"] is not None else "NULL"
            sql_lines.append("    lcn_new := lcn_new + 1;")
            sql_lines.append(
                f"    INSERT INTO public.location (id, id_type_location, id_storage, id_consignment) "
                f"VALUES ({loc_ref}, 1, {id_storage}, {id_consignment});"
            )
            sql_lines.append(
                f"    INSERT INTO public.relocate (id_location_old, id_location_new, date, id_user, id_active, "
                f"reason, date_current, id_order) VALUES ({old_loc_val}, {loc_ref}, '{date_str}', {id_user}, "
                f"{vr['id_active']}, {reason_val}, '{date_str}', NULL);"
            )
            design_number_clause = f"id_design_number = {id_design_number}, " if id_design_number is not None else ""
            active_number_comment = str(vr["active_number"]).replace("\n", " ").replace("\r", " ")
            sql_lines.append(
                f"    UPDATE public.actives SET id_location = {loc_ref}, id_actves_parent = NULL, "
                f"id_actives_root = NULL, {design_number_clause}lcn = ('S{id_storage}.' || lcn_new)::ltree "
                f"WHERE id = {vr['id_active']}; -- {sql_escape(active_number_comment)}"
            )

        sql_lines.append(f"    UPDATE public.storage SET last_lcn = lcn_new WHERE id = {id_storage};")
        sql_lines.append("END $$;")
        return sql_lines

    @staticmethod
    def _build_insert_sql_lines(valid_rows: list[tuple[int, int, int, str, bool]]) -> list[str]:
        sql_lines: list[str] = []
        for train_type_id, car_place_id, design_number_id, lcn, is_default in valid_rows:
            isdefault_val = "TRUE" if is_default else "FALSE"
            sql_lines.append(
                f"INSERT INTO public.models (id_train_type, id_car_place, id_design_number, lcn, is_default) "
                f"VALUES ({train_type_id}, {car_place_id}, {design_number_id}, '{sql_escape(lcn)}', {isdefault_val});"
            )
        return sql_lines

    @staticmethod
    def _parse_delete_ids(rows: list[dict], progress: dict | None = None) -> tuple[list[dict], list[int]]:
        errors: list[dict] = []
        valid_ids: list[int] = []
        if progress is not None:
            progress.update(processed=0, total=len(rows), phase="validating")
        for idx, row in enumerate(rows):
            row_num = idx + 1
            if progress is not None:
                progress["processed"] = row_num
            row_id = row.get("id")
            if not row_id:
                errors.append({"row": row_num, "field": "id",
                               "message": "Поле 'id' отсутствует или пустое"})
                continue
            valid_ids.append(int(row_id))
        return errors, valid_ids

    @staticmethod
    def _build_delete_sql_lines(valid_ids: list[int]) -> list[str]:
        return [f"DELETE FROM public.models WHERE id = {rid};" for rid in valid_ids]

    @post("/generate-sql/start")
    async def generate_sql_start(
        self,
        request: Request,
        data: dict = Body(media_type=RequestEncodingType.MULTI_PART),
    ) -> Response:
        """Start background generation of the SQL file inserting rows into models; returns a task_id to poll."""
        rows: list[dict] = json.loads(data.get("rows", "[]"))
        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_insert_generate(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task
        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_insert_generate(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                errors, valid_rows = await self._validate_and_build_rows(session, rows, progress=progress)
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка валидации: {e}"}])
            return

        if errors:
            progress.update(status="error", errors=errors)
            return

        sql_lines = self._build_insert_sql_lines(valid_rows)
        progress.update(status="done", sql="\n".join(sql_lines), count=len(valid_rows))

    @post("/execute-sql/start")
    async def execute_sql_start(
        self,
        request: Request,
        data: dict = Body(media_type=RequestEncodingType.MULTI_PART),
    ) -> Response:
        """Start the background atomic insert of rows into public.models; returns a task_id to poll."""
        rows: list[dict] = json.loads(data.get("rows", "[]"))
        skip_errors = str(data.get("skip_errors", "")).strip().lower() == "true"
        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_insert_execute(task_id, rows, skip_errors=skip_errors))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task
        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_insert_execute(self, task_id: str, rows: list[dict], skip_errors: bool = False) -> None:
        progress = _progress[task_id]
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                try:
                    errors, valid_rows = await self._validate_and_build_rows(session, rows, progress=progress)
                except Exception as e:
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка валидации: {e}"}])
                    return

                if errors and not skip_errors:
                    progress.update(status="confirm_errors", errors=errors, valid_count=len(valid_rows))
                    return

                if not valid_rows:
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": "Нет валидных строк для вставки"}])
                    return

                progress.update(processed=0, total=len(valid_rows), phase="executing")
                try:
                    for i, (train_type_id, car_place_id, design_number_id, lcn, is_default) in enumerate(valid_rows, start=1):
                        await session.execute(
                            text(
                                "INSERT INTO public.models (id_train_type, id_car_place, id_design_number, lcn, is_default) "
                                "VALUES (:tt, :cp, :dn, :lcn, :def)"
                            ),
                            {"tt": train_type_id, "cp": car_place_id, "dn": design_number_id, "lcn": lcn, "def": is_default},
                        )
                        progress["processed"] = i
                    await session.commit()
                except Exception as e:
                    await session.rollback()
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
                    return
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
            return

        now = datetime.now()
        log_lines = [
            f"=== Execute SQL: {now.strftime('%Y-%m-%d %H:%M:%S')} ===",
            f"Rows inserted: {len(valid_rows)}",
            "",
            *self._build_insert_sql_lines(valid_rows),
            "",
        ]
        log_file = LOG_DIR / f"insert_models_{now.strftime('%Y-%m-%d_%H-%M-%S')}.log"
        with open(log_file, "a", encoding="utf-8") as f:
            f.write("\n".join(log_lines))
        logger.info("SQL executed: %d rows inserted, log saved to %s", len(valid_rows), log_file)

        message = f"Успешно вставлено {len(valid_rows)} строк"
        if errors:
            message += f" (пропущено с ошибками: {len(errors)})"
        progress.update(status="done", count=len(valid_rows), message=message, errors=errors)

    @post("/delete-rows/start")
    async def delete_rows_start(
        self,
        request: Request,
        data: dict = Body(media_type=RequestEncodingType.MULTI_PART),
    ) -> Response:
        """Start background generation of the SQL file deleting rows from models; returns a task_id to poll."""
        rows: list[dict] = json.loads(data.get("rows", "[]"))
        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_delete_generate(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task
        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_delete_generate(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        errors, valid_ids = self._parse_delete_ids(rows, progress=progress)
        if errors:
            progress.update(status="error", errors=errors)
            return

        sql_lines = self._build_delete_sql_lines(valid_ids)
        progress.update(status="done", sql="\n".join(sql_lines), count=len(sql_lines))

    @post("/execute-delete/start")
    async def execute_delete_start(
        self,
        request: Request,
        data: dict = Body(media_type=RequestEncodingType.MULTI_PART),
    ) -> Response:
        """Start the background atomic delete of rows from public.models; returns a task_id to poll."""
        rows: list[dict] = json.loads(data.get("rows", "[]"))
        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_delete_execute(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task
        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_delete_execute(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        errors, valid_ids = self._parse_delete_ids(rows, progress=progress)
        if errors:
            progress.update(status="error", errors=errors)
            return

        if not valid_ids:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": "Нет валидных строк для удаления"}])
            return

        progress.update(processed=0, total=len(valid_ids), phase="executing")
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                try:
                    for i, rid in enumerate(valid_ids, start=1):
                        await session.execute(text("DELETE FROM public.models WHERE id = :id"), {"id": rid})
                        progress["processed"] = i
                    await session.commit()
                except Exception as e:
                    await session.rollback()
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
                    return
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
            return

        now = datetime.now()
        log_lines = [
            f"=== Execute Delete: {now.strftime('%Y-%m-%d %H:%M:%S')} ===",
            f"Rows deleted: {len(valid_ids)}",
            "",
            *self._build_delete_sql_lines(valid_ids),
            "",
        ]
        log_file = LOG_DIR / f"delete_models_{now.strftime('%Y-%m-%d_%H-%M-%S')}.log"
        with open(log_file, "a", encoding="utf-8") as f:
            f.write("\n".join(log_lines))
        logger.info("SQL executed: %d rows deleted, log saved to %s", len(valid_ids), log_file)

        progress.update(status="done", count=len(valid_ids), message=f"Успешно удалено {len(valid_ids)} строк")

    @post("/serial-none/generate-sql/start")
    async def serial_none_generate_sql_start(
        self,
        request: Request,
        data: dict = Body(media_type=RequestEncodingType.MULTI_PART),
    ) -> Response:
        """Start background generation of the 'set serial=none lcn' SQL file; returns a task_id to poll."""
        rows: list[dict] = json.loads(data.get("rows", "[]"))
        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_serial_none_generate(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task
        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_serial_none_generate(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                errors, valid_rows = await self._validate_and_build_serial_none_rows(session, rows, progress=progress)
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка валидации: {e}"}])
            return

        if errors:
            progress.update(status="error", errors=errors)
            return

        sql_lines = self._build_serial_none_sql_lines(valid_rows)
        progress.update(status="done", sql="\n".join(sql_lines), count=len(valid_rows))

    @post("/serial-none/execute-sql/start")
    async def serial_none_execute_sql_start(
        self,
        request: Request,
        data: dict = Body(media_type=RequestEncodingType.MULTI_PART),
    ) -> Response:
        """Start the background atomic update of serial_number='none'; returns a task_id to poll."""
        rows: list[dict] = json.loads(data.get("rows", "[]"))
        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_serial_none_execute(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task
        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_serial_none_execute(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        total_updated = 0
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                try:
                    errors, valid_rows = await self._validate_and_build_serial_none_rows(session, rows, progress=progress)
                except Exception as e:
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка валидации: {e}"}])
                    return

                if errors:
                    progress.update(status="error", errors=errors)
                    return

                if not valid_rows:
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": "Нет валидных строк для обновления"}])
                    return

                # One shared UPDATE over the merged, duplicate-free list of lcns from
                # every file row — instead of a query per row, which with identical or
                # overlapping lsns produced several identical UPDATEs in a row.
                progress.update(processed=0, total=1, phase="executing")
                try:
                    lcns = self._merge_serial_none_lcns(valid_rows)
                    stmt = text(
                        "UPDATE public.actives SET serial_number = 'none' WHERE lcn::text IN :lcns"
                    ).bindparams(bindparam("lcns", expanding=True))
                    result = await session.execute(stmt, {"lcns": lcns})
                    total_updated = result.rowcount
                    progress["processed"] = 1
                    await session.commit()
                except Exception as e:
                    await session.rollback()
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
                    return
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
            return

        now = datetime.now()
        log_lines = [
            f"=== Execute serial-none update: {now.strftime('%Y-%m-%d %H:%M:%S')} ===",
            f"Rows processed: {len(valid_rows)}, actives updated: {total_updated}",
            "",
            *self._build_serial_none_sql_lines(valid_rows),
            "",
        ]
        log_file = LOG_DIR / f"serial_none_{now.strftime('%Y-%m-%d_%H-%M-%S')}.log"
        with open(log_file, "a", encoding="utf-8") as f:
            f.write("\n".join(log_lines))
        logger.info("Serial-none updated for %d rows (%d actives), log: %s", len(valid_rows), total_updated, log_file)

        progress.update(status="done", count=len(valid_rows),
                         message=f"Обновлено активов: {total_updated} (строк файла: {len(valid_rows)})")

    @post("/change-lcn/generate-sql/start")
    async def change_lcn_generate_sql_start(
        self,
        request: Request,
        data: dict = Body(media_type=RequestEncodingType.MULTI_PART),
    ) -> Response:
        """Start background generation of the 'изменить lcn в модели' SQL file; returns a task_id to poll."""
        rows: list[dict] = json.loads(data.get("rows", "[]"))
        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_change_lcn_generate(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task
        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_change_lcn_generate(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                errors, valid_rows = await self._validate_and_build_change_model_lcn_rows(session, rows, progress=progress)
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка валидации: {e}"}])
            return

        if errors:
            progress.update(status="error", errors=errors)
            return

        sql_lines = self._build_change_lcn_sql_lines(valid_rows)
        full_sql = "\n".join(["BEGIN;", *sql_lines, "COMMIT;"])
        progress.update(status="done", sql=full_sql, count=len(valid_rows))

    @post("/change-lcn/execute-sql/start")
    async def change_lcn_execute_sql_start(
        self,
        request: Request,
        data: dict = Body(media_type=RequestEncodingType.MULTI_PART),
    ) -> Response:
        """Start the background atomic lcn change on models; returns a task_id to poll."""
        rows: list[dict] = json.loads(data.get("rows", "[]"))
        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_change_lcn_execute(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task
        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_change_lcn_execute(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        total_updated = 0
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                try:
                    errors, valid_rows = await self._validate_and_build_change_model_lcn_rows(session, rows, progress=progress)
                except Exception as e:
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка валидации: {e}"}])
                    return

                if errors:
                    progress.update(status="error", errors=errors)
                    return

                if not valid_rows:
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": "Нет валидных строк для изменения"}])
                    return

                progress.update(processed=0, total=1, phase="executing")
                try:
                    # Two-phase UPDATE (see _build_change_lcn_sql_lines) — both steps must
                    # run in one transaction, or after the first step the models rows are
                    # left carrying the temporary 'Z' prefix in lcn.
                    sql_lines = self._build_change_lcn_sql_lines(valid_rows)
                    await session.execute(text(sql_lines[0]))
                    result = await session.execute(text(sql_lines[1]))
                    total_updated = result.rowcount
                    progress["processed"] = 1
                    await session.commit()
                except Exception as e:
                    await session.rollback()
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
                    return
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
            return

        now = datetime.now()
        log_lines = [
            f"=== Execute change-lcn update: {now.strftime('%Y-%m-%d %H:%M:%S')} ===",
            f"Rows processed: {len(valid_rows)}, models updated: {total_updated}",
            "",
            *self._build_change_lcn_sql_lines(valid_rows),
            "",
        ]
        log_file = LOG_DIR / f"change_lcn_{now.strftime('%Y-%m-%d_%H-%M-%S')}.log"
        with open(log_file, "a", encoding="utf-8") as f:
            f.write("\n".join(log_lines))
        logger.info("Changed lcn for %d models, log: %s", total_updated, log_file)

        progress.update(status="done", count=len(valid_rows),
                         message=f"Изменено lcn у моделей: {total_updated} (строк файла: {len(valid_rows)})")

    @post("/is-default/generate-sql/start")
    async def is_default_generate_sql_start(
        self,
        request: Request,
        data: dict = Body(media_type=RequestEncodingType.MULTI_PART),
    ) -> Response:
        """Start background generation of the 'Изменить серийность в модели' SQL file; returns a task_id."""
        rows: list[dict] = json.loads(data.get("rows", "[]"))
        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_is_default_generate(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task
        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_is_default_generate(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                errors, valid_rows = await self._validate_and_build_is_default_rows(session, rows, progress=progress)
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка валидации: {e}"}])
            return

        if errors:
            progress.update(status="error", errors=errors)
            return

        sql_lines = self._build_is_default_sql_lines(valid_rows)
        full_sql = "\n".join(["BEGIN;", *sql_lines, "COMMIT;"])
        progress.update(status="done", sql=full_sql, count=len(valid_rows))

    @post("/is-default/execute-sql/start")
    async def is_default_execute_sql_start(
        self,
        request: Request,
        data: dict = Body(media_type=RequestEncodingType.MULTI_PART),
    ) -> Response:
        """Start the background atomic is_default change on models; returns a task_id."""
        rows: list[dict] = json.loads(data.get("rows", "[]"))
        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_is_default_execute(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task
        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_is_default_execute(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        total_updated = 0
        valid_rows: list[dict] = []
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                try:
                    errors, valid_rows = await self._validate_and_build_is_default_rows(session, rows, progress=progress)
                except Exception as e:
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка валидации: {e}"}])
                    return

                if errors:
                    progress.update(status="error", errors=errors)
                    return

                if not valid_rows:
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": "Нет валидных строк для изменения"}])
                    return

                # FALSE rows before TRUE ones (see _build_is_default_sql_lines) — one
                # UPDATE per model, each its own transaction step, so the order stays
                # predictable for the partial UNIQUE index over is_default=true.
                sql_lines = self._build_is_default_sql_lines(valid_rows)
                progress.update(processed=0, total=len(sql_lines), phase="executing")
                try:
                    for i, line in enumerate(sql_lines, start=1):
                        result = await session.execute(text(line))
                        total_updated += result.rowcount
                        progress["processed"] = i
                    await session.commit()
                except Exception as e:
                    await session.rollback()
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
                    return
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
            return

        now = datetime.now()
        log_lines = [
            f"=== Execute is-default update: {now.strftime('%Y-%m-%d %H:%M:%S')} ===",
            f"Rows processed: {len(valid_rows)}, models updated: {total_updated}",
            "",
            *self._build_is_default_sql_lines(valid_rows),
            "",
        ]
        log_file = LOG_DIR / f"is_default_{now.strftime('%Y-%m-%d_%H-%M-%S')}.log"
        with open(log_file, "a", encoding="utf-8") as f:
            f.write("\n".join(log_lines))
        logger.info("Changed is_default for %d models, log: %s", total_updated, log_file)

        progress.update(status="done", count=len(valid_rows),
                         message=f"Изменена серийность (is_default) у моделей: {total_updated} (строк файла: {len(valid_rows)})")

    @post("/move-no-relocate/generate-sql/start")
    async def move_no_relocate_generate_sql_start(
        self,
        request: Request,
        data: dict = Body(media_type=RequestEncodingType.MULTI_PART),
    ) -> Response:
        """Start background generation of the 'Переместить активы без relocate' SQL file; returns a task_id."""
        rows: list[dict] = json.loads(data.get("rows", "[]"))
        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_move_no_relocate_generate(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task
        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_move_no_relocate_generate(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                # The same lcn pair validation and parsing as 'Изменить lcn в модели'
                # (Старый lsn -> lsn, both columns resolved as in 'set serial=none lcn').
                errors, valid_rows = await self._validate_and_build_move_no_relocate_rows(session, rows, progress=progress)
                if not errors:
                    errors.extend(await self._check_lcn_collisions(session, valid_rows))
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка валидации: {e}"}])
            return

        if errors:
            progress.update(status="error", errors=errors)
            return

        sql_lines = self._build_move_no_relocate_sql_lines(valid_rows)
        full_sql = "\n".join(["BEGIN;", *sql_lines, "COMMIT;"])
        progress.update(status="done", sql=full_sql, count=len(valid_rows))

    @post("/move-no-relocate/execute-sql/start")
    async def move_no_relocate_execute_sql_start(
        self,
        request: Request,
        data: dict = Body(media_type=RequestEncodingType.MULTI_PART),
    ) -> Response:
        """Start the background atomic asset move without relocate; returns a task_id."""
        rows: list[dict] = json.loads(data.get("rows", "[]"))
        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_move_no_relocate_execute(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task
        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_move_no_relocate_execute(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        total_updated = 0
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                try:
                    errors, valid_rows = await self._validate_and_build_move_no_relocate_rows(session, rows, progress=progress)
                    if not errors:
                        errors.extend(await self._check_lcn_collisions(session, valid_rows))
                except Exception as e:
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка валидации: {e}"}])
                    return

                if errors:
                    progress.update(status="error", errors=errors)
                    return

                if not valid_rows:
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": "Нет валидных строк для перемещения"}])
                    return

                progress.update(processed=0, total=1, phase="executing")
                try:
                    # Two-phase UPDATE as in change-lcn, but additionally resetting
                    # id_actves_parent/id_actives_root; id_location and relocate are untouched.
                    sql_lines = self._build_move_no_relocate_sql_lines(valid_rows)
                    await session.execute(text(sql_lines[0]))
                    result = await session.execute(text(sql_lines[1]))
                    total_updated = result.rowcount
                    progress["processed"] = 1
                    await session.commit()
                except Exception as e:
                    await session.rollback()
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
                    return
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
            return

        now = datetime.now()
        log_lines = [
            f"=== Execute move-no-relocate update: {now.strftime('%Y-%m-%d %H:%M:%S')} ===",
            f"Pairs processed: {len(valid_rows)}, actives updated: {total_updated}",
            "",
            *self._build_move_no_relocate_sql_lines(valid_rows),
            "",
        ]
        log_file = LOG_DIR / f"move_no_relocate_{now.strftime('%Y-%m-%d_%H-%M-%S')}.log"
        with open(log_file, "a", encoding="utf-8") as f:
            f.write("\n".join(log_lines))
        logger.info("Moved (no relocate) %d pairs (%d actives), log: %s", len(valid_rows), total_updated, log_file)

        progress.update(status="done", count=len(valid_rows),
                         message=f"Перемещено активов: {total_updated} (пар: {len(valid_rows)})")

    @post("/move-actives/generate-sql/start")
    async def move_actives_generate_sql_start(
        self,
        request: Request,
        data: dict = Body(media_type=RequestEncodingType.MULTI_PART),
    ) -> Response:
        """Start background generation of the asset move SQL file; returns a task_id to poll."""
        rows: list[dict] = json.loads(data.get("rows", "[]"))
        reason = str(data.get("reason", "") or "")
        storage_name = str(data.get("storage_name", "") or "")
        consignment_name = str(data.get("consignment_name", "") or "")
        user_fullname = str(data.get("user_fullname", "") or "")
        set_nocm = str(data.get("set_nocm", "") or "").strip().lower() in ("1", "true", "on")

        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(
            self._run_move_actives_generate(task_id, rows, reason, storage_name, consignment_name, user_fullname, set_nocm)
        )
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task
        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_move_actives_generate(
        self, task_id: str, rows: list[dict],
        reason: str, storage_name: str, consignment_name: str, user_fullname: str, set_nocm: bool,
    ) -> None:
        progress = _progress[task_id]
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                errors, valid_rows, id_storage, id_consignment, id_user, id_design_number = await self._validate_and_build_move_rows(
                    session, rows, storage_name, consignment_name, user_fullname, set_nocm, progress=progress,
                )
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка валидации: {e}"}])
            return

        if errors:
            progress.update(status="error", errors=errors)
            return
        if not valid_rows:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": "Нет валидных строк для перемещения"}])
            return

        move_date = datetime.now() - MOVE_TZ_SHIFT
        sql_lines = self._build_move_actives_sql_body(
            valid_rows, id_storage, id_consignment, id_user, reason, move_date, id_design_number,
        )
        full_sql = "\n".join(["BEGIN;", *sql_lines, "COMMIT;"])
        progress.update(status="done", sql=full_sql, count=len(valid_rows))

    @post("/move-actives/execute-sql/start")
    async def move_actives_execute_sql_start(
        self,
        request: Request,
        data: dict = Body(media_type=RequestEncodingType.MULTI_PART),
    ) -> Response:
        """Start the background atomic asset move; returns a task_id to poll."""
        rows: list[dict] = json.loads(data.get("rows", "[]"))
        reason = str(data.get("reason", "") or "")
        storage_name = str(data.get("storage_name", "") or "")
        consignment_name = str(data.get("consignment_name", "") or "")
        user_fullname = str(data.get("user_fullname", "") or "")
        set_nocm = str(data.get("set_nocm", "") or "").strip().lower() in ("1", "true", "on")

        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(
            self._run_move_actives_execute(task_id, rows, reason, storage_name, consignment_name, user_fullname, set_nocm)
        )
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task
        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_move_actives_execute(
        self, task_id: str, rows: list[dict],
        reason: str, storage_name: str, consignment_name: str, user_fullname: str, set_nocm: bool,
    ) -> None:
        progress = _progress[task_id]
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                try:
                    errors, valid_rows, id_storage, id_consignment, id_user, id_design_number = await self._validate_and_build_move_rows(
                        session, rows, storage_name, consignment_name, user_fullname, set_nocm, progress=progress,
                    )
                except Exception as e:
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка валидации: {e}"}])
                    return

                if errors:
                    progress.update(status="error", errors=errors)
                    return
                if not valid_rows:
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": "Нет валидных строк для перемещения"}])
                    return

                progress.update(processed=0, total=1, phase="executing")
                move_date = datetime.now() - MOVE_TZ_SHIFT
                sql_body = "\n".join(
                    self._build_move_actives_sql_body(
                        valid_rows, id_storage, id_consignment, id_user, reason, move_date, id_design_number,
                    )
                )
                try:
                    # DO $$ ... $$ containing several statements cannot go through an
                    # ordinary execute() (asyncpg will not prepare several commands in one
                    # prepared statement) — the same trick as in create_actives and
                    # create_named_actives: a raw connection, which session.rollback()
                    # below also rolls back, since the session opened a transaction earlier
                    # (the validation queries).
                    conn = await session.connection()
                    raw_conn = await conn.get_raw_connection()
                    await raw_conn.driver_connection.execute(sql_body)
                    progress["processed"] = 1
                    await session.commit()
                except Exception as e:
                    await session.rollback()
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
                    return
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
            return

        now = datetime.now()
        log_lines = [
            f"=== Execute move-actives: {now.strftime('%Y-%m-%d %H:%M:%S')} ===",
            f"Storage={storage_name} (id={id_storage}), Consignment={consignment_name} (id={id_consignment}), "
            f"User={user_fullname} (id={id_user})",
            f"Rows processed: {len(valid_rows)}",
            "",
            sql_body,
            "",
        ]
        log_file = LOG_DIR / f"move_actives_{now.strftime('%Y-%m-%d_%H-%M-%S')}.log"
        with open(log_file, "a", encoding="utf-8") as f:
            f.write("\n".join(log_lines))
        logger.info("Moved %d actives to storage '%s', log: %s", len(valid_rows), storage_name, log_file)

        progress.update(status="done", count=len(valid_rows),
                         message=f"Перемещено активов: {len(valid_rows)} (склад: {storage_name})")

    @get("/progress/{task_id:str}")
    async def get_progress(self, task_id: str) -> Response:
        state = _progress.get(task_id)
        if state is None:
            return Response(
                content=json.dumps({"status": "error", "errors": [{"row": 0, "field": "*", "message": "Задача не найдена или устарела"}]}),
                status_code=200,
                media_type="application/json",
            )
        return Response(
            content=json.dumps({k: v for k, v in state.items() if k != "created_at"}),
            status_code=200,
            media_type="application/json",
        )
