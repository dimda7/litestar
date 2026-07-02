import json
import logging
import tempfile
import uuid
from datetime import datetime
from pathlib import Path

import openpyxl
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession
from litestar import Controller, get, post
from litestar.connection.request import Request
from litestar.datastructures import UploadFile
from litestar.enums import RequestEncodingType
from litestar.params import Body
from litestar.response import Template, Response, Redirect

from models import DesignNumber, CounterGroup
from schemas import DesignNumberSelectSheetRequest
from sql_utils import sql_escape

PARSER_DATA_DIR = Path(__file__).parent.parent / "parser_data"
PARSER_DATA_DIR.mkdir(exist_ok=True)

LOG_DIR = Path(__file__).parent.parent / "log"
LOG_DIR.mkdir(exist_ok=True)

logger = logging.getLogger("design_number_parser")

PREFIX = "dn_parser"


def _load_data(session_id: str) -> dict | None:
    path = PARSER_DATA_DIR / f"{session_id}.json"
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    return None


def _save_data(session_id: str, data: dict) -> None:
    path = PARSER_DATA_DIR / f"{session_id}.json"
    path.write_text(json.dumps(data, ensure_ascii=False, default=str), encoding="utf-8")


def _cleanup_old_files() -> None:
    import time
    now = time.time()
    for f in PARSER_DATA_DIR.glob("*.json"):
        if now - f.stat().st_mtime > 3600:
            f.unlink(missing_ok=True)


class DesignNumberParserController(Controller):
    path = "/design-number-parser"

    @get("/")
    async def index(
        self,
        request: Request,
        page: int = 1,
        per_page: int = 10,
        select_sheet: bool = False,
    ) -> Template:
        page = max(page, 1)
        per_page = min(per_page, 200)
        error: str = request.session.pop(f"{PREFIX}_error", "")

        pending_sheets: list[str] = []
        pending_filename: str = ""
        if select_sheet:
            pending_sheets = request.session.get(f"{PREFIX}_pending_sheets", [])
            pending_filename = request.session.get(f"{PREFIX}_pending_filename", "")

        session_id = request.session.get(f"{PREFIX}_session_id", "")
        stored = _load_data(session_id) if session_id else None

        all_rows: list[dict] = stored["rows"] if stored else []
        headers: list[str] = stored["headers"] if stored else []
        filename: str = stored["filename"] if stored else ""

        total = len(all_rows)
        total_pages = max((total + per_page - 1) // per_page, 1)
        if page > total_pages:
            page = total_pages

        offset = (page - 1) * per_page
        rows = all_rows[offset:offset + per_page]

        return Template(
            template_name="design_number_parser.html",
            context={
                "headers": headers,
                "rows": rows,
                "all_rows": all_rows,
                "filename": filename,
                "error": error,
                "page": page,
                "per_page": per_page,
                "total": total,
                "total_pages": total_pages,
                "user_id": request.session.get("user_id"),
                "fullname": request.session.get("fullname", ""),
                "active_page": "design_number_parser",
                "pending_sheets": pending_sheets,
                "pending_filename": pending_filename,
            },
        )

    @post("/upload")
    async def upload(self, request: Request) -> Redirect:
        form = await request.form()
        upload_file: UploadFile | None = form.get("file")

        if not upload_file or not upload_file.filename:
            request.session[f"{PREFIX}_error"] = "Файл не выбран"
            return Redirect("/design-number-parser")

        suffix = Path(upload_file.filename).suffix.lower()
        if suffix not in (".xlsx", ".xls"):
            request.session[f"{PREFIX}_error"] = "Поддерживаются только .xlsx и .xls файлы"
            return Redirect("/design-number-parser")

        try:
            _cleanup_old_files()

            content = await upload_file.read()
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                tmp.write(content)
                tmp_path = tmp.name

            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()

            if len(sheet_names) > 1:
                request.session[f"{PREFIX}_pending_file"] = tmp_path
                request.session[f"{PREFIX}_pending_sheets"] = sheet_names
                request.session[f"{PREFIX}_pending_filename"] = upload_file.filename
                return Redirect("/design-number-parser?select_sheet=1")

            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            ws = wb.active

            rows: list[dict[str, str | None]] = []
            headers: list[str] = []

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c) if c else f"col_{i}" for i, c in enumerate(row)]
                    continue
                rows.append({headers[j]: row[j] for j in range(len(row))})

            wb.close()
            Path(tmp_path).unlink(missing_ok=True)

            session_id = uuid.uuid4().hex
            _save_data(session_id, {
                "headers": headers,
                "rows": rows,
                "filename": upload_file.filename,
            })
            request.session[f"{PREFIX}_session_id"] = session_id

            return Redirect("/design-number-parser")
        except Exception as e:
            request.session[f"{PREFIX}_error"] = f"Ошибка чтения файла: {e}"
            return Redirect("/design-number-parser")

    @post("/select-sheet")
    async def select_sheet(
        self,
        request: Request,
        data: DesignNumberSelectSheetRequest = Body(media_type=RequestEncodingType.URL_ENCODED),
    ) -> Redirect:
        sheet_name = data.sheet_name

        tmp_path = request.session.get(f"{PREFIX}_pending_file", "")
        filename = request.session.get(f"{PREFIX}_pending_filename", "")
        sheet_names = request.session.get(f"{PREFIX}_pending_sheets", [])

        if not tmp_path or not Path(tmp_path).exists():
            request.session[f"{PREFIX}_error"] = "Временный файл истёк. Загрузите файл заново."
            return Redirect("/design-number-parser")

        if sheet_name not in sheet_names:
            request.session[f"{PREFIX}_error"] = "Выбранный лист не найден в файле."
            return Redirect("/design-number-parser")

        try:
            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            ws = wb[sheet_name]

            rows: list[dict[str, str | None]] = []
            headers: list[str] = []

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c) if c else f"col_{i}" for i, c in enumerate(row)]
                    continue
                rows.append({headers[j]: row[j] for j in range(len(row))})

            wb.close()
            Path(tmp_path).unlink(missing_ok=True)

            request.session.pop(f"{PREFIX}_pending_file", None)
            request.session.pop(f"{PREFIX}_pending_sheets", None)
            request.session.pop(f"{PREFIX}_pending_filename", None)

            session_id = uuid.uuid4().hex
            _save_data(session_id, {
                "headers": headers,
                "rows": rows,
                "filename": f"{filename} [{sheet_name}]",
            })
            request.session[f"{PREFIX}_session_id"] = session_id

            return Redirect("/design-number-parser")
        except Exception as e:
            request.session[f"{PREFIX}_error"] = f"Ошибка чтения листа: {e}"
            return Redirect("/design-number-parser")

    async def _validate_counter_group(
        self, db_session: AsyncSession, rows: list[dict]
    ) -> tuple[list[dict], list[tuple[int, str, int]]]:
        """Validate rows for counter_group update.
        Returns (errors, valid_rows) where valid_rows is [(design_number_id, number, counter_group_id), ...]
        """
        errors: list[dict] = []
        valid_rows: list[tuple[int, str, int]] = []

        cg_result = await db_session.execute(select(CounterGroup.id, CounterGroup.name))
        cg_map: dict[str, int] = {}
        for cg in cg_result.all():
            if cg[1]:
                cg_map[cg[1].strip().lower()] = cg[0]

        for idx, row in enumerate(rows):
            row_num = idx + 1
            number = str(row.get("number", "") or "").strip()
            cg_name = str(row.get("counter_group", "") or "").strip()

            if not number:
                errors.append({"row": row_num, "field": "number", "message": "Поле 'number' пустое"})
                continue

            dn_result = await db_session.execute(
                select(DesignNumber.id).where(DesignNumber.number == number)
            )
            dn_id = dn_result.scalar_one_or_none()
            if dn_id is None:
                errors.append({"row": row_num, "field": "number", "message": f"design_number не найден: '{number}'"})
                continue

            if not cg_name:
                errors.append({"row": row_num, "field": "counter_group", "message": "Поле 'counter_group' пустое"})
                continue

            cg_id = cg_map.get(cg_name.lower())
            if cg_id is None:
                errors.append({"row": row_num, "field": "counter_group", "message": f"counter_group не найден: '{cg_name}'"})
                continue

            valid_rows.append((dn_id, number, cg_id))

        return errors, valid_rows

    async def _validate_is_serial_1c(
        self, db_session: AsyncSession, rows: list[dict]
    ) -> tuple[list[dict], list[tuple[int, str, bool]]]:
        """Validate rows for is_serial_1c update.
        Returns (errors, valid_rows) where valid_rows is [(design_number_id, number, is_serial_1c), ...]
        """
        errors: list[dict] = []
        valid_rows: list[tuple[int, str, bool]] = []

        for idx, row in enumerate(rows):
            row_num = idx + 1
            number = str(row.get("number", "") or "").strip()
            is_serial_1c_raw = str(row.get("is_serial_1c", "") or "").strip().lower()

            if not number:
                errors.append({"row": row_num, "field": "number", "message": "Поле 'number' пустое"})
                continue

            dn_result = await db_session.execute(
                select(DesignNumber.id).where(DesignNumber.number == number)
            )
            dn_id = dn_result.scalar_one_or_none()
            if dn_id is None:
                errors.append({"row": row_num, "field": "number", "message": f"design_number не найден: '{number}'"})
                continue

            if is_serial_1c_raw not in ("true", "false", "1", "0", "да", "нет"):
                errors.append({"row": row_num, "field": "is_serial_1c",
                               "message": f"Неверное значение is_serial_1c: '{is_serial_1c_raw}' (ожидается true/false)"})
                continue

            is_serial_1c = is_serial_1c_raw in ("true", "1", "да")
            valid_rows.append((dn_id, number, is_serial_1c))

        return errors, valid_rows

    @post("/update-counter-group")
    async def update_counter_group(
        self,
        request: Request,
        db_session: AsyncSession,
        data: dict = Body(media_type=RequestEncodingType.MULTI_PART),
    ) -> Response:
        rows: list[dict] = json.loads(data.get("rows", "[]"))
        errors, valid_rows = await self._validate_counter_group(db_session, rows)

        if errors:
            return Response(
                content=json.dumps({"status": "error", "errors": errors}),
                status_code=200,
                media_type="application/json",
            )

        if not valid_rows:
            return Response(
                content=json.dumps({"status": "error", "errors": [{"row": 0, "field": "*", "message": "Нет валидных строк для обновления"}]}),
                status_code=200,
                media_type="application/json",
            )

        try:
            for dn_id, number, cg_id in valid_rows:
                await db_session.execute(
                    text("UPDATE public.design_number SET id_counter_group = :cg_id WHERE number = :number"),
                    {"cg_id": cg_id, "number": number},
                )
            await db_session.commit()
        except Exception as e:
            await db_session.rollback()
            return Response(
                content=json.dumps({"status": "error", "errors": [{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}]}),
                status_code=200,
                media_type="application/json",
            )

        now = datetime.now()
        log_lines = [
            f"=== Update id_counter_group: {now.strftime('%Y-%m-%d %H:%M:%S')} ===",
            f"Rows updated: {len(valid_rows)}",
            "",
        ]
        for dn_id, number, cg_id in valid_rows:
            log_lines.append(f"UPDATE design_number SET id_counter_group = {cg_id} WHERE number = '{sql_escape(number)}';")
        log_lines.append("")

        log_file = LOG_DIR / f"update_counter_group_{now.strftime('%Y-%m-%d_%H-%M-%S')}.log"
        with open(log_file, "a", encoding="utf-8") as f:
            f.write("\n".join(log_lines))
        logger.info("Updated counter_group for %d rows, log: %s", len(valid_rows), log_file)

        return Response(
            content=json.dumps({"status": "ok", "count": len(valid_rows), "message": f"Успешно обновлено id_counter_group для {len(valid_rows)} записей"}),
            status_code=200,
            media_type="application/json",
        )

    @post("/sync-counter-active")
    async def sync_counter_active(
        self,
        request: Request,
        db_session: AsyncSession,
        data: dict = Body(media_type=RequestEncodingType.MULTI_PART),
    ) -> Response:
        rows: list[dict] = json.loads(data.get("rows", "[]"))
        errors, valid_rows = await self._validate_counter_group(db_session, rows)

        if errors:
            return Response(
                content=json.dumps({"status": "error", "errors": errors}),
                status_code=200,
                media_type="application/json",
            )

        if not valid_rows:
            return Response(
                content=json.dumps({"status": "error", "errors": [{"row": 0, "field": "*", "message": "Нет валидных строк для синхронизации"}]}),
                status_code=200,
                media_type="application/json",
            )

        deleted_total = 0
        inserted_total = 0
        sync_details: list[str] = []

        try:
            for dn_id, number, cg_id in valid_rows:
                ct_result = await db_session.execute(
                    text(
                        "SELECT ctg.id_counter_type "
                        "FROM public.counter_type_to_group ctg "
                        "WHERE ctg.id_counter_group = :cg_id"
                    ),
                    {"cg_id": cg_id},
                )
                new_types = {row[0] for row in ct_result.all()}

                active_result = await db_session.execute(
                    text("SELECT id FROM public.actives WHERE id_design_number = :dn_id"),
                    {"dn_id": dn_id},
                )
                active_ids = [row[0] for row in active_result.all()]

                if not active_ids:
                    continue

                deleted_dn = 0
                inserted_dn = 0

                for aid in active_ids:
                    existing_result = await db_session.execute(
                        text(
                            "SELECT id, id_counter_type FROM public.counter_active "
                            "WHERE id_active = :aid"
                        ),
                        {"aid": aid},
                    )
                    existing_rows = existing_result.all()

                    for ca_id, ca_type in existing_rows:
                        if ca_type not in new_types:
                            await db_session.execute(
                                text("DELETE FROM public.counter_active WHERE id = :id"),
                                {"id": ca_id},
                            )
                            deleted_dn += 1

                    existing_types = {row[1] for row in existing_rows}
                    missing_types = new_types - existing_types

                    for ct_id in missing_types:
                        if ct_id == 2:
                            continue
                        if ct_id == 1:
                            freq = 7
                        elif ct_id == 3:
                            freq = 5
                        else:
                            freq = 8
                        now_ts = datetime.now()
                        await db_session.execute(
                            text(
                                "INSERT INTO public.counter_active "
                                "(id_active, date, value, id_counter_type, id_frequency_type, is_train, date_create) "
                                "VALUES (:aid, :dt, 0, :ct_id, :freq, false, :dt)"
                            ),
                            {"aid": aid, "dt": now_ts, "ct_id": ct_id, "freq": freq},
                        )
                        inserted_dn += 1

                deleted_total += deleted_dn
                inserted_total += inserted_dn
                sync_details.append(
                    f"design_number id={dn_id} number='{number}' counter_group={cg_id}: "
                    f"типы={sorted(new_types)}, actives={len(active_ids)}, "
                    f"удалено={deleted_dn}, добавлено={inserted_dn}"
                )

            await db_session.commit()
        except Exception as e:
            await db_session.rollback()
            return Response(
                content=json.dumps({"status": "error", "errors": [{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}]}),
                status_code=200,
                media_type="application/json",
            )

        now = datetime.now()
        log_lines = [
            f"=== Sync counter_active: {now.strftime('%Y-%m-%d %H:%M:%S')} ===",
            f"Rows processed: {len(valid_rows)}",
            f"Counter_active удалено: {deleted_total}, добавлено: {inserted_total}",
            "",
        ]
        for dn_id, number, cg_id in valid_rows:
            log_lines.append(f"design_number id={dn_id} number='{number}' counter_group={cg_id}")
        if sync_details:
            log_lines.append("")
            log_lines.append("Детали синхронизации:")
            for detail in sync_details:
                log_lines.append(f"  {detail}")
        log_lines.append("")

        log_file = LOG_DIR / f"sync_counter_active_{now.strftime('%Y-%m-%d_%H-%M-%S')}.log"
        with open(log_file, "a", encoding="utf-8") as f:
            f.write("\n".join(log_lines))
        logger.info("Synced counter_active for %d rows, deleted %d, inserted %d, log: %s", len(valid_rows), deleted_total, inserted_total, log_file)

        return Response(
            content=json.dumps({
                "status": "ok",
                "count": len(valid_rows),
                "deleted_counter_active": deleted_total,
                "inserted_counter_active": inserted_total,
                "message": f"Синхронизация counter_active: {len(valid_rows)} design_number, удалено {deleted_total}, добавлено {inserted_total}",
            }),
            status_code=200,
            media_type="application/json",
        )

    @post("/generate-sql-counter-active")
    async def generate_sql_counter_active(
        self,
        request: Request,
        db_session: AsyncSession,
        data: dict = Body(media_type=RequestEncodingType.MULTI_PART),
    ) -> Response:
        rows: list[dict] = json.loads(data.get("rows", "[]"))
        errors, valid_rows = await self._validate_counter_group(db_session, rows)

        if errors:
            return Response(
                content=json.dumps({"status": "error", "errors": errors}),
                status_code=200,
                media_type="application/json",
            )

        sql_lines: list[str] = []

        for dn_id, number, cg_id in valid_rows:
            ct_result = await db_session.execute(
                text(
                    "SELECT ctg.id_counter_type "
                    "FROM public.counter_type_to_group ctg "
                    "WHERE ctg.id_counter_group = :cg_id"
                ),
                {"cg_id": cg_id},
            )
            new_types = {row[0] for row in ct_result.all()}

            active_result = await db_session.execute(
                text("SELECT id FROM public.actives WHERE id_design_number = :dn_id"),
                {"dn_id": dn_id},
            )
            active_ids = [row[0] for row in active_result.all()]

            if not active_ids:
                continue

            for aid in active_ids:
                existing_result = await db_session.execute(
                    text(
                        "SELECT id, id_counter_type FROM public.counter_active "
                        "WHERE id_active = :aid"
                    ),
                    {"aid": aid},
                )
                existing_rows = existing_result.all()

                for ca_id, ca_type in existing_rows:
                    if ca_type not in new_types:
                        sql_lines.append(f"DELETE FROM public.counter_active WHERE id = {ca_id};")

                existing_types = {row[1] for row in existing_rows}
                missing_types = new_types - existing_types

                for ct_id in missing_types:
                    if ct_id == 2:
                        continue
                    if ct_id == 1:
                        freq = 7
                    elif ct_id == 3:
                        freq = 5
                    else:
                        freq = 8
                    sql_lines.append(
                        f"INSERT INTO public.counter_active "
                        f"(id_active, date, value, id_counter_type, id_frequency_type, is_train, date_create) "
                        f"VALUES ({aid}, NOW(), 0, {ct_id}, {freq}, false, NOW());"
                    )

        content = "\n".join(sql_lines)
        return Response(
            content=json.dumps({"status": "ok", "sql": content, "count": len(sql_lines)}),
            status_code=200,
            media_type="application/json",
        )

    @post("/generate-sql-counter-group")
    async def generate_sql_counter_group(
        self,
        request: Request,
        db_session: AsyncSession,
        data: dict = Body(media_type=RequestEncodingType.MULTI_PART),
    ) -> Response:
        rows: list[dict] = json.loads(data.get("rows", "[]"))
        errors, valid_rows = await self._validate_counter_group(db_session, rows)

        if errors:
            return Response(
                content=json.dumps({"status": "error", "errors": errors}),
                status_code=200,
                media_type="application/json",
            )

        sql_lines = [
            f"UPDATE design_number SET id_counter_group = {cg_id} WHERE number = '{sql_escape(number)}';"
            for _, number, cg_id in valid_rows
        ]
        content = "\n".join(sql_lines)
        return Response(
            content=json.dumps({"status": "ok", "sql": content, "count": len(sql_lines)}),
            status_code=200,
            media_type="application/json",
        )

    @post("/update-is-serial-1c")
    async def update_is_serial_1c(
        self,
        request: Request,
        db_session: AsyncSession,
        data: dict = Body(media_type=RequestEncodingType.MULTI_PART),
    ) -> Response:
        rows: list[dict] = json.loads(data.get("rows", "[]"))
        errors, valid_rows = await self._validate_is_serial_1c(db_session, rows)

        if errors:
            return Response(
                content=json.dumps({"status": "error", "errors": errors}),
                status_code=200,
                media_type="application/json",
            )

        if not valid_rows:
            return Response(
                content=json.dumps({"status": "error", "errors": [{"row": 0, "field": "*", "message": "Нет валидных строк для обновления"}]}),
                status_code=200,
                media_type="application/json",
            )

        try:
            for dn_id, number, is_serial_1c in valid_rows:
                await db_session.execute(
                    text("UPDATE public.design_number SET is_serial_1c = :val WHERE number = :number"),
                    {"val": is_serial_1c, "number": number},
                )
            await db_session.commit()
        except Exception as e:
            await db_session.rollback()
            return Response(
                content=json.dumps({"status": "error", "errors": [{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}]}),
                status_code=200,
                media_type="application/json",
            )

        now = datetime.now()
        log_lines = [
            f"=== Update is_serial_1c: {now.strftime('%Y-%m-%d %H:%M:%S')} ===",
            f"Rows updated: {len(valid_rows)}",
            "",
        ]
        for dn_id, number, is_serial_1c in valid_rows:
            log_lines.append(f"UPDATE design_number SET is_serial_1c = {'TRUE' if is_serial_1c else 'FALSE'} WHERE number = '{sql_escape(number)}';")
        log_lines.append("")

        log_file = LOG_DIR / f"update_is_serial_1c_{now.strftime('%Y-%m-%d_%H-%M-%S')}.log"
        with open(log_file, "a", encoding="utf-8") as f:
            f.write("\n".join(log_lines))
        logger.info("Updated is_serial_1c for %d rows, log: %s", len(valid_rows), log_file)

        return Response(
            content=json.dumps({"status": "ok", "count": len(valid_rows), "message": f"Успешно обновлено is_serial_1c для {len(valid_rows)} записей"}),
            status_code=200,
            media_type="application/json",
        )

    @post("/generate-sql-is-serial-1c")
    async def generate_sql_is_serial_1c(
        self,
        request: Request,
        db_session: AsyncSession,
        data: dict = Body(media_type=RequestEncodingType.MULTI_PART),
    ) -> Response:
        rows: list[dict] = json.loads(data.get("rows", "[]"))
        errors, valid_rows = await self._validate_is_serial_1c(db_session, rows)

        if errors:
            return Response(
                content=json.dumps({"status": "error", "errors": errors}),
                status_code=200,
                media_type="application/json",
            )

        sql_lines = [
            f"UPDATE design_number SET is_serial_1c = {'TRUE' if is_serial_1c else 'FALSE'} WHERE number = '{sql_escape(number)}';"
            for _, number, is_serial_1c in valid_rows
        ]
        content = "\n".join(sql_lines)
        return Response(
            content=json.dumps({"status": "ok", "sql": content, "count": len(sql_lines)}),
            status_code=200,
            media_type="application/json",
        )
