import asyncio
import base64
import json
import logging
import re
import tempfile
import time
import uuid
from io import BytesIO
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from advanced_alchemy.repository import SQLAlchemyAsyncRepository
from sqlalchemy import bindparam, func, or_, select, text
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import aliased
from litestar import Controller, get, post
from litestar.connection.request import Request
from litestar.datastructures import UploadFile
from litestar.enums import RequestEncodingType
from litestar.params import Body
from litestar.response import Template, Response, Redirect

from db_manager import get_session_maker
from models import (
    Actives, ActiveAdditionalField, ActivesToMainPtoir, CounterActive, Consignment,
    DesignNumber, IteratorNumberLast, Location, MaterialsToActives,
    MileageHistoryActives, MileageStart, MileageTrain, Orders, OrderToActives,
    Ptoir, Relocate, Storage, StoragePlace, Train,
)
from schemas import SelectSheetRequest
from sql_utils import sql_escape
from parser_storage import (
    LOG_DIR,
    load_data as _load_data,
    save_data as _save_data,
    cleanup_old_files as _cleanup_old_files,
)

logger = logging.getLogger("actives_parser")

PREFIX = "actives_parser"

ACTIVE_NUMBER_COUNTER_DESCRIPTION = "Номер следующего актива"
# active_number must be exactly this long: a letter prefix (asset type) plus the
# counter digits, zero-padded on the left between prefix and number to that length.
ACTIVE_NUMBER_LENGTH = 10

# Date of the historical mileage accounting migration: asset moves before it feed
# the recomputed mileage_start.milage, and train mileage is summed up to it inclusive.
MILEAGE_RECOUNT_CUTOFF = date(2022, 5, 13)
# relocate.date is stored without a timezone (UTC); for comparison against the
# cutoff it is converted to Moscow time, as in the old peewee script.
MILEAGE_TZ_SHIFT = timedelta(hours=3)
# counter_type of the mileage counter in counter_active
MILEAGE_COUNTER_TYPE_ID = 3

# Tables referencing actives by FK: a single related record blocks deleting the
# asset (a strict DELETE — only assets without history are removed).
# counter_active and mileage_start are deliberately absent: the counter_active row
# is created by the actives_trgger trigger on every asset INSERT (every asset has a
# counter, so blocking on it would forbid deletion outright), which is why both
# tables are deleted along with the asset. ptoir goes with the asset too (together
# with its ptoir_level_warning), as do the asset's "empty" orders — but an order
# with related records in any of ORDERS_DEPENDENCY_CHECKS blocks the deletion
# (checked separately).
DELETE_ACTIVES_BLOCKERS: list[tuple[str, object]] = [
    ("relocate", Relocate.id_active),
    ("relocate (root)", Relocate.id_root_active),
    ("order_to_actives", OrderToActives.id_active),
    ("active_additional_field", ActiveAdditionalField.id_active),
    ("actives_to_main_ptoir", ActivesToMainPtoir.id_actives),
    ("materials_to_actives", MaterialsToActives.id_actives),
    ("mileage_history_actives", MileageHistoryActives.id_actives),
]

# Tables referencing orders by FK (taken from information_schema of the real DB):
# an asset's order is deleted along with it only when all of these are empty —
# otherwise the order carries real working data and the asset is blocked.
ORDERS_DEPENDENCY_CHECKS: list[tuple[str, str]] = [
    ("consumption_rate_order", "id_order"),
    ("counter_order", "id_order"),
    ("labor_costs", "id_order"),
    ("material_1c", "id_order"),
    ("order_diagnostic", "id"),
    ("order_status_bin", "id"),
    ("order_to_actives", "id_order"),
    ("order_to_attachment", "id_order"),
    ("order_to_order", "id_parent"),
    ("order_to_order", "id_child"),
    ("order_to_order_executor", "id_order"),
    ("order_used_tools", "id_order"),
    ("orders_ref_data", "id_order"),
    ("orders_required_tools", "id_order"),
    ("orders_to_classifier", "id_order"),
    ("orders_to_labor_costs", "id_order"),
    ("orders_to_orders_work_operation", "id_orders"),
    ("orders_to_specification", "id_order"),
    ("orders_work_operation", "id_order"),
    ("relocate", "id_order"),
]

# A model lcn like 'M9.1.6.4': the digits after the letter prefix and before the
# first dot are id_train_type; the rest of the path is carried over as is.
# The same pattern as in parser.py (_parse_model_lcn) — not reused from there
# directly, since the project's controllers are self-contained (see other modules).
_MODEL_LCN_RE = re.compile(r"^\D*(\d+)(?:\.(.*))?$")


def _parse_model_lcn(lcn: str) -> tuple[int, str] | None:
    """Extract (id_train_type, rest_of_path) from an lcn like 'M9.6.5' -> (9, '6.5'); 'M9' -> (9, '')."""
    match = _MODEL_LCN_RE.match(lcn)
    if not match:
        return None
    return int(match.group(1)), match.group(2) or ""


# As in train_parser: validating material rows issues several database queries per
# row, and on large files that takes seconds — progress is served through a separate
# poll rather than holding one HTTP request open all that time.
PROGRESS_TTL_SECONDS = 15 * 60
_progress: dict[str, dict] = {}
_tasks: dict[str, asyncio.Task] = {}


def _cleanup_progress() -> None:
    cutoff = time.time() - PROGRESS_TTL_SECONDS
    stale = [tid for tid, state in _progress.items() if state["created_at"] < cutoff]
    for tid in stale:
        _progress.pop(tid, None)


class StorageRepository(SQLAlchemyAsyncRepository[Storage]):
    model_type = Storage


class StoragePlaceRepository(SQLAlchemyAsyncRepository[StoragePlace]):
    model_type = StoragePlace


class ConsignmentRepository(SQLAlchemyAsyncRepository[Consignment]):
    model_type = Consignment


class DesignNumberRepository(SQLAlchemyAsyncRepository[DesignNumber]):
    model_type = DesignNumber


class IteratorNumberLastRepository(SQLAlchemyAsyncRepository[IteratorNumberLast]):
    model_type = IteratorNumberLast


class ActivesRepository(SQLAlchemyAsyncRepository[Actives]):
    model_type = Actives


class MileageStartRepository(SQLAlchemyAsyncRepository[MileageStart]):
    model_type = MileageStart


class CounterActiveRepository(SQLAlchemyAsyncRepository[CounterActive]):
    model_type = CounterActive


class ActivesParserController(Controller):
    path = "/actives-parser"

    @get("/")
    async def index(
        self,
        request: Request,
        page: int = 1,
        per_page: int = 10,
        select_sheet: bool = False,
    ) -> Template:
        page = max(page, 1)
        per_page = min(per_page, 200)
        error: str = request.session.pop(f"{PREFIX}_error", "")

        pending_sheets: list[str] = []
        pending_filename: str = ""
        if select_sheet:
            pending_sheets = request.session.get(f"{PREFIX}_pending_sheets", [])
            pending_filename = request.session.get(f"{PREFIX}_pending_filename", "")

        session_id = request.session.get(f"{PREFIX}_session_id", "")
        stored = _load_data(session_id) if session_id else None

        all_rows: list[dict] = stored["rows"] if stored else []
        headers: list[str] = stored["headers"] if stored else []
        filename: str = stored["filename"] if stored else ""

        total = len(all_rows)
        total_pages = max((total + per_page - 1) // per_page, 1)
        if page > total_pages:
            page = total_pages

        offset = (page - 1) * per_page
        rows = all_rows[offset:offset + per_page]

        return Template(
            template_name="actives_parser.html",
            context={
                "headers": headers,
                "rows": rows,
                "all_rows": all_rows,
                "filename": filename,
                "error": error,
                "page": page,
                "per_page": per_page,
                "total": total,
                "total_pages": total_pages,
                "user_id": request.session.get("user_id"),
                "fullname": request.session.get("fullname", ""),
                "active_page": "actives_parser",
                "pending_sheets": pending_sheets,
                "pending_filename": pending_filename,
            },
        )

    @post("/upload")
    async def upload(self, request: Request) -> Redirect:
        form = await request.form()
        upload_file: UploadFile | None = form.get("file")

        if not upload_file or not upload_file.filename:
            request.session[f"{PREFIX}_error"] = "Файл не выбран"
            return Redirect("/actives-parser")

        suffix = Path(upload_file.filename).suffix.lower()
        if suffix not in (".xlsx", ".xls"):
            request.session[f"{PREFIX}_error"] = "Поддерживаются только .xlsx и .xls файлы"
            return Redirect("/actives-parser")

        try:
            _cleanup_old_files()

            content = await upload_file.read()
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                tmp.write(content)
                tmp_path = tmp.name

            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()

            if len(sheet_names) > 1:
                request.session[f"{PREFIX}_pending_file"] = tmp_path
                request.session[f"{PREFIX}_pending_sheets"] = sheet_names
                request.session[f"{PREFIX}_pending_filename"] = upload_file.filename
                return Redirect("/actives-parser?select_sheet=1")

            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            ws = wb.active

            rows: list[dict[str, str | None]] = []
            headers: list[str] = []

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c) if c else f"col_{i}" for i, c in enumerate(row)]
                    continue
                if all(c is None or str(c).strip() == "" for c in row):
                    continue
                rows.append({headers[j]: row[j] for j in range(len(row))})

            wb.close()
            Path(tmp_path).unlink(missing_ok=True)

            session_id = uuid.uuid4().hex
            _save_data(session_id, {
                "headers": headers,
                "rows": rows,
                "filename": upload_file.filename,
            })
            request.session[f"{PREFIX}_session_id"] = session_id

            return Redirect("/actives-parser")
        except Exception as e:
            request.session[f"{PREFIX}_error"] = f"Ошибка чтения файла: {e}"
            return Redirect("/actives-parser")

    @post("/select-sheet")
    async def select_sheet(
        self,
        request: Request,
        data: SelectSheetRequest = Body(media_type=RequestEncodingType.URL_ENCODED),
    ) -> Redirect:
        sheet_name = data.sheet_name

        tmp_path = request.session.get(f"{PREFIX}_pending_file", "")
        filename = request.session.get(f"{PREFIX}_pending_filename", "")
        sheet_names = request.session.get(f"{PREFIX}_pending_sheets", [])

        if not tmp_path or not Path(tmp_path).exists():
            request.session[f"{PREFIX}_error"] = "Временный файл истёк. Загрузите файл заново."
            return Redirect("/actives-parser")

        if sheet_name not in sheet_names:
            request.session[f"{PREFIX}_error"] = "Выбранный лист не найден в файле."
            return Redirect("/actives-parser")

        try:
            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            ws = wb[sheet_name]

            rows: list[dict[str, str | None]] = []
            headers: list[str] = []

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c) if c else f"col_{i}" for i, c in enumerate(row)]
                    continue
                if all(c is None or str(c).strip() == "" for c in row):
                    continue
                rows.append({headers[j]: row[j] for j in range(len(row))})

            wb.close()
            Path(tmp_path).unlink(missing_ok=True)

            request.session.pop(f"{PREFIX}_pending_file", None)
            request.session.pop(f"{PREFIX}_pending_sheets", None)
            request.session.pop(f"{PREFIX}_pending_filename", None)

            session_id = uuid.uuid4().hex
            _save_data(session_id, {
                "headers": headers,
                "rows": rows,
                "filename": f"{filename} [{sheet_name}]",
            })
            request.session[f"{PREFIX}_session_id"] = session_id

            return Redirect("/actives-parser")
        except Exception as e:
            request.session[f"{PREFIX}_error"] = f"Ошибка чтения листа: {e}"
            return Redirect("/actives-parser")

    async def _validate_serial_number(
        self, db_session: AsyncSession, rows: list[dict], progress: dict | None = None,
    ) -> tuple[list[dict], list[tuple[str, str]]]:
        """Validate rows for actives.serial_number update.
        Returns (errors, valid_rows) where valid_rows is [(active_number, serial_number), ...]
        """
        errors: list[dict] = []
        valid_rows: list[tuple[str, str]] = []
        batch_numbers: set[str] = set()

        if rows and "Новый с/н" not in rows[0] and "Новый Серийный номер" not in rows[0]:
            errors.append({
                "row": 0,
                "field": "Новый с/н",
                "message": "В файле не найдена колонка 'Новый с/н' (или 'Новый Серийный номер')",
            })
            return errors, valid_rows

        if progress is not None:
            progress.update(processed=0, total=len(rows), phase="validating")

        for idx, row in enumerate(rows):
            row_num = idx + 1
            if progress is not None and (idx % 20 == 0 or row_num == len(rows)):
                progress["processed"] = row_num
            active_number = str(row.get("Актив", "") or "").strip()
            serial_number = str(row.get("Новый с/н") or row.get("Новый Серийный номер") or "").strip()

            if not active_number:
                errors.append({"row": row_num, "field": "Актив", "message": "Поле 'Актив' пустое"})
                continue

            if active_number in batch_numbers:
                errors.append({"row": row_num, "field": "Актив",
                                "message": f"Дубликат внутри файла: '{active_number}'"})
                continue

            result = await db_session.execute(
                select(Actives.id).where(Actives.active_number == active_number)
            )
            active_id = result.scalar_one_or_none()
            if active_id is None:
                errors.append({"row": row_num, "field": "Актив",
                                "message": f"Актив не найден: '{active_number}'"})
                continue

            if not serial_number:
                errors.append({"row": row_num, "field": "Новый с/н", "message": "Поле 'Новый с/н' пустое"})
                continue

            batch_numbers.add(active_number)
            valid_rows.append((active_number, serial_number))

        return errors, valid_rows

    async def _validate_design_number(
        self, db_session: AsyncSession, rows: list[dict], progress: dict | None = None,
    ) -> tuple[list[dict], list[tuple[str, int, str]]]:
        """Validate rows for actives.id_design_number update.
        Returns (errors, valid_rows) where valid_rows is [(active_number, design_number_id, design_number), ...]
        """
        errors: list[dict] = []
        valid_rows: list[tuple[str, int, str]] = []
        batch_numbers: set[str] = set()

        if rows and not ({"Новая Позиция ТМЦ", "Новый ТМЦ номер", "Позиция ТМЦ"} & rows[0].keys()):
            errors.append({
                "row": 0,
                "field": "Новая Позиция ТМЦ",
                "message": "В файле не найдена колонка 'Новая Позиция ТМЦ' (или 'Новый ТМЦ номер', 'Позиция ТМЦ')",
            })
            return errors, valid_rows

        if progress is not None:
            progress.update(processed=0, total=len(rows), phase="validating")

        for idx, row in enumerate(rows):
            row_num = idx + 1
            if progress is not None and (idx % 20 == 0 or row_num == len(rows)):
                progress["processed"] = row_num
            active_number = str(row.get("Актив", "") or "").strip()
            design_number = str(
                row.get("Новая Позиция ТМЦ") or row.get("Новый ТМЦ номер") or row.get("Позиция ТМЦ") or ""
            ).strip()

            if not active_number:
                errors.append({"row": row_num, "field": "Актив", "message": "Поле 'Актив' пустое"})
                continue

            if active_number in batch_numbers:
                errors.append({"row": row_num, "field": "Актив",
                                "message": f"Дубликат внутри файла: '{active_number}'"})
                continue

            result = await db_session.execute(
                select(Actives.id).where(Actives.active_number == active_number)
            )
            active_id = result.scalar_one_or_none()
            if active_id is None:
                errors.append({"row": row_num, "field": "Актив",
                                "message": f"Актив не найден: '{active_number}'"})
                continue

            if not design_number:
                errors.append({"row": row_num, "field": "Новая Позиция ТМЦ", "message": "Поле 'Новая Позиция ТМЦ' ('Новый ТМЦ номер') пустое"})
                continue

            result = await db_session.execute(
                select(DesignNumber.id).where(DesignNumber.number == design_number)
            )
            design_number_id = result.scalar_one_or_none()
            if design_number_id is None:
                errors.append({"row": row_num, "field": "Новая Позиция ТМЦ",
                                "message": f"Позиция ТМЦ не найдена: '{design_number}'"})
                continue

            batch_numbers.add(active_number)
            valid_rows.append((active_number, design_number_id, design_number))

        return errors, valid_rows

    @post("/design-number/generate-sql/start")
    async def design_number_generate_sql_start(self, request: Request) -> Response:
        """Start background generation of the id_design_number update SQL file; returns a task_id."""
        session_id = request.session.get(f"{PREFIX}_session_id", "")
        stored = _load_data(session_id) if session_id else None
        if not stored:
            return Response(
                content=json.dumps({"status": "error", "errors": [{"row": 0, "field": "*", "message": "Данные не загружены"}]}),
                status_code=200,
                media_type="application/json",
            )

        rows: list[dict] = stored["rows"]

        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_design_number_generate(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task

        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_design_number_generate(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                errors, valid_rows = await self._validate_design_number(session, rows, progress=progress)
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка: {e}"}])
            return

        if errors:
            progress.update(status="error", errors=errors)
            return
        if not valid_rows:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": "Нет валидных строк для обновления"}])
            return

        sql_lines = [
            f"UPDATE public.actives SET id_design_number = {design_number_id} "
            f"WHERE active_number = '{sql_escape(active_number)}';"
            for active_number, design_number_id, _ in valid_rows
        ]
        progress.update(status="done", sql="\n".join(sql_lines), count=len(sql_lines))

    @post("/design-number/execute/start")
    async def design_number_execute_start(self, request: Request) -> Response:
        """Start the background id_design_number update in the database; returns a task_id."""
        session_id = request.session.get(f"{PREFIX}_session_id", "")
        stored = _load_data(session_id) if session_id else None
        if not stored:
            return Response(
                content=json.dumps({"status": "error", "errors": [{"row": 0, "field": "*", "message": "Данные не загружены"}]}),
                status_code=200,
                media_type="application/json",
            )

        rows: list[dict] = stored["rows"]

        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_design_number_execute(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task

        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_design_number_execute(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        valid_rows: list[tuple[str, int, str]] = []
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                errors, valid_rows = await self._validate_design_number(session, rows, progress=progress)

                if errors:
                    progress.update(status="error", errors=errors)
                    return
                if not valid_rows:
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": "Нет валидных строк для обновления"}])
                    return

                progress.update(processed=0, total=len(valid_rows), phase="executing")

                try:
                    for i, (active_number, design_number_id, _) in enumerate(valid_rows, start=1):
                        await session.execute(
                            text("UPDATE public.actives SET id_design_number = :dn_id WHERE active_number = :an"),
                            {"dn_id": design_number_id, "an": active_number},
                        )
                        if i % 20 == 0 or i == len(valid_rows):
                            progress["processed"] = i
                    await session.commit()
                except Exception as e:
                    await session.rollback()
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
                    return
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
            return

        now = datetime.now()
        log_lines = [
            f"=== Update id_design_number: {now.strftime('%Y-%m-%d %H:%M:%S')} ===",
            f"Rows updated: {len(valid_rows)}",
            "",
        ]
        for active_number, design_number_id, design_number in valid_rows:
            log_lines.append(
                f"UPDATE public.actives SET id_design_number = {design_number_id} "
                f"WHERE active_number = '{sql_escape(active_number)}'; -- '{sql_escape(design_number)}'"
            )
        log_lines.append("")

        log_file = LOG_DIR / f"update_design_number_{now.strftime('%Y-%m-%d_%H-%M-%S')}.log"
        with open(log_file, "a", encoding="utf-8") as f:
            f.write("\n".join(log_lines))
        logger.info("Updated id_design_number for %d rows, log: %s", len(valid_rows), log_file)

        progress.update(status="done", count=len(valid_rows),
                         message=f"Успешно обновлено id_design_number для {len(valid_rows)} записей")

    @post("/serial-number/generate-sql/start")
    async def serial_number_generate_sql_start(self, request: Request) -> Response:
        """Start background generation of the serial_number update SQL file; returns a task_id."""
        session_id = request.session.get(f"{PREFIX}_session_id", "")
        stored = _load_data(session_id) if session_id else None
        if not stored:
            return Response(
                content=json.dumps({"status": "error", "errors": [{"row": 0, "field": "*", "message": "Данные не загружены"}]}),
                status_code=200,
                media_type="application/json",
            )

        rows: list[dict] = stored["rows"]

        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_serial_number_generate(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task

        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_serial_number_generate(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                errors, valid_rows = await self._validate_serial_number(session, rows, progress=progress)
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка: {e}"}])
            return

        if errors:
            progress.update(status="error", errors=errors)
            return
        if not valid_rows:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": "Нет валидных строк для обновления"}])
            return

        sql_lines = [
            f"UPDATE public.actives SET serial_number = '{sql_escape(serial_number)}' "
            f"WHERE active_number = '{sql_escape(active_number)}';"
            for active_number, serial_number in valid_rows
        ]
        progress.update(status="done", sql="\n".join(sql_lines), count=len(sql_lines))

    @post("/serial-number/execute/start")
    async def serial_number_execute_start(self, request: Request) -> Response:
        """Start the background serial_number update in the database; returns a task_id."""
        session_id = request.session.get(f"{PREFIX}_session_id", "")
        stored = _load_data(session_id) if session_id else None
        if not stored:
            return Response(
                content=json.dumps({"status": "error", "errors": [{"row": 0, "field": "*", "message": "Данные не загружены"}]}),
                status_code=200,
                media_type="application/json",
            )

        rows: list[dict] = stored["rows"]

        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_serial_number_execute(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task

        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_serial_number_execute(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        valid_rows: list[tuple[str, str]] = []
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                errors, valid_rows = await self._validate_serial_number(session, rows, progress=progress)

                if errors:
                    progress.update(status="error", errors=errors)
                    return
                if not valid_rows:
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": "Нет валидных строк для обновления"}])
                    return

                progress.update(processed=0, total=len(valid_rows), phase="executing")

                try:
                    for i, (active_number, serial_number) in enumerate(valid_rows, start=1):
                        await session.execute(
                            text("UPDATE public.actives SET serial_number = :sn WHERE active_number = :an"),
                            {"sn": serial_number, "an": active_number},
                        )
                        if i % 20 == 0 or i == len(valid_rows):
                            progress["processed"] = i
                    await session.commit()
                except Exception as e:
                    await session.rollback()
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
                    return
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
            return

        now = datetime.now()
        log_lines = [
            f"=== Update serial_number: {now.strftime('%Y-%m-%d %H:%M:%S')} ===",
            f"Rows updated: {len(valid_rows)}",
            "",
        ]
        for active_number, serial_number in valid_rows:
            log_lines.append(
                f"UPDATE public.actives SET serial_number = '{sql_escape(serial_number)}' "
                f"WHERE active_number = '{sql_escape(active_number)}';"
            )
        log_lines.append("")

        log_file = LOG_DIR / f"update_serial_number_{now.strftime('%Y-%m-%d_%H-%M-%S')}.log"
        with open(log_file, "a", encoding="utf-8") as f:
            f.write("\n".join(log_lines))
        logger.info("Updated serial_number for %d rows, log: %s", len(valid_rows), log_file)

        progress.update(status="done", count=len(valid_rows),
                         message=f"Успешно обновлено serial_number для {len(valid_rows)} записей")

    async def _validate_recount_mileage(
        self, db_session: AsyncSession, rows: list[dict], progress: dict | None = None,
    ) -> tuple[list[dict], list[dict]]:
        """Validate the rows of a "пересчитать пробег" file (replaces update_milage_start + recount_counter).

        For each asset it computes a total correction to mileage_start.milage from the
        relocate history: moves before MILEAGE_RECOUNT_CUTOFF from storage onto a train
        add that train's mileage (mileage_train.mileage_average over the period from the
        move to the cutoff), moves from a train into storage subtract it. The
        milage_const and counter_active.value values themselves are not computed here —
        they are read and computed in the database when the SQL runs (see
        _build_recount_mileage_sql_body).

        The optional 'milage_const' column (replaces update_milage_start_const): a
        non-empty, non-zero value is written into mileage_start.milage_const before the
        recount; for an asset without a mileage_start row one is created (INSERT), in
        which case a missing row is not an error. A value of 0 is ignored, as in the old
        script. Column headers are matched case-insensitively and ignoring edge
        whitespace (real files contain 'mileage_const\\xa0'), and alternative spellings
        are accepted: the asset column may be 'Актив' or 'active_number', the constant
        milage_const or mileage_const.
        """
        actives_repo = ActivesRepository(session=db_session)
        mileage_start_repo = MileageStartRepository(session=db_session)
        counter_repo = CounterActiveRepository(session=db_session)

        errors: list[dict] = []
        valid_rows: list[dict] = []
        batch_numbers: set[str] = set()

        active_column: str | None = next(
            (k for k in (rows[0] if rows else {})
             if str(k).strip().lower() in ("актив", "active_number")),
            None,
        )
        if rows and active_column is None:
            errors.append({"row": 0, "field": "Актив",
                            "message": "В файле не найдена колонка 'Актив' (или 'active_number')"})
            return errors, valid_rows
        const_column: str | None = next(
            (k for k in (rows[0] if rows else {})
             if str(k).strip().lower() in ("milage_const", "mileage_const")),
            None,
        )

        if progress is not None:
            progress.update(processed=0, total=len(rows), phase="validating")

        loc_old = aliased(Location)
        loc_new = aliased(Location)

        for idx, row in enumerate(rows):
            row_num = idx + 1
            if progress is not None:
                progress["processed"] = row_num

            active_number = str(row.get(active_column, "") or "").strip()
            if not active_number:
                errors.append({"row": row_num, "field": "Актив", "message": "Поле 'Актив' пустое"})
                continue

            if active_number in batch_numbers:
                errors.append({"row": row_num, "field": "Актив",
                                "message": f"Дубликат внутри файла: '{active_number}'"})
                continue

            active = await actives_repo.get_one_or_none(Actives.active_number == active_number)
            if active is None:
                errors.append({"row": row_num, "field": "Актив",
                                "message": f"Актив не найден: '{active_number}'"})
                continue
            is_train = active.id_unit_type == 1

            milage_const: int | None = None
            if const_column is not None:
                const_raw = row.get(const_column)
                if const_raw is not None and str(const_raw).strip() != "":
                    try:
                        milage_const = int(float(const_raw))
                    except (TypeError, ValueError):
                        errors.append({"row": row_num, "field": "milage_const",
                                        "message": f"Некорректное значение milage_const: '{const_raw}'"})
                        continue
                    if milage_const == 0:
                        milage_const = None

            mileage_start_rows = await mileage_start_repo.get_many(MileageStart.id_active == active.id)
            if len(mileage_start_rows) > 1:
                errors.append({"row": row_num, "field": "Актив",
                                "message": f"Несколько записей mileage_start для актива '{active_number}' — неоднозначно"})
                continue
            # Without milage_const the mileage_start row must exist; with it the row is
            # created by the SQL (INSERT), so a missing one is not an error.
            if not mileage_start_rows and milage_const is None:
                errors.append({"row": row_num, "field": "Актив",
                                "message": f"Запись mileage_start не найдена для актива '{active_number}'"})
                continue

            if not is_train:
                counters = await counter_repo.get_many(
                    CounterActive.id_active == active.id,
                    CounterActive.id_counter_type == MILEAGE_COUNTER_TYPE_ID,
                )
                if not counters:
                    errors.append({"row": row_num, "field": "Актив",
                                    "message": f"Счётчик пробега (counter_type={MILEAGE_COUNTER_TYPE_ID}) не найден для актива '{active_number}'"})
                    continue
                if len(counters) > 1:
                    errors.append({"row": row_num, "field": "Актив",
                                    "message": f"Несколько счётчиков пробега для актива '{active_number}' — неоднозначно"})
                    continue

            relocates = (await db_session.execute(
                select(
                    Relocate.date,
                    loc_old.id_train.label("id_train_old"),
                    loc_new.id_train.label("id_train_new"),
                )
                .join(loc_old, Relocate.id_location_old == loc_old.id, isouter=True)
                .join(loc_new, Relocate.id_location_new == loc_new.id, isouter=True)
                .where(Relocate.id_active == active.id)
                .order_by(Relocate.date)
            )).all()

            total = 0
            for relocate_date, id_train_old, id_train_new in relocates:
                if relocate_date is None:
                    continue
                if (relocate_date + MILEAGE_TZ_SHIFT).date() >= MILEAGE_RECOUNT_CUTOFF:
                    continue
                # Only storage<->train moves count; train->train and storage->storage
                # do not change mileage (as in the old script).
                if id_train_old is not None and id_train_new is None:
                    train_id, sign = id_train_old, -1
                elif id_train_old is None and id_train_new is not None:
                    train_id, sign = id_train_new, 1
                else:
                    continue

                train_mileage = (await db_session.execute(
                    select(func.sum(MileageTrain.mileage_average)).where(
                        MileageTrain.id_train == train_id,
                        MileageTrain.date_average > relocate_date,
                        MileageTrain.date_average <= MILEAGE_RECOUNT_CUTOFF,
                    )
                )).scalar()
                total += sign * (train_mileage or 0)

            batch_numbers.add(active_number)
            valid_rows.append({
                "row_num": row_num,
                "active_number": active_number,
                "id_active": active.id,
                "total": total,
                "is_train": is_train,
                "milage_const": milage_const,
                "insert_mileage_start": not mileage_start_rows,
            })

        return errors, valid_rows

    @staticmethod
    def _build_recount_mileage_sql_body(valid_rows: list[dict]) -> list[str]:
        """Build the mileage recount SQL (without BEGIN/COMMIT), shared by download and execution.

        The order is mandatory: first the milage_const from the file (UPDATE, or INSERT
        for assets without mileage_start), then UPDATE mileage_start.milage — which
        reads milage_const at execution time and must already see the new value — then
        UPDATE counter_active, whose recount through function_get_mileage() reads
        mileage_start.milage. All in one transaction, so a file downloaded and run later
        cannot drift from the database. For trains (id_unit_type=1) the counter is not
        recomputed (as in the old script).
        """
        sql_lines: list[str] = []
        for vr in valid_rows:
            if vr["milage_const"] is None:
                continue
            if vr["insert_mileage_start"]:
                sql_lines.append(
                    f"INSERT INTO public.mileage_start (id_active, milage, milage_const, is_recount) "
                    f"VALUES ({vr['id_active']}, {vr['milage_const']}, {vr['milage_const']}, true); "
                    f"-- '{sql_escape(vr['active_number'])}'"
                )
            else:
                sql_lines.append(
                    f"UPDATE public.mileage_start SET milage_const = {vr['milage_const']}, is_recount = true "
                    f"WHERE id_active = {vr['id_active']}; -- '{sql_escape(vr['active_number'])}'"
                )
        for vr in valid_rows:
            sql_lines.append(
                f"UPDATE public.mileage_start SET milage = COALESCE(milage_const, 0) + ({vr['total']}), "
                f"is_recount = true WHERE id_active = {vr['id_active']}; -- '{sql_escape(vr['active_number'])}'"
            )
        for vr in valid_rows:
            if vr["is_train"]:
                continue
            sql_lines.append(
                f"UPDATE public.counter_active SET value = COALESCE("
                f"(SELECT sum FROM public.function_get_mileage(id_active, date::date)), 0) "
                f"WHERE id_active = {vr['id_active']} AND id_counter_type = {MILEAGE_COUNTER_TYPE_ID}; "
                f"-- '{sql_escape(vr['active_number'])}'"
            )
        return sql_lines

    @post("/recount-mileage/generate-sql/start")
    async def recount_mileage_generate_sql_start(self, request: Request) -> Response:
        """Start background generation of the mileage recount SQL file; returns a task_id."""
        session_id = request.session.get(f"{PREFIX}_session_id", "")
        stored = _load_data(session_id) if session_id else None
        if not stored:
            return Response(
                content=json.dumps({"status": "error", "errors": [{"row": 0, "field": "*", "message": "Данные не загружены"}]}),
                status_code=200,
                media_type="application/json",
            )

        rows: list[dict] = stored["rows"]

        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_recount_mileage_generate(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task

        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_recount_mileage_generate(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                errors, valid_rows = await self._validate_recount_mileage(session, rows, progress=progress)
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка: {e}"}])
            return

        if errors:
            progress.update(status="error", errors=errors)
            return
        if not valid_rows:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": "Нет валидных строк для пересчёта"}])
            return

        sql_lines = self._build_recount_mileage_sql_body(valid_rows)
        full_sql = "\n".join(["BEGIN;", *sql_lines, "COMMIT;"])
        progress.update(status="done", sql=full_sql, count=len(valid_rows))

    @post("/recount-mileage/execute/start")
    async def recount_mileage_execute_start(self, request: Request) -> Response:
        """Start the background atomic mileage recount in the database; returns a task_id."""
        session_id = request.session.get(f"{PREFIX}_session_id", "")
        stored = _load_data(session_id) if session_id else None
        if not stored:
            return Response(
                content=json.dumps({"status": "error", "errors": [{"row": 0, "field": "*", "message": "Данные не загружены"}]}),
                status_code=200,
                media_type="application/json",
            )

        rows: list[dict] = stored["rows"]

        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_recount_mileage_execute(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task

        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_recount_mileage_execute(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        valid_rows: list[dict] = []
        sql_lines: list[str] = []
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                errors, valid_rows = await self._validate_recount_mileage(session, rows, progress=progress)

                if errors:
                    progress.update(status="error", errors=errors)
                    return
                if not valid_rows:
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": "Нет валидных строк для пересчёта"}])
                    return

                progress.update(processed=0, total=1, phase="executing")

                sql_lines = self._build_recount_mileage_sql_body(valid_rows)
                sql_body = "\n".join(sql_lines)

                try:
                    # IMPORTANT: as in create-actives, session.rollback() below also rolls this
                    # raw call back only because the session opened a real transaction earlier
                    # (the repository queries inside _validate_recount_mileage). Do not remove
                    # the session usage before this point.
                    conn = await session.connection()
                    raw_conn = await conn.get_raw_connection()
                    await raw_conn.driver_connection.execute(sql_body)

                    await session.commit()
                except Exception as e:
                    await session.rollback()
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
                    return

                progress["processed"] = 1
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
            return

        now = datetime.now()
        log_lines = [
            f"=== Recount Mileage: {now.strftime('%Y-%m-%d %H:%M:%S')} ===",
            f"Actives recounted: {len(valid_rows)}",
            "",
            *sql_lines,
            "",
        ]
        log_file = LOG_DIR / f"recount_mileage_{now.strftime('%Y-%m-%d_%H-%M-%S')}.log"
        with open(log_file, "a", encoding="utf-8") as f:
            f.write("\n".join(log_lines))
        logger.info("Recounted mileage for %d actives, log: %s", len(valid_rows), log_file)

        progress.update(status="done", count=len(valid_rows),
                         message=f"Пробег пересчитан для {len(valid_rows)} активов")

    async def _validate_create_actives_rows(
        self, db_session: AsyncSession, rows: list[dict], progress: dict | None = None,
    ) -> tuple[list[dict], list[dict]]:
        """Validate the material rows and build the asset creation plan (replaces add_active_spcial).

        valid_rows holds one element per asset to create ('Количество' expanded row by
        row), and each asset gets its own location record.

        The current storage.last_lcn and iterator_number_last.number are deliberately not
        read here for use in the SQL — only to check that the counter exists. The values
        themselves are read and locked (FOR UPDATE) inside the DO block when it runs (see
        _build_create_actives_sql_body) rather than snapshotted on the Python side during
        validation — otherwise by the time of the real run (especially for a file
        downloaded and executed later) the value in the database could have moved on.
        """
        storage_repo = StorageRepository(session=db_session)
        storage_place_repo = StoragePlaceRepository(session=db_session)
        consignment_repo = ConsignmentRepository(session=db_session)
        design_number_repo = DesignNumberRepository(session=db_session)
        iterator_repo = IteratorNumberLastRepository(session=db_session)

        errors: list[dict] = []
        valid_rows: list[dict] = []

        storage_cache: dict[str, Storage | None] = {}
        storage_place_cache: dict[str, int | None] = {}
        consignment_cache: dict[str, int | None] = {}
        design_number_cache: dict[str, int | None] = {}

        counter_row = await iterator_repo.get_one_or_none(
            IteratorNumberLast.description == ACTIVE_NUMBER_COUNTER_DESCRIPTION
        )
        if counter_row is None or counter_row.number is None:
            errors.append({
                "row": 0, "field": "*",
                "message": f"Не найден счётчик '{ACTIVE_NUMBER_COUNTER_DESCRIPTION}' в iterator_number_last",
            })
            return errors, valid_rows

        # The material column is called 'АРТИКУЛ' in newer files and
        # 'Номер ТМЦ (DU,KP,A2V)' in older ones; matched ignoring case and edge spaces.
        tmc_column: str | None = next(
            (k for k in (rows[0] if rows else {})
             if str(k).strip().lower() in ("артикул", "номер тмц (du,kp,a2v)")),
            None,
        )
        if rows and tmc_column is None:
            errors.append({"row": 0, "field": "АРТИКУЛ",
                            "message": "В файле не найдена колонка 'АРТИКУЛ' (или 'Номер ТМЦ (DU,KP,A2V)')"})
            return errors, valid_rows

        if progress is not None:
            progress.update(processed=0, total=len(rows), phase="validating")

        for idx, row in enumerate(rows):
            row_num = idx + 1
            if progress is not None and (idx % 20 == 0 or row_num == len(rows)):
                progress["processed"] = row_num

            design_number_raw = str(row.get(tmc_column) or "").strip()
            if not design_number_raw or design_number_raw == "None":
                continue

            count_raw = row.get("Количество")
            try:
                count = int(count_raw)
            except (TypeError, ValueError):
                errors.append({"row": row_num, "field": "Количество", "message": f"Некорректное количество: '{count_raw}'"})
                continue
            if count < 1:
                errors.append({"row": row_num, "field": "Количество", "message": f"Количество должно быть больше 0: '{count_raw}'"})
                continue

            storage_name = str(row.get("Склад") or "").strip()
            if not storage_name:
                errors.append({"row": row_num, "field": "Склад", "message": "Поле 'Склад' пустое"})
                continue
            if storage_name not in storage_cache:
                storage_cache[storage_name] = await storage_repo.get_one_or_none(Storage.name == storage_name)
            storage = storage_cache[storage_name]
            if storage is None:
                errors.append({"row": row_num, "field": "Склад", "message": f"Склад не найден: '{storage_name}'"})
                continue
            type_active = str(row.get("Тип актива") or "").strip()
            if not type_active:
                errors.append({"row": row_num, "field": "Тип актива", "message": "Поле 'Тип актива' пустое"})
                continue
            if len(type_active) >= ACTIVE_NUMBER_LENGTH:
                errors.append({
                    "row": row_num, "field": "Тип актива",
                    "message": (f"'{type_active}' слишком длинный для {ACTIVE_NUMBER_LENGTH}-значного "
                                f"номера актива (префикс + цифры)"),
                })
                continue

            special_raw = row.get("Особый учет")
            special = str(special_raw).strip() if special_raw not in (None, "") else None

            storage_place_name = str(row.get("Ячейка") or "").strip()
            id_storage_place: int | None = None
            if storage_place_name and storage_place_name != "None":
                if storage_place_name not in storage_place_cache:
                    sp = await storage_place_repo.get_one_or_none(StoragePlace.name == storage_place_name)
                    storage_place_cache[storage_place_name] = sp.id if sp else None
                id_storage_place = storage_place_cache[storage_place_name]
                if id_storage_place is None:
                    errors.append({"row": row_num, "field": "Ячейка", "message": f"Ячейка не найдена: '{storage_place_name}'"})
                    continue

            consignment_name = str(row.get("Партия") or "").strip()
            if not consignment_name:
                errors.append({"row": row_num, "field": "Партия", "message": "Поле 'Партия' пустое"})
                continue
            if consignment_name not in consignment_cache:
                c = await consignment_repo.get_one_or_none(Consignment.name == consignment_name)
                consignment_cache[consignment_name] = c.id if c else None
            id_consignment = consignment_cache[consignment_name]
            if id_consignment is None:
                errors.append({"row": row_num, "field": "Партия", "message": f"Партия не найдена: '{consignment_name}'"})
                continue

            if design_number_raw not in design_number_cache:
                dn = await design_number_repo.get_one_or_none(DesignNumber.number == design_number_raw)
                design_number_cache[design_number_raw] = dn.id if dn else None
            id_design_number = design_number_cache[design_number_raw]
            if id_design_number is None:
                errors.append({"row": row_num, "field": tmc_column, "message": f"ТМЦ не найдена: '{design_number_raw}'"})
                continue

            serial_number = (str(row.get("Серийный номер") or "").strip() if count == 1 else "none")

            for _ in range(count):
                valid_rows.append({
                    "row_num": row_num,
                    "id_design_number": id_design_number,
                    "id_storage": storage.id,
                    "id_storage_place": id_storage_place,
                    "id_consignment": id_consignment,
                    "type_active": type_active,
                    "special_account": special,
                    "serial_number": serial_number or None,
                })

        return errors, valid_rows

    @staticmethod
    def _build_create_actives_sql_body(valid_rows: list[dict]) -> list[str]:
        """Build the body of the script creating assets from materials (without BEGIN/COMMIT).

        Every asset gets its own location record (materials is not used). Shared by
        "Скачать SQL-файл" and "Выполнить в базе данных". All counters are read and
        consumed inside the DO block itself as it runs, not on the Python side during
        generation:
        - ids for location/actives come from nextval() (as in train_parser);
        - storage.last_lcn and iterator_number_last.number come from
          `SELECT ... INTO var ... FOR UPDATE`, are incremented locally, and the same
          variable is written back with `UPDATE` at the end of that block (FOR UPDATE
          holds the row lock until the transaction ends, which rules out a race on
          concurrent runs).
        Without this a file downloaded but executed later (or two runs in a row) could
        drift from the value in the database: an id from nextval() is always unique on its
        own, while last_lcn/number are ordinary integer columns, and the old peewee code
        computed them once on the Python side.
        """
        total_actives = len(valid_rows)
        storage_ids = sorted({vr["id_storage"] for vr in valid_rows})

        sql_lines: list[str] = ["DO $$", "DECLARE"]
        sql_lines.append(
            f"    loc_ids bigint[] := ARRAY(SELECT nextval('public.location_id_seq') "
            f"FROM generate_series(1, {total_actives}));"
        )
        sql_lines.append(
            f"    act_ids bigint[] := ARRAY(SELECT nextval('public.actives_id_seq') "
            f"FROM generate_series(1, {total_actives}));"
        )
        sql_lines.append("    active_num bigint;")
        for storage_id in storage_ids:
            sql_lines.append(f"    lcn_{storage_id} bigint;")

        sql_lines.append("BEGIN")
        sql_lines.append(
            f"    SELECT number INTO active_num FROM public.iterator_number_last "
            f"WHERE description = '{sql_escape(ACTIVE_NUMBER_COUNTER_DESCRIPTION)}' FOR UPDATE;"
        )
        sql_lines.append(
            f"    IF active_num IS NULL THEN RAISE EXCEPTION "
            f"'Счётчик ''{sql_escape(ACTIVE_NUMBER_COUNTER_DESCRIPTION)}'' не найден или пуст'; END IF;"
        )
        for storage_id in storage_ids:
            sql_lines.append(
                f"    SELECT last_lcn INTO lcn_{storage_id} FROM public.storage WHERE id = {storage_id} FOR UPDATE;"
            )

        body_lines: list[str] = []

        for i, vr in enumerate(valid_rows, start=1):
            loc_ref = f"loc_ids[{i}]"
            act_ref = f"act_ids[{i}]"
            sp_val = str(vr["id_storage_place"]) if vr["id_storage_place"] is not None else "NULL"

            # The increments do not touch the location row, but are moved ahead of both
            # INSERTs so location and actives always sit next to each other as one pair.
            lcn_var = f"lcn_{vr['id_storage']}"
            body_lines.append(f"    {lcn_var} := {lcn_var} + 1;")
            body_lines.append("    active_num := active_num + 1;")

            body_lines.append(
                f"    INSERT INTO public.location (id, id_type_location, id_storage, id_storage_place, id_consignment) "
                f"VALUES ({loc_ref}, 1, {vr['id_storage']}, {sp_val}, {vr['id_consignment']});"
            )

            sn_val = f"'{sql_escape(vr['serial_number'])}'" if vr["serial_number"] else "NULL"
            sa_val = f"'{sql_escape(vr['special_account'])}'" if vr["special_account"] else "NULL"
            # An asset number of fixed length ACTIVE_NUMBER_LENGTH: the letter prefix plus
            # the counter digits, zero-padded on the left to the required width (validation
            # above guarantees len(type_active) < ACTIVE_NUMBER_LENGTH, so the width > 0).
            number_width = ACTIVE_NUMBER_LENGTH - len(vr["type_active"])
            active_number_expr = (
                f"'{sql_escape(vr['type_active'])}' || lpad(active_num::text, {number_width}, '0')"
            )
            lcn_expr = f"('S{vr['id_storage']}.' || {lcn_var})::ltree"

            body_lines.append(
                f"    INSERT INTO public.actives (id, active_number, id_design_number, id_location, "
                f"serial_number, lcn, special_account) "
                f"VALUES ({act_ref}, {active_number_expr}, {vr['id_design_number']}, {loc_ref}, "
                f"{sn_val}, {lcn_expr}, {sa_val});"
            )

        sql_lines.extend(body_lines)

        for storage_id in storage_ids:
            sql_lines.append(f"    UPDATE public.storage SET last_lcn = lcn_{storage_id} WHERE id = {storage_id};")
        sql_lines.append(
            f"    UPDATE public.iterator_number_last SET number = active_num "
            f"WHERE description = '{sql_escape(ACTIVE_NUMBER_COUNTER_DESCRIPTION)}';"
        )

        sql_lines.append("END $$;")

        return sql_lines

    async def _validate_create_active_from_model_rows(
        self, db_session: AsyncSession, rows: list[dict], progress: dict | None = None,
    ) -> tuple[list[dict], list[dict], int]:
        """Validate the rows for 'Создать актив из модели lcn'.

        In an lsn like 'M9.1.6.4': 9 is id_train_type, the first path segment ('1') is
        car_number, and the whole path after id_train_type ('1.6.4') is the remainder used
        to build the real asset lcn on a specific train. id_car_place is taken from
        public.models by that same lcn (WHERE id_train_type=... AND lcn::text=lsn) — the
        same lookup principle train_parser.py uses when creating a new train.

        One lcn often appears in several models rows with different id_design_number
        (alternative material positions for the same place in the model) — in practice
        their id_car_place usually matches, and is_default may be false on ALL of the rows
        at once (verified against the real database: 'M9.2.5.4' has 2 rows, both
        is_default=false but with the same id_car_place). So all distinct id_car_place
        values are collected first, without filtering on is_default; if there is only one,
        it is used. Only when car_place genuinely differs between rows does is_default=true
        come in as a tie-break; if that still leaves it ambiguous, the row is an error.

        The number of assets created per position does NOT come from a column in the file
        but from the number of trains of that id_train_type (public.train): each train gets
        its own asset with its own location (id_type_location=2, id_train, car_number,
        id_car_place). Rows with identical or overlapping lsn share one cache of trains by
        type (as in _validate_and_build_serial_none_rows in parser.py).

        Assets whose computed lcn is already taken by an existing one (the position is
        already filled on some of the fleet) are skipped silently — not a validation error
        but the ordinary case of partial fleet coverage. Returns
        (errors, valid_rows, skipped_count).
        """
        errors: list[dict] = []
        candidates: list[dict] = []

        tmc_column: str | None = next(
            (k for k in (rows[0] if rows else {})
             if str(k).strip().lower() in ("артикул", "номер тмц (du,kp,a2v)")),
            None,
        )
        if rows and tmc_column is None:
            errors.append({"row": 0, "field": "АРТИКУЛ",
                            "message": "В файле не найдена колонка 'АРТИКУЛ' (или 'Номер ТМЦ (DU,KP,A2V)')"})
            return errors, [], 0

        lsn_column: str | None = next(
            (k for k in (rows[0] if rows else {}) if str(k).strip().lower() in ("lsn", "lcn")),
            None,
        )
        if rows and lsn_column is None:
            errors.append({"row": 0, "field": "lcn", "message": "В файле не найдена колонка 'lsn' (или 'lcn')"})
            return errors, [], 0

        counter_repo = IteratorNumberLastRepository(session=db_session)
        counter_row = await counter_repo.get_one_or_none(
            IteratorNumberLast.description == ACTIVE_NUMBER_COUNTER_DESCRIPTION
        )
        if counter_row is None or counter_row.number is None:
            errors.append({
                "row": 0, "field": "*",
                "message": f"Не найден счётчик '{ACTIVE_NUMBER_COUNTER_DESCRIPTION}' в iterator_number_last",
            })
            return errors, [], 0

        design_number_cache: dict[str, int | None] = {}
        train_ids_by_type: dict[int, list[int]] = {}
        model_cache: dict[tuple[int, str], list[int | None]] = {}

        if progress is not None:
            progress.update(processed=0, total=len(rows), phase="validating")

        for idx, row in enumerate(rows):
            row_num = idx + 1
            if progress is not None and (idx % 20 == 0 or row_num == len(rows)):
                progress["processed"] = row_num

            design_number_raw = str(row.get(tmc_column) or "").strip()
            if not design_number_raw or design_number_raw == "None":
                continue

            type_active = str(row.get("Тип актива") or "").strip()
            if not type_active:
                errors.append({"row": row_num, "field": "Тип актива", "message": "Поле 'Тип актива' пустое"})
                continue
            if len(type_active) >= ACTIVE_NUMBER_LENGTH:
                errors.append({
                    "row": row_num, "field": "Тип актива",
                    "message": (f"'{type_active}' слишком длинный для {ACTIVE_NUMBER_LENGTH}-значного "
                                f"номера актива (префикс + цифры)"),
                })
                continue

            lsn_raw = str(row.get(lsn_column) or "").strip()
            if not lsn_raw:
                errors.append({"row": row_num, "field": "lcn", "message": "Пустой lsn"})
                continue

            parsed = _parse_model_lcn(lsn_raw)
            if parsed is None:
                errors.append({"row": row_num, "field": "lcn",
                                "message": f"Не удалось распознать id_train_type в lcn '{lsn_raw}'"})
                continue
            id_train_type, rest = parsed

            if design_number_raw not in design_number_cache:
                design_number_cache[design_number_raw] = await db_session.scalar(
                    select(DesignNumber.id).where(DesignNumber.number == design_number_raw)
                )
            id_design_number = design_number_cache[design_number_raw]
            if id_design_number is None:
                errors.append({"row": row_num, "field": tmc_column, "message": f"ТМЦ не найдена: '{design_number_raw}'"})
                continue

            model_key = (id_train_type, lsn_raw)
            if model_key not in model_cache:
                model_result = await db_session.execute(
                    text(
                        "SELECT id_car_place, is_default FROM public.models "
                        "WHERE id_train_type = :tt AND lcn::text = :lcn"
                    ),
                    {"tt": id_train_type, "lcn": lsn_raw},
                )
                model_cache[model_key] = model_result.all()
            model_rows = model_cache[model_key]
            if not model_rows:
                errors.append({"row": row_num, "field": "lcn", "message": f"Модель с lcn '{lsn_raw}' не найдена"})
                continue
            distinct_places = {r[0] for r in model_rows}
            if len(distinct_places) == 1:
                id_car_place = next(iter(distinct_places))
            else:
                # One lcn may map to several models rows with different id_design_number
                # (alternative positions) — when those also differ on id_car_place, the row
                # with is_default=true wins. is_default=true is not guaranteed for every lcn
                # (see the case above where both rows are non-default but agree on
                # car_place — that path never reaches here).
                default_places = {r[0] for r in model_rows if r[1]}
                if len(default_places) == 1:
                    id_car_place = next(iter(default_places))
                else:
                    errors.append({"row": row_num, "field": "lcn",
                                    "message": (f"Модель с lcn '{lsn_raw}' неоднозначна: несколько разных "
                                                f"id_car_place {sorted(p for p in distinct_places if p is not None)}")})
                    continue

            car_number: int | None = None
            if rest:
                first_segment = rest.split(".")[0]
                try:
                    car_number = int(first_segment)
                except ValueError:
                    errors.append({"row": row_num, "field": "lcn",
                                    "message": f"Не удалось распознать номер вагона в lcn '{lsn_raw}'"})
                    continue

            if id_train_type not in train_ids_by_type:
                result = await db_session.execute(select(Train.id).where(Train.id_train_type == id_train_type))
                train_ids_by_type[id_train_type] = [r[0] for r in result.all()]
            train_ids = train_ids_by_type[id_train_type]
            if not train_ids:
                errors.append({"row": row_num, "field": "lcn",
                                "message": f"Поезда с id_train_type={id_train_type} не найдены"})
                continue

            # The serial number from the file is applied only when exactly one asset is
            # created for the position (a single train of that type) — otherwise the same
            # serial would be duplicated across assets (as in _validate_create_actives_rows).
            serial_raw = str(row.get("Серийный номер") or "").strip()
            serial_number = serial_raw if len(train_ids) == 1 else "none"

            for train_id in train_ids:
                new_lcn = f"{train_id}.{rest}" if rest else str(train_id)
                candidates.append({
                    "row_num": row_num,
                    "id_design_number": id_design_number,
                    "type_active": type_active,
                    "id_train": train_id,
                    "car_number": car_number,
                    "id_car_place": id_car_place,
                    "lcn": new_lcn,
                    "serial_number": serial_number or None,
                })

        if not candidates:
            return errors, [], 0

        # A batched check of target lcn occupancy (one query for the whole file rather than
        # per row) plus deduplication within the batch — both are skipped silently, not errors.
        all_lcns = [c["lcn"] for c in candidates]
        stmt = text(
            "SELECT lcn::text FROM public.actives WHERE lcn::text IN :lcns"
        ).bindparams(bindparam("lcns", expanding=True))
        existing_lcns = set((await db_session.execute(stmt, {"lcns": all_lcns})).scalars().all())

        valid_rows: list[dict] = []
        seen_lcns: set[str] = set()
        skipped_count = 0
        for c in candidates:
            if c["lcn"] in existing_lcns or c["lcn"] in seen_lcns:
                skipped_count += 1
                continue
            seen_lcns.add(c["lcn"])
            valid_rows.append(c)

        return errors, valid_rows, skipped_count

    @staticmethod
    def _build_create_active_from_model_sql_body(valid_rows: list[dict]) -> list[str]:
        """Build the SQL creating assets from model positions (without BEGIN/COMMIT).

        Unlike _build_create_actives_sql_body there is no storage counter here
        (storage.last_lcn under FOR UPDATE) — each asset's real lcn is already determined
        by the model (id_train plus the path from lsn), so FOR UPDATE is only needed on the
        shared asset number counter iterator_number_last.
        """
        total = len(valid_rows)

        sql_lines: list[str] = ["DO $$", "DECLARE"]
        sql_lines.append(
            f"    loc_ids bigint[] := ARRAY(SELECT nextval('public.location_id_seq') "
            f"FROM generate_series(1, {total}));"
        )
        sql_lines.append(
            f"    act_ids bigint[] := ARRAY(SELECT nextval('public.actives_id_seq') "
            f"FROM generate_series(1, {total}));"
        )
        sql_lines.append("    active_num bigint;")
        sql_lines.append("BEGIN")
        sql_lines.append(
            f"    SELECT number INTO active_num FROM public.iterator_number_last "
            f"WHERE description = '{sql_escape(ACTIVE_NUMBER_COUNTER_DESCRIPTION)}' FOR UPDATE;"
        )
        sql_lines.append(
            f"    IF active_num IS NULL THEN RAISE EXCEPTION "
            f"'Счётчик ''{sql_escape(ACTIVE_NUMBER_COUNTER_DESCRIPTION)}'' не найден или пуст'; END IF;"
        )

        for i, vr in enumerate(valid_rows, start=1):
            loc_ref = f"loc_ids[{i}]"
            act_ref = f"act_ids[{i}]"
            sql_lines.append("    active_num := active_num + 1;")

            car_number_val = str(vr["car_number"]) if vr["car_number"] is not None else "NULL"
            car_place_val = str(vr["id_car_place"]) if vr["id_car_place"] is not None else "NULL"
            sql_lines.append(
                f"    INSERT INTO public.location (id, id_type_location, id_train, car_number, id_car_place) "
                f"VALUES ({loc_ref}, 2, {vr['id_train']}, {car_number_val}, {car_place_val});"
            )

            sn_val = f"'{sql_escape(vr['serial_number'])}'" if vr["serial_number"] else "NULL"
            number_width = ACTIVE_NUMBER_LENGTH - len(vr["type_active"])
            active_number_expr = (
                f"'{sql_escape(vr['type_active'])}' || lpad(active_num::text, {number_width}, '0')"
            )
            sql_lines.append(
                f"    INSERT INTO public.actives (id, active_number, id_design_number, id_location, "
                f"serial_number, lcn) VALUES ({act_ref}, {active_number_expr}, {vr['id_design_number']}, {loc_ref}, "
                f"{sn_val}, '{sql_escape(vr['lcn'])}'::ltree);"
            )

        sql_lines.append(
            f"    UPDATE public.iterator_number_last SET number = active_num "
            f"WHERE description = '{sql_escape(ACTIVE_NUMBER_COUNTER_DESCRIPTION)}';"
        )
        sql_lines.append("END $$;")

        return sql_lines

    async def _validate_delete_actives(
        self, db_session: AsyncSession, rows: list[dict], progress: dict | None = None,
    ) -> tuple[list[dict], list[dict]]:
        """Validate the rows of an "удалить активы" file (a strict DELETE).

        Only assets without history are removed: if a single record from
        DELETE_ACTIVES_BLOCKERS (maintenance, moves, orders and so on) references the
        asset, the row is an error and the asset stays. Blockers are checked in batches:
        one query per table for all the file's assets, not one query per row.
        """
        actives_repo = ActivesRepository(session=db_session)

        errors: list[dict] = []
        batch_numbers: set[str] = set()
        # id_active -> the row's data; the valid ones remain after the batched blocker check
        candidates: dict[int, dict] = {}

        active_column: str | None = next(
            (k for k in (rows[0] if rows else {})
             if str(k).strip().lower() in ("актив", "active_number")),
            None,
        )
        if rows and active_column is None:
            errors.append({"row": 0, "field": "Актив",
                            "message": "В файле не найдена колонка 'Актив' (или 'active_number')"})
            return errors, []

        if progress is not None:
            progress.update(processed=0, total=len(rows), phase="validating")

        for idx, row in enumerate(rows):
            row_num = idx + 1
            if progress is not None and (idx % 20 == 0 or row_num == len(rows)):
                progress["processed"] = row_num

            active_number = str(row.get(active_column, "") or "").strip()
            if not active_number:
                errors.append({"row": row_num, "field": "Актив", "message": "Поле 'Актив' пустое"})
                continue

            if active_number in batch_numbers:
                errors.append({"row": row_num, "field": "Актив",
                                "message": f"Дубликат внутри файла: '{active_number}'"})
                continue

            active = await actives_repo.get_one_or_none(Actives.active_number == active_number)
            if active is None:
                errors.append({"row": row_num, "field": "Актив",
                                "message": f"Актив не найден: '{active_number}'"})
                continue

            batch_numbers.add(active_number)
            candidates[active.id] = {
                "row_num": row_num,
                "active_number": active_number,
                "id_active": active.id,
                "id_location": active.id_location,
            }

        if candidates:
            ids = list(candidates)
            for table_name, column in DELETE_ACTIVES_BLOCKERS:
                blocked_ids = (await db_session.execute(
                    select(column).where(column.in_(ids)).distinct()
                )).scalars().all()
                for blocked_id in blocked_ids:
                    vr = candidates.pop(blocked_id, None)
                    if vr is not None:
                        errors.append({
                            "row": vr["row_num"], "field": "Актив",
                            "message": (f"У актива '{vr['active_number']}' есть связанные записи "
                                        f"в {table_name} — удаление запрещено"),
                        })

        if candidates:
            # An asset's orders (directly by id_active or through its maintenance record)
            # are deleted with it, but only the "empty" ones: an order with related records
            # in any of ORDERS_DEPENDENCY_CHECKS is real working history, and blocks the asset.
            ids = list(candidates)
            order_rows = (await db_session.execute(
                select(Orders.id, Orders.id_active, Ptoir.id_active)
                .outerjoin(Ptoir, Orders.id_ptoir == Ptoir.id)
                .where(or_(Orders.id_active.in_(ids), Ptoir.id_active.in_(ids)))
            )).all()
            order_owners: dict[int, set[int]] = {}
            for order_id, direct_active, ptoir_active in order_rows:
                owners = order_owners.setdefault(order_id, set())
                for owner in (direct_active, ptoir_active):
                    if owner in candidates:
                        owners.add(owner)

            if order_owners:
                order_ids = list(order_owners)
                for table_name, column in ORDERS_DEPENDENCY_CHECKS:
                    stmt = text(
                        f"SELECT DISTINCT {column} FROM public.{table_name} WHERE {column} IN :ids"
                    ).bindparams(bindparam("ids", expanding=True))
                    blocked_orders = (await db_session.execute(stmt, {"ids": order_ids})).scalars().all()
                    for order_id in blocked_orders:
                        for owner in order_owners.get(order_id, ()):
                            vr = candidates.pop(owner, None)
                            if vr is not None:
                                errors.append({
                                    "row": vr["row_num"], "field": "Актив",
                                    "message": (f"У заказов актива '{vr['active_number']}' есть "
                                                f"связанные записи в {table_name} — удаление запрещено"),
                                })

        valid_rows = sorted(candidates.values(), key=lambda vr: vr["row_num"])
        return errors, valid_rows

    @staticmethod
    def _build_delete_actives_sql_body(valid_rows: list[dict]) -> list[str]:
        """Build the asset deletion SQL (without BEGIN/COMMIT), shared by download and execution.

        Per asset: its "empty" orders (directly and through maintenance — validation
        guarantees they have no related records), the ptoir_level_warning of its
        maintenance records and the ptoir rows themselves, the counter_active rows (created
        by the trigger on asset INSERT) and mileage_start (no FK — they would be orphaned
        otherwise), the asset itself, and then its location — but only when nothing else
        references it any more (other actives, materials, relocate); the NOT EXISTS check
        runs when the SQL executes, after the asset is already gone. Orders are deleted
        before ptoir because of the FK orders.id_ptoir -> ptoir.

        DELETE from counter_active is forbidden by the DBA trigger tr_abort_delete
        (dba.fn_abort_delete, an unconditional RAISE) — without disabling it temporarily an
        asset cannot be deleted at all: every asset gets a counter automatically, and the FK
        counter_active->actives blocks deleting the asset itself. The trigger is disabled
        only within this transaction (ALTER TABLE takes ACCESS EXCLUSIVE until it ends) and
        requires table owner privileges.
        """
        sql_lines: list[str] = [
            "ALTER TABLE public.counter_active DISABLE TRIGGER tr_abort_delete;"
        ]
        for vr in valid_rows:
            comment = f" -- '{sql_escape(vr['active_number'])}'"
            sql_lines.append(
                f"DELETE FROM public.orders WHERE id_active = {vr['id_active']} OR id_ptoir IN "
                f"(SELECT id FROM public.ptoir WHERE id_active = {vr['id_active']});{comment}"
            )
            sql_lines.append(
                f"DELETE FROM public.ptoir_level_warning WHERE id_ptoir IN "
                f"(SELECT id FROM public.ptoir WHERE id_active = {vr['id_active']});{comment}"
            )
            sql_lines.append(f"DELETE FROM public.ptoir WHERE id_active = {vr['id_active']};{comment}")
            sql_lines.append(f"DELETE FROM public.counter_active WHERE id_active = {vr['id_active']};{comment}")
            sql_lines.append(f"DELETE FROM public.mileage_start WHERE id_active = {vr['id_active']};{comment}")
            sql_lines.append(f"DELETE FROM public.actives WHERE id = {vr['id_active']};{comment}")
            if vr["id_location"] is not None:
                sql_lines.append(
                    f"DELETE FROM public.location l WHERE l.id = {vr['id_location']} "
                    f"AND NOT EXISTS (SELECT 1 FROM public.actives a WHERE a.id_location = l.id) "
                    f"AND NOT EXISTS (SELECT 1 FROM public.materials m WHERE m.id_location = l.id) "
                    f"AND NOT EXISTS (SELECT 1 FROM public.relocate r WHERE r.id_location_old = l.id "
                    f"OR r.id_location_new = l.id);{comment}"
                )
        sql_lines.append("ALTER TABLE public.counter_active ENABLE TRIGGER tr_abort_delete;")
        return sql_lines

    @post("/delete-actives/generate-sql/start")
    async def delete_actives_generate_sql_start(self, request: Request) -> Response:
        """Start background generation of the asset deletion SQL file; returns a task_id."""
        session_id = request.session.get(f"{PREFIX}_session_id", "")
        stored = _load_data(session_id) if session_id else None
        if not stored:
            return Response(
                content=json.dumps({"status": "error", "errors": [{"row": 0, "field": "*", "message": "Данные не загружены"}]}),
                status_code=200,
                media_type="application/json",
            )

        rows: list[dict] = stored["rows"]

        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_delete_actives_generate(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task

        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_delete_actives_generate(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                errors, valid_rows = await self._validate_delete_actives(session, rows, progress=progress)
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка: {e}"}])
            return

        if errors:
            progress.update(status="error", errors=errors)
            return
        if not valid_rows:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": "Нет валидных строк для удаления"}])
            return

        sql_lines = self._build_delete_actives_sql_body(valid_rows)
        full_sql = "\n".join(["BEGIN;", *sql_lines, "COMMIT;"])
        progress.update(status="done", sql=full_sql, count=len(valid_rows))

    @post("/delete-actives/execute/start")
    async def delete_actives_execute_start(self, request: Request) -> Response:
        """Start the background atomic asset deletion; returns a task_id."""
        session_id = request.session.get(f"{PREFIX}_session_id", "")
        stored = _load_data(session_id) if session_id else None
        if not stored:
            return Response(
                content=json.dumps({"status": "error", "errors": [{"row": 0, "field": "*", "message": "Данные не загружены"}]}),
                status_code=200,
                media_type="application/json",
            )

        rows: list[dict] = stored["rows"]

        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_delete_actives_execute(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task

        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_delete_actives_execute(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        valid_rows: list[dict] = []
        sql_lines: list[str] = []
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                errors, valid_rows = await self._validate_delete_actives(session, rows, progress=progress)

                if errors:
                    progress.update(status="error", errors=errors)
                    return
                if not valid_rows:
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": "Нет валидных строк для удаления"}])
                    return

                progress.update(processed=0, total=1, phase="executing")

                sql_lines = self._build_delete_actives_sql_body(valid_rows)
                sql_body = "\n".join(sql_lines)

                try:
                    # IMPORTANT: as in create-actives, session.rollback() below also rolls this
                    # raw call back only because the session opened a real transaction earlier
                    # (the repository queries inside _validate_delete_actives). Do not remove
                    # the session usage before this point.
                    conn = await session.connection()
                    raw_conn = await conn.get_raw_connection()
                    await raw_conn.driver_connection.execute(sql_body)

                    await session.commit()
                except Exception as e:
                    await session.rollback()
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
                    return

                progress["processed"] = 1
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
            return

        now = datetime.now()
        log_lines = [
            f"=== Delete Actives: {now.strftime('%Y-%m-%d %H:%M:%S')} ===",
            f"Actives deleted: {len(valid_rows)}",
            "",
            *sql_lines,
            "",
        ]
        log_file = LOG_DIR / f"delete_actives_{now.strftime('%Y-%m-%d_%H-%M-%S')}.log"
        with open(log_file, "a", encoding="utf-8") as f:
            f.write("\n".join(log_lines))
        logger.info("Deleted %d actives, log: %s", len(valid_rows), log_file)

        progress.update(status="done", count=len(valid_rows),
                         message=f"Успешно удалено активов: {len(valid_rows)}")

    async def _validate_create_named_actives(
        self, db_session: AsyncSession, rows: list[dict], progress: dict | None = None,
    ) -> tuple[list[dict], list[dict]]:
        """Validate the rows creating named assets (the asset number comes from the file).

        Unlike creating assets from materials, no iterator_number_last counter is needed
        here: active_number is taken from the 'Актив' column and must not already exist in
        the database. Each row is one asset with its own location record; the lcn is still
        handed out from storage.last_lcn when the SQL runs.
        """
        actives_repo = ActivesRepository(session=db_session)
        storage_repo = StorageRepository(session=db_session)
        consignment_repo = ConsignmentRepository(session=db_session)
        design_number_repo = DesignNumberRepository(session=db_session)

        errors: list[dict] = []
        valid_rows: list[dict] = []
        batch_numbers: set[str] = set()

        required_columns = ("Актив", "ТМЦ", "Положение", "Партия")
        if rows:
            missing = [c for c in required_columns if c not in rows[0]]
            if missing:
                errors.append({
                    "row": 0, "field": ", ".join(missing),
                    "message": f"В файле не найдены колонки: {', '.join(missing)}",
                })
                return errors, valid_rows

        storage_cache: dict[str, int | None] = {}
        consignment_cache: dict[str, int | None] = {}
        design_number_cache: dict[str, int | None] = {}

        if progress is not None:
            progress.update(processed=0, total=len(rows), phase="validating")

        for idx, row in enumerate(rows):
            row_num = idx + 1
            if progress is not None and (idx % 20 == 0 or row_num == len(rows)):
                progress["processed"] = row_num

            active_number = str(row.get("Актив") or "").strip()
            if not active_number:
                errors.append({"row": row_num, "field": "Актив", "message": "Поле 'Актив' пустое"})
                continue

            if active_number in batch_numbers:
                errors.append({"row": row_num, "field": "Актив",
                                "message": f"Дубликат внутри файла: '{active_number}'"})
                continue

            existing = await actives_repo.get_one_or_none(Actives.active_number == active_number)
            if existing is not None:
                errors.append({"row": row_num, "field": "Актив",
                                "message": f"Актив уже существует: '{active_number}'"})
                continue

            design_number = str(row.get("ТМЦ") or "").strip()
            if not design_number:
                errors.append({"row": row_num, "field": "ТМЦ", "message": "Поле 'ТМЦ' пустое"})
                continue
            if design_number not in design_number_cache:
                dn = await design_number_repo.get_one_or_none(DesignNumber.number == design_number)
                design_number_cache[design_number] = dn.id if dn else None
            id_design_number = design_number_cache[design_number]
            if id_design_number is None:
                errors.append({"row": row_num, "field": "ТМЦ", "message": f"ТМЦ не найдена: '{design_number}'"})
                continue

            storage_name = str(row.get("Положение") or "").strip()
            if not storage_name:
                errors.append({"row": row_num, "field": "Положение", "message": "Поле 'Положение' пустое"})
                continue
            if storage_name not in storage_cache:
                storage = await storage_repo.get_one_or_none(Storage.name == storage_name)
                storage_cache[storage_name] = storage.id if storage else None
            id_storage = storage_cache[storage_name]
            if id_storage is None:
                errors.append({"row": row_num, "field": "Положение", "message": f"Склад не найден: '{storage_name}'"})
                continue

            consignment_name = str(row.get("Партия") or "").strip()
            if not consignment_name:
                errors.append({"row": row_num, "field": "Партия", "message": "Поле 'Партия' пустое"})
                continue
            if consignment_name not in consignment_cache:
                c = await consignment_repo.get_one_or_none(Consignment.name == consignment_name)
                consignment_cache[consignment_name] = c.id if c else None
            id_consignment = consignment_cache[consignment_name]
            if id_consignment is None:
                errors.append({"row": row_num, "field": "Партия", "message": f"Партия не найдена: '{consignment_name}'"})
                continue

            batch_numbers.add(active_number)
            valid_rows.append({
                "row_num": row_num,
                "active_number": active_number,
                "id_design_number": id_design_number,
                "id_storage": id_storage,
                "id_consignment": id_consignment,
            })

        return errors, valid_rows

    @staticmethod
    def _build_create_named_actives_sql_body(valid_rows: list[dict]) -> list[str]:
        """Build the body of the named asset creation SQL (without BEGIN/COMMIT).

        As in _build_create_actives_sql_body: ids for location/actives come from nextval(),
        and storage.last_lcn is read and locked FOR UPDATE inside the DO block as it runs.
        The difference is that active_number is inlined as a literal from the file and the
        iterator_number_last counter is not used.
        """
        total_actives = len(valid_rows)
        storage_ids = sorted({vr["id_storage"] for vr in valid_rows})

        sql_lines: list[str] = ["DO $$", "DECLARE"]
        sql_lines.append(
            f"    loc_ids bigint[] := ARRAY(SELECT nextval('public.location_id_seq') "
            f"FROM generate_series(1, {total_actives}));"
        )
        sql_lines.append(
            f"    act_ids bigint[] := ARRAY(SELECT nextval('public.actives_id_seq') "
            f"FROM generate_series(1, {total_actives}));"
        )
        for storage_id in storage_ids:
            sql_lines.append(f"    lcn_{storage_id} bigint;")

        sql_lines.append("BEGIN")
        for storage_id in storage_ids:
            sql_lines.append(
                f"    SELECT last_lcn INTO lcn_{storage_id} FROM public.storage WHERE id = {storage_id} FOR UPDATE;"
            )

        for i, vr in enumerate(valid_rows, start=1):
            loc_ref = f"loc_ids[{i}]"
            act_ref = f"act_ids[{i}]"
            lcn_var = f"lcn_{vr['id_storage']}"
            sql_lines.append(f"    {lcn_var} := {lcn_var} + 1;")
            sql_lines.append(
                f"    INSERT INTO public.location (id, id_type_location, id_storage, id_consignment) "
                f"VALUES ({loc_ref}, 1, {vr['id_storage']}, {vr['id_consignment']});"
            )
            sql_lines.append(
                f"    INSERT INTO public.actives (id, active_number, id_design_number, id_location, lcn) "
                f"VALUES ({act_ref}, '{sql_escape(vr['active_number'])}', {vr['id_design_number']}, {loc_ref}, "
                f"('S{vr['id_storage']}.' || {lcn_var})::ltree);"
            )

        for storage_id in storage_ids:
            sql_lines.append(f"    UPDATE public.storage SET last_lcn = lcn_{storage_id} WHERE id = {storage_id};")

        sql_lines.append("END $$;")

        return sql_lines

    @post("/create-named-actives/generate-sql/start")
    async def create_named_actives_generate_sql_start(self, request: Request) -> Response:
        """Start background generation of the named asset creation SQL file; returns a task_id."""
        session_id = request.session.get(f"{PREFIX}_session_id", "")
        stored = _load_data(session_id) if session_id else None
        if not stored:
            return Response(
                content=json.dumps({"status": "error", "errors": [{"row": 0, "field": "*", "message": "Данные не загружены"}]}),
                status_code=200,
                media_type="application/json",
            )

        rows: list[dict] = stored["rows"]

        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_create_named_actives_generate(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task

        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_create_named_actives_generate(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                errors, valid_rows = await self._validate_create_named_actives(session, rows, progress=progress)
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка: {e}"}])
            return

        if errors:
            progress.update(status="error", errors=errors)
            return
        if not valid_rows:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": "Нет валидных строк для создания активов"}])
            return

        sql_lines = self._build_create_named_actives_sql_body(valid_rows)
        full_sql = "\n".join(["BEGIN;", *sql_lines, "COMMIT;"])
        progress.update(status="done", sql=full_sql, count=len(valid_rows))

    @post("/create-named-actives/execute/start")
    async def create_named_actives_execute_start(self, request: Request) -> Response:
        """Start the background atomic creation of named assets; returns a task_id."""
        session_id = request.session.get(f"{PREFIX}_session_id", "")
        stored = _load_data(session_id) if session_id else None
        if not stored:
            return Response(
                content=json.dumps({"status": "error", "errors": [{"row": 0, "field": "*", "message": "Данные не загружены"}]}),
                status_code=200,
                media_type="application/json",
            )

        rows: list[dict] = stored["rows"]

        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_create_named_actives_execute(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task

        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_create_named_actives_execute(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        valid_rows: list[dict] = []
        sql_lines: list[str] = []
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                errors, valid_rows = await self._validate_create_named_actives(session, rows, progress=progress)

                if errors:
                    progress.update(status="error", errors=errors)
                    return
                if not valid_rows:
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": "Нет валидных строк для создания активов"}])
                    return

                progress.update(processed=0, total=1, phase="executing")

                sql_lines = self._build_create_named_actives_sql_body(valid_rows)
                sql_body = "\n".join(sql_lines)

                try:
                    # IMPORTANT: as in create-actives, session.rollback() below also rolls this
                    # raw call back only because the session opened a real transaction earlier
                    # (the repository queries inside _validate_create_named_actives). Do not
                    # remove the session usage before this point.
                    conn = await session.connection()
                    raw_conn = await conn.get_raw_connection()
                    await raw_conn.driver_connection.execute(sql_body)

                    await session.commit()
                except Exception as e:
                    await session.rollback()
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
                    return

                progress["processed"] = 1
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
            return

        now = datetime.now()
        log_lines = [
            f"=== Create Named Actives: {now.strftime('%Y-%m-%d %H:%M:%S')} ===",
            f"Actives created: {len(valid_rows)}",
            "",
            *sql_lines,
            "",
        ]
        log_file = LOG_DIR / f"create_named_actives_{now.strftime('%Y-%m-%d_%H-%M-%S')}.log"
        with open(log_file, "a", encoding="utf-8") as f:
            f.write("\n".join(log_lines))
        logger.info("Created %d named actives, log: %s", len(valid_rows), log_file)

        progress.update(status="done", count=len(valid_rows),
                         message=f"Успешно создано именных активов: {len(valid_rows)}")

    @post("/create-actives/generate-sql/start")
    async def create_actives_generate_sql_start(self, request: Request) -> Response:
        """Start background generation of the SQL file creating assets from materials; returns a task_id."""
        session_id = request.session.get(f"{PREFIX}_session_id", "")
        stored = _load_data(session_id) if session_id else None
        if not stored:
            return Response(
                content=json.dumps({"status": "error", "errors": [{"row": 0, "field": "*", "message": "Данные не загружены"}]}),
                status_code=200,
                media_type="application/json",
            )

        rows: list[dict] = stored["rows"]

        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_create_actives_generate(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task

        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_create_actives_generate(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                errors, valid_rows = \
                    await self._validate_create_actives_rows(session, rows, progress=progress)
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка: {e}"}])
            return

        if errors:
            progress.update(status="error", errors=errors)
            return
        if not valid_rows:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": "Нет валидных строк для создания активов"}])
            return

        sql_lines = self._build_create_actives_sql_body(valid_rows)
        full_sql = "\n".join(["BEGIN;", *sql_lines, "COMMIT;"])
        progress.update(status="done", sql=full_sql, count=len(valid_rows))

    @post("/create-actives/execute/start")
    async def create_actives_execute_start(self, request: Request) -> Response:
        """Start the background atomic insert of assets from materials; returns a task_id."""
        session_id = request.session.get(f"{PREFIX}_session_id", "")
        stored = _load_data(session_id) if session_id else None
        if not stored:
            return Response(
                content=json.dumps({"status": "error", "errors": [{"row": 0, "field": "*", "message": "Данные не загружены"}]}),
                status_code=200,
                media_type="application/json",
            )

        rows: list[dict] = stored["rows"]

        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_create_actives_execute(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task

        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    @staticmethod
    def _reconstruct_created_active_numbers(valid_rows: list[dict], counter_after: int) -> list[str]:
        """Reconstruct the created assets' numbers from the counter's final value.

        The DO block hands out numbers strictly in valid_rows order (active_num is
        incremented per asset), so the numbers can be reconstructed deterministically from
        the counter value after execution. counter_after must be read in the same
        transaction as the DO block — FOR UPDATE still holds the counter row, so no
        concurrent run could have slipped in between.
        """
        counter_before = counter_after - len(valid_rows)
        numbers: list[str] = []
        for i, vr in enumerate(valid_rows, start=1):
            width = ACTIVE_NUMBER_LENGTH - len(vr["type_active"])
            numbers.append(vr["type_active"] + str(counter_before + i).rjust(width, "0"))
        return numbers

    @staticmethod
    def _build_created_actives_xlsx(rows: list[tuple[str, str | None]]) -> str:
        """Build an xlsx of (active_number, serial_number) and return it base64-encoded."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["active_number", "serial_number"])
        for active_number, serial_number in rows:
            ws.append([active_number, serial_number])
        buf = BytesIO()
        wb.save(buf)
        return base64.b64encode(buf.getvalue()).decode()

    async def _run_create_actives_execute(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        valid_rows: list[dict] = []
        created_numbers: list[str] = []
        xlsx_rows: list[tuple[str, str | None]] = []
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                errors, valid_rows = \
                    await self._validate_create_actives_rows(session, rows, progress=progress)

                if errors:
                    progress.update(status="error", errors=errors)
                    return
                if not valid_rows:
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": "Нет валидных строк для создания активов"}])
                    return

                progress.update(processed=0, total=1, phase="executing")

                sql_lines = self._build_create_actives_sql_body(valid_rows)
                sql_body = "\n".join(sql_lines)

                try:
                    # IMPORTANT: as in train_parser, session.rollback() below also rolls this
                    # raw call back only because the session opened a real transaction earlier
                    # (the repository queries inside _validate_create_actives_rows). Do not
                    # remove the session usage before this point.
                    conn = await session.connection()
                    raw_conn = await conn.get_raw_connection()
                    await raw_conn.driver_connection.execute(sql_body)

                    counter_after = (await session.execute(
                        text("SELECT number FROM public.iterator_number_last WHERE description = :d"),
                        {"d": ACTIVE_NUMBER_COUNTER_DESCRIPTION},
                    )).scalar()
                    created_numbers = self._reconstruct_created_active_numbers(valid_rows, counter_after)

                    await session.commit()
                except Exception as e:
                    await session.rollback()
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
                    return

                progress["processed"] = 1

                # Serial numbers are read from the database rather than the file, so the
                # report reflects the records actually created
                db_rows = (await session.execute(
                    text("SELECT active_number, serial_number FROM public.actives "
                         "WHERE active_number = ANY(:nums)"),
                    {"nums": created_numbers},
                )).all()
                serial_by_number = {an: sn for an, sn in db_rows}
                xlsx_rows = [(n, serial_by_number.get(n)) for n in created_numbers]
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
            return

        now = datetime.now()
        log_lines = [
            f"=== Create Actives From TMC: {now.strftime('%Y-%m-%d %H:%M:%S')} ===",
            f"Actives created: {len(valid_rows)}",
            "",
            *(f"{an}\t{sn or ''}" for an, sn in xlsx_rows),
            "",
        ]
        log_file = LOG_DIR / f"create_actives_{now.strftime('%Y-%m-%d_%H-%M-%S')}.log"
        with open(log_file, "a", encoding="utf-8") as f:
            f.write("\n".join(log_lines))
        logger.info("Created %d actives from TMC, log saved to %s", len(valid_rows), log_file)

        progress.update(status="done", count=len(valid_rows),
                         xlsx=self._build_created_actives_xlsx(xlsx_rows),
                         xlsx_filename="actives.xlsx",
                         message=f"Успешно создано активов: {len(valid_rows)}")

    @post("/create-active-from-model/generate-sql/start")
    async def create_active_from_model_generate_sql_start(self, request: Request) -> Response:
        """Start background generation of the SQL file creating assets from a model lcn; returns a task_id."""
        session_id = request.session.get(f"{PREFIX}_session_id", "")
        stored = _load_data(session_id) if session_id else None
        if not stored:
            return Response(
                content=json.dumps({"status": "error", "errors": [{"row": 0, "field": "*", "message": "Данные не загружены"}]}),
                status_code=200,
                media_type="application/json",
            )

        rows: list[dict] = stored["rows"]

        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_create_active_from_model_generate(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task

        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_create_active_from_model_generate(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                errors, valid_rows, skipped_count = \
                    await self._validate_create_active_from_model_rows(session, rows, progress=progress)
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка: {e}"}])
            return

        if errors:
            progress.update(status="error", errors=errors)
            return
        if not valid_rows:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": "Нет валидных строк для создания активов"}])
            return

        sql_lines = self._build_create_active_from_model_sql_body(valid_rows)
        header = [f"-- Пропущено (lcn уже занят на части поездов): {skipped_count}"] if skipped_count else []
        full_sql = "\n".join([*header, "BEGIN;", *sql_lines, "COMMIT;"])
        progress.update(status="done", sql=full_sql, count=len(valid_rows))

    @post("/create-active-from-model/execute/start")
    async def create_active_from_model_execute_start(self, request: Request) -> Response:
        """Start the background atomic insert of assets from a model lcn; returns a task_id."""
        session_id = request.session.get(f"{PREFIX}_session_id", "")
        stored = _load_data(session_id) if session_id else None
        if not stored:
            return Response(
                content=json.dumps({"status": "error", "errors": [{"row": 0, "field": "*", "message": "Данные не загружены"}]}),
                status_code=200,
                media_type="application/json",
            )

        rows: list[dict] = stored["rows"]

        _cleanup_progress()
        task_id = uuid.uuid4().hex
        _progress[task_id] = {"processed": 0, "total": len(rows), "phase": "validating",
                               "status": "running", "created_at": time.time()}
        task = asyncio.ensure_future(self._run_create_active_from_model_execute(task_id, rows))
        task.add_done_callback(lambda t: _tasks.pop(task_id, None))
        _tasks[task_id] = task

        return Response(
            content=json.dumps({"task_id": task_id}),
            status_code=200,
            media_type="application/json",
        )

    async def _run_create_active_from_model_execute(self, task_id: str, rows: list[dict]) -> None:
        progress = _progress[task_id]
        valid_rows: list[dict] = []
        skipped_count = 0
        sql_lines: list[str] = []
        try:
            session_maker = get_session_maker()
            async with session_maker() as session:
                errors, valid_rows, skipped_count = \
                    await self._validate_create_active_from_model_rows(session, rows, progress=progress)

                if errors:
                    progress.update(status="error", errors=errors)
                    return
                if not valid_rows:
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": "Нет валидных строк для создания активов"}])
                    return

                progress.update(processed=0, total=1, phase="executing")

                sql_lines = self._build_create_active_from_model_sql_body(valid_rows)
                sql_body = "\n".join(sql_lines)

                try:
                    # IMPORTANT: as in create-actives, session.rollback() below also rolls this
                    # raw call back only because the session opened a real transaction earlier
                    # (the repository queries inside _validate_create_active_from_model_rows).
                    # Do not remove the session usage before this point.
                    conn = await session.connection()
                    raw_conn = await conn.get_raw_connection()
                    await raw_conn.driver_connection.execute(sql_body)

                    await session.commit()
                except Exception as e:
                    await session.rollback()
                    progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
                    return

                progress["processed"] = 1
        except Exception as e:
            progress.update(status="error", errors=[{"row": 0, "field": "*", "message": f"Ошибка выполнения: {e}"}])
            return

        now = datetime.now()
        log_lines = [
            f"=== Create Active From Model LCN: {now.strftime('%Y-%m-%d %H:%M:%S')} ===",
            f"Actives created: {len(valid_rows)}",
            f"Skipped (lcn already occupied): {skipped_count}",
            "",
            *sql_lines,
            "",
        ]
        log_file = LOG_DIR / f"create_active_from_model_{now.strftime('%Y-%m-%d_%H-%M-%S')}.log"
        with open(log_file, "a", encoding="utf-8") as f:
            f.write("\n".join(log_lines))
        logger.info("Created %d actives from model lcn (%d skipped), log: %s",
                     len(valid_rows), skipped_count, log_file)

        message = f"Успешно создано активов: {len(valid_rows)}"
        if skipped_count:
            message += f" (пропущено {skipped_count} — позиция уже занята на части поездов)"
        progress.update(status="done", count=len(valid_rows), message=message)

    @get("/progress/{task_id:str}")
    async def task_progress(self, task_id: str) -> Response:
        state = _progress.get(task_id)
        if state is None:
            return Response(
                content=json.dumps({"status": "error", "errors": [{"row": 0, "field": "*", "message": "Задача не найдена или устарела"}]}),
                status_code=200,
                media_type="application/json",
            )
        return Response(
            content=json.dumps({k: v for k, v in state.items() if k != "created_at"}),
            status_code=200,
            media_type="application/json",
        )
