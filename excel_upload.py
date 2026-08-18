import tempfile
import uuid
from pathlib import Path

import openpyxl
from litestar.connection.request import Request
from litestar.datastructures import UploadFile
from litestar.response import Redirect

from parser_storage import cleanup_old_files, load_data, save_data

SUPPORTED_SUFFIXES = (".xlsx", ".xls")


def read_sheet(
    tmp_path: str, sheet_name: str | None = None, skip_blank_rows: bool = False,
) -> tuple[list[str], list[dict]]:
    """Read one sheet of an Excel file into (headers, rows).

    The first row holds the headers; a column with an empty header keeps its
    positional name ('col_3'). skip_blank_rows is per parser rather than global:
    dropping fully empty rows shifts the row numbers reported in validation
    errors, so parsers that never skipped them keep counting every row.
    """
    wb = openpyxl.load_workbook(tmp_path, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows: list[dict[str, str | None]] = []
    headers: list[str] = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c) if c else f"col_{i}" for i, c in enumerate(row)]
            continue
        if skip_blank_rows and all(c is None or str(c).strip() == "" for c in row):
            continue
        rows.append({headers[j]: row[j] for j in range(len(row))})

    wb.close()
    return headers, rows


async def handle_upload(
    request: Request, prefix: str, redirect_path: str,
    allow_sheet_choice: bool = True, skip_blank_rows: bool = False,
) -> Redirect:
    """Handle the Excel upload form of a parser page.

    The file is written to a temporary file (openpyxl needs a real path), and a
    workbook with several sheets redirects to the sheet picker with the temporary
    path kept in the session — the file is only read and deleted once a sheet is
    chosen (see handle_sheet_choice). The parsed rows go to parser_storage under a
    fresh session id; everything the page needs afterwards is '<prefix>_session_id'.
    """
    form = await request.form()
    upload_file: UploadFile | None = form.get("file")

    if not upload_file or not upload_file.filename:
        request.session[f"{prefix}_error"] = "Файл не выбран"
        return Redirect(redirect_path)

    suffix = Path(upload_file.filename).suffix.lower()
    if suffix not in SUPPORTED_SUFFIXES:
        request.session[f"{prefix}_error"] = "Поддерживаются только .xlsx и .xls файлы"
        return Redirect(redirect_path)

    try:
        cleanup_old_files()

        content = await upload_file.read()
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(content)
            tmp_path = tmp.name

        if allow_sheet_choice:
            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()

            if len(sheet_names) > 1:
                request.session[f"{prefix}_pending_file"] = tmp_path
                request.session[f"{prefix}_pending_sheets"] = sheet_names
                request.session[f"{prefix}_pending_filename"] = upload_file.filename
                return Redirect(f"{redirect_path}?select_sheet=1")

        headers, rows = read_sheet(tmp_path, skip_blank_rows=skip_blank_rows)
        Path(tmp_path).unlink(missing_ok=True)

        _store(request, prefix, headers, rows, upload_file.filename)
        return Redirect(redirect_path)
    except Exception as e:
        request.session[f"{prefix}_error"] = f"Ошибка чтения файла: {e}"
        return Redirect(redirect_path)


def handle_sheet_choice(
    request: Request, prefix: str, redirect_path: str, sheet_name: str,
    skip_blank_rows: bool = False,
) -> Redirect:
    """Parse the sheet chosen on the picker page of a multi-sheet upload."""
    tmp_path = request.session.get(f"{prefix}_pending_file", "")
    filename = request.session.get(f"{prefix}_pending_filename", "")
    sheet_names = request.session.get(f"{prefix}_pending_sheets", [])

    if not tmp_path or not Path(tmp_path).exists():
        request.session[f"{prefix}_error"] = "Временный файл истёк. Загрузите файл заново."
        return Redirect(redirect_path)

    if sheet_name not in sheet_names:
        request.session[f"{prefix}_error"] = "Выбранный лист не найден в файле."
        return Redirect(redirect_path)

    try:
        headers, rows = read_sheet(tmp_path, sheet_name, skip_blank_rows=skip_blank_rows)
        Path(tmp_path).unlink(missing_ok=True)

        request.session.pop(f"{prefix}_pending_file", None)
        request.session.pop(f"{prefix}_pending_sheets", None)
        request.session.pop(f"{prefix}_pending_filename", None)

        _store(request, prefix, headers, rows, f"{filename} [{sheet_name}]")
        return Redirect(redirect_path)
    except Exception as e:
        request.session[f"{prefix}_error"] = f"Ошибка чтения листа: {e}"
        return Redirect(redirect_path)


def _store(request: Request, prefix: str, headers: list[str], rows: list[dict], filename: str) -> None:
    session_id = uuid.uuid4().hex
    save_data(session_id, {"headers": headers, "rows": rows, "filename": filename})
    request.session[f"{prefix}_session_id"] = session_id


def stored_data(request: Request, prefix: str) -> dict | None:
    """The parsed file behind this page's session, or None when nothing is loaded."""
    session_id = request.session.get(f"{prefix}_session_id", "")
    return load_data(session_id) if session_id else None


def stored_rows(request: Request, prefix: str) -> list[dict] | None:
    stored = stored_data(request, prefix)
    return stored["rows"] if stored else None
